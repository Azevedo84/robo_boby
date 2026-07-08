import os
from openpyxl import load_workbook, Workbook
from core.banco import conecta
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def buscar_filhos(cursor, codigo_pai, visitados, resultado):
    if codigo_pai in visitados:
        return

    visitados.add(codigo_pai)

    cursor.execute("""
        SELECT
            p.codigo,
            p.descricao,
            p.unidade,
            COALESCE(p.obs, ''),

            f.codigo,
            f.descricao,
            COALESCE(f.obs, ''),
            f.unidade,
            ep.quantidade

        FROM produto p
        JOIN estrutura e ON e.id = p.id_versao
        JOIN estrutura_produto ep ON ep.id_estrutura = e.id
        JOIN produto f ON f.id = ep.id_prod_filho

        WHERE p.codigo = ?
    """, (codigo_pai,))

    rows = cursor.fetchall()

    for row in rows:
        resultado.append(row)
        buscar_filhos(cursor, row[4], visitados, resultado)


def processar_estoque_excel():
    cursor = conecta.cursor()

    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    caminho_excel = os.path.join(desktop, "Estoque Final 31-03-2026.xlsx")

    wb = load_workbook(caminho_excel)
    ws = wb.active

    resultado = []
    visitados = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        codigo = str(row[0]).strip()

        # 🔎 busca produto
        cursor.execute("""
            SELECT id, codigo, descricao, unidade, conjunto
            FROM produto
            WHERE codigo = ?
        """, (codigo,))

        prod = cursor.fetchone()

        if not prod:
            print(f"❌ Não encontrado: {codigo}")
            continue

        id_prod, cod, desc, um, conjunto = prod

        # 🔥 SE FOR CONJUNTO → explode estrutura
        if conjunto == 10:
            print(f"🔁 Conjunto: {codigo}")
            buscar_filhos(cursor, codigo, visitados, resultado)

        else:
            print(f"⏭ Ignorado (não é conjunto): {codigo}")

    # remove duplicados
    resultado = list(dict.fromkeys(resultado))

    gerar_excel(resultado)


def gerar_excel(resultado):
    wb = Workbook()
    ws = wb.active
    ws.title = "Estrutura"

    headers = [
        "Cod Pai", "Desc Pai", "UM Pai", "Ref Pai",
        "Cod Filho", "Desc Filho", "Ref Filho", "UM Filho", "Qtde"
    ]

    ws.append(headers)

    # 🎨 estilo cabeçalho
    fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center", vertical="center")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align

    # 📊 dados
    for row in resultado:
        ws.append([
            int(row[0]),
            row[1],
            row[2],
            row[3],
            int(row[4]),
            row[5],
            row[6],
            row[7],
            float(row[8])
        ])

    # 📏 largura automática
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2

    # 📌 alinhar tudo central
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = align

    # 💾 salvar
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    caminho = os.path.join(desktop, "estrutura_estoque.xlsx")

    wb.save(caminho)

    print(f"Excel gerado em: {caminho}")


# 🚀 EXECUTAR
processar_estoque_excel()