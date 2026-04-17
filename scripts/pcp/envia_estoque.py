import os
from pathlib import Path
import sys

os.chdir(r"C:\Users\Anderson\PycharmProjects\robo_boby")

BASE_DIR = Path(__file__).resolve().parents[2]

if str(BASE_DIR) not in sys.path:
    sys.path.append(str(BASE_DIR))

from core.banco import conecta, conecta_robo
from core.erros import trata_excecao
from core.email_service import dados_email
import smtplib
from email.mime.multipart import MIMEMultipart
from email.header import Header
from email.mime.base import MIMEBase
from email import encoders
from datetime import date, timedelta
import locale
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


class EnviaEstoqueFinal:
    def __init__(self):
        self.destinatario = ['<maquinas@unisold.com.br>', '<estoque@acinplas.com.br>']

    def manipula_dados_acinplas(self, ultimo_mes, data_string, nome_mes, ano):
        try:
            cursor = conecta.cursor()
            cursor.execute("SELECT prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, prod.unidade, "
                           "SUM(CASE WHEN sal.local_estoque = 1 THEN sal.saldo ELSE 0 END) as saldo_local_1, "
                           "SUM(CASE WHEN sal.local_estoque = 2 THEN sal.saldo ELSE 0 END) as saldo_local_2, "
                           "SUM(CASE WHEN sal.local_estoque = 1 "
                           "OR sal.local_estoque = 2 THEN sal.saldo ELSE 0 END) as saldo_total "
                           "FROM saldo_estoque as sal "
                           "INNER JOIN produto prod ON sal.produto_id = prod.id "
                           "WHERE saldo <> 0 and (local_estoque = 1 or local_estoque = 2) "
                           "GROUP BY prod.codigo, prod.descricao, prod.obs, prod.unidade;")
            dados_produto = cursor.fetchall()

            cursor.execute("SELECT prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, prod.unidade, "
                           f"COALESCE(CASE WHEN m.tipo < 200 THEN m.quantidade END, 0) AS Qtde_Entrada, "
                           f"COALESCE(CASE WHEN m.tipo > 200 THEN m.quantidade END, 0) AS Qtde_Saida, "
                           f"m.localestoque "
                           f"FROM movimentacao m "
                           "INNER JOIN produto prod ON m.produto = prod.id "
                           f"WHERE m.data > '{ultimo_mes}' "
                           f"and (m.localestoque = 1 or m.localestoque = 2);")
            select_mov = cursor.fetchall()
            if select_mov:
                for movimentacao in select_mov:
                    cod_mov, des_mov, ref_mov, um_mov, entrada_mov, saida_mov, local_mov = movimentacao

                    if cod_mov not in [item[0] for item in dados_produto]:
                        dados_produto.append((cod_mov, des_mov, ref_mov, um_mov, 0, 0, 0))

            saldos_atualizados = []
            for i in dados_produto:
                cod_saldo, des_saldo, ref_saldo, um_saldo, l1_saldo, l2_saldo, t_saldo = i

                if select_mov:
                    for movimentacao in select_mov:
                        cod_mov, des_mov, ref_mov, um_mov, entrada_mov, saida_mov, local_mov = movimentacao
                        if cod_saldo == cod_mov and local_mov == 1:
                            if entrada_mov:
                                l1_saldo -= entrada_mov
                            if saida_mov:
                                l1_saldo += saida_mov
                        elif cod_saldo == cod_mov and local_mov == 2:
                            if entrada_mov:
                                l2_saldo -= entrada_mov
                            if saida_mov:
                                l2_saldo += saida_mov

                saldo_total = l1_saldo + l2_saldo
                saldos_atualizados.append((cod_saldo, des_saldo, ref_saldo, um_saldo, l1_saldo, l2_saldo, saldo_total))

            saldos_atualizados = [produto for produto in saldos_atualizados if produto[6] != 0]
            saldos_atualizados_ordenados = sorted(saldos_atualizados, key=lambda x: x[1])

            arquivo = f'Estoque Final {data_string}.xlsx'
            caminho_arquivo = fr'C:\Users\Anderson\PycharmProjects\robo_boby\{arquivo}'

            self.gerar_excel(saldos_atualizados_ordenados, data_string, caminho_arquivo)
            self.envia_email(nome_mes, ano, arquivo, caminho_arquivo)
            self.remover_anexo(caminho_arquivo)
            self.inserir_no_banco(ultimo_mes, arquivo)

        except Exception as e:
            trata_excecao(e)
            raise

    def gerar_excel(self, extrai_dados_tabela, date_string, caminho_arquivo):
        try:
            if extrai_dados_tabela:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Estoque Final"

                headers = ["Código", "Descrição", "Referência", "UM", "Almox", "Obsoleto", "Total"]
                sheet.append(headers)

                header_row = sheet[1]
                for cell in header_row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                for dados_ex in extrai_dados_tabela:
                    codigo, descr, ref, um, saldo_local_1, saldo_local_2, saldo_total = dados_ex
                    codigu = int(codigo)

                    if saldo_local_1 == "":
                        saldo1_e = 0.00
                    else:
                        saldo1_e = float(saldo_local_1)

                    if saldo_local_2 == "":
                        saldo2_e = 0.00
                    else:
                        saldo2_e = float(saldo_local_2)

                    if saldo_total == "":
                        total = 0.00
                    else:
                        total = float(saldo_total)

                    sheet.append([codigu, descr, ref, um, saldo1_e, saldo2_e, total])

                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                             top=Side(style='thin'), bottom=Side(style='thin'))
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                for column in sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        if isinstance(cell.value, (int, float)):
                            cell_value_str = "{:.3f}".format(cell.value)
                        else:
                            cell_value_str = str(cell.value)
                        if len(cell_value_str) > max_length:
                            max_length = len(cell_value_str)

                    adjusted_width = (max_length + 2)
                    sheet.column_dimensions[column_letter].width = adjusted_width

                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=7, max_col=9):
                    for cell in row:
                        cell.number_format = '0.000'

                workbook.save(caminho_arquivo)

                print(f'Relatório do Estoque Final do dia {date_string} criado com sucesso!!')

        except Exception as e:
            trata_excecao(e)
            raise

    def envia_email(self, mes_escrito, ano, arquivo, caminho_arquivo):
        try:
            saudacao, msg_final, email_user, password = dados_email()

            subject = f'Est - Estoque Final {mes_escrito}/{ano}!'

            msg = MIMEMultipart()
            msg['From'] = email_user
            msg['Subject'] = subject

            corpo_email = saudacao
            corpo_email += "\n\n"
            corpo_email += f"Segue estoque do final do mês de {mes_escrito}."
            corpo_email += "\n\n"

            corpo_email += "\n\n"
            corpo_email += msg_final

            attachment = open(caminho_arquivo, 'rb')

            part = MIMEBase('application', "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment', filename=Header(arquivo, 'utf-8').encode())
            msg.attach(part)

            text = msg.as_string()
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(email_user, password)

            server.sendmail(email_user, self.destinatario, text)
            server.quit()

            print("email enviado")

        except Exception as e:
            trata_excecao(e)
            raise

    def remover_anexo(self, nome_do_arquivo_excel):
        try:
            if os.path.exists(nome_do_arquivo_excel):
                os.remove(nome_do_arquivo_excel)
                print(f"Arquivo {nome_do_arquivo_excel} removido com sucesso.")
            else:
                print(f"O arquivo {nome_do_arquivo_excel} não existe na pasta do projeto.")

        except Exception as e:
            trata_excecao(e)
            raise

    def inserir_no_banco(self, ultimo_mes, arquivo):
        try:
            cursor = conecta_robo.cursor()
            cursor.execute(f"Insert into envia_relat_estoque (ID, data_relatorio) "
                           f"values (GEN_ID(GEN_envia_relat_estoque_ID,1), '{ultimo_mes}');")
            print(f"Estoque Final de {arquivo} gravado com sucesso!")
            conecta_robo.commit()

        except Exception as e:
            trata_excecao(e)
            raise

    def verifica_banco(self):
        try:
            data_hoje = date.today()

            if data_hoje.day >= 1:
                primeiro_dia_do_mes_atual = date(data_hoje.year, data_hoje.month, 1)
                ultimo_dia_mes = primeiro_dia_do_mes_atual - timedelta(days=1)
                data_string = ultimo_dia_mes.strftime("%d-%m-%Y")

                locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
                nome_mes = ultimo_dia_mes.strftime('%B')
                ano = ultimo_dia_mes.strftime('%y')

                cursor = conecta_robo.cursor()
                cursor.execute(f"SELECT * FROM envia_relat_estoque where data_relatorio = '{ultimo_dia_mes}';")
                select_envio = cursor.fetchall()

                if not select_envio:
                    self.manipula_dados_acinplas(ultimo_dia_mes, data_string, nome_mes, ano)

        except Exception as e:
            trata_excecao(e)
            raise


chama_classe = EnviaEstoqueFinal()
chama_classe.verifica_banco()
