import sys
from banco_dados.conexao import conecta
from banco_dados.controle_erros import grava_erro_banco
import os
import time
import inspect
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.header import Header
from email import encoders
from datetime import datetime, date, timedelta
import traceback
import pandas as pd
from openpyxl import load_workbook, drawing
from openpyxl.styles import Side, Alignment, Border, Font
from threading import Thread


para_sem_saldo = None


class EnviaMovimentacao:
    def __init__(self):
        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

    def trata_excecao(self, nome_funcao, mensagem, arquivo, excecao):
        try:
            tb = traceback.extract_tb(excecao)
            num_linha_erro = tb[-1][1]

            traceback.print_exc()
            print(f'Houve um problema no arquivo: {arquivo} na função: "{nome_funcao}"\n{mensagem} {num_linha_erro}')

            grava_erro_banco(nome_funcao, mensagem, arquivo, num_linha_erro)

        except Exception as e:
            nome_funcao_trat = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            tb = traceback.extract_tb(exc_traceback)
            num_linha_erro = tb[-1][1]
            print(f'Houve um problema no arquivo: {self.nome_arquivo} na função: "{nome_funcao_trat}"\n'
                  f'{e} {num_linha_erro}')
            grava_erro_banco(nome_funcao_trat, e, self.nome_arquivo, num_linha_erro)

    def mensagem_email(self):
        try:
            current_time = (datetime.now())
            horario = current_time.strftime('%H')
            hora_int = int(horario)
            saudacao = "teste"
            if 4 < hora_int < 13:
                saudacao = "Bom Dia!"
            elif 12 < hora_int < 19:
                saudacao = "Boa Tarde!"
            elif hora_int > 18:
                saudacao = "Boa Noite!"
            elif hora_int < 5:
                saudacao = "Boa Noite!"

            msg_final = f"Att,\n" \
                        f"Suzuki Máquinas Ltda\n" \
                        f"Fone (51) 3561.2583/(51) 3170.0965\n\n" \
                        f"Mensagem enviada automaticamente, por favor não responda.\n\n" \
                        f"Caso haja divergências com a movimentação dos itens, favor entrar em contato pelo email " \
                        f"fat_maq@unisold.com.br. " \
                        f"Se houver algum problema com o recebimento de emails ou conflitos com o arquivo excel, " \
                        f"favor entrar em contato pelo email maquinas@unisold.com.br.\n\n"

            email_user = 'ti.ahcmaq@gmail.com'

            to = ['<maquinas@unisold.com.br>']

            password = 'poswxhqkeaacblku'

            return saudacao, msg_final, email_user, to, password

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def datas_relatorio(self):
        try:
            cursor = conecta.cursor()
            cursor.execute(f"SELECT MAX(data_mov) FROM envia_mov where tipo = 1;")
            ultima_data = cursor.fetchall()
            ult_data = ultima_data[0]
            u_data = ult_data[0]

            data_hoje = date.today()

            data_relatorio = u_data + timedelta(days=1)

            dia_relatorio = data_relatorio.strftime("%d")
            mes_relatorio = data_relatorio.strftime("%m")
            ano_relatorio = data_relatorio.strftime("%Y")
            data_str = f"{dia_relatorio}/{mes_relatorio}/{ano_relatorio}"

            if mes_relatorio == "01":
                mes_certo = "Janeiro"
            elif mes_relatorio == "02":
                mes_certo = "Fevereiro"
            elif mes_relatorio == "03":
                mes_certo = "Marco"
            elif mes_relatorio == "04":
                mes_certo = "Abril"
            elif mes_relatorio == "05":
                mes_certo = "Maio"
            elif mes_relatorio == "06":
                mes_certo = "Junho"
            elif mes_relatorio == "07":
                mes_certo = "Julho"
            elif mes_relatorio == "08":
                mes_certo = "Agosto"
            elif mes_relatorio == "09":
                mes_certo = "Setembro"
            elif mes_relatorio == "10":
                mes_certo = "Outubro"
            elif mes_relatorio == "11":
                mes_certo = "Novembro"
            elif mes_relatorio == "12":
                mes_certo = "Dezembro"
            else:
                mes_certo = ""

            return data_hoje, data_relatorio, dia_relatorio, mes_certo, ano_relatorio, data_str

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def envia_email_com_anexo(self):
        try:
            saudacao, msg_final, email_user, to, password = self.mensagem_email()

            data_hoje, data_relatorio, dia_relatorio, mes_certo, ano_relatorio, data_str = self.datas_relatorio()

            subject = f'Mov - Movimentação do dia {data_str}'

            msg = MIMEMultipart()
            msg['From'] = email_user
            msg['Subject'] = subject

            body = f"{saudacao}\n\n" \
                   f"Segue movimentação do dia {data_str}.\n\n\n\n" \
                   f"{msg_final}"

            msg.attach(MIMEText(body, 'plain'))

            nome_arquivo = f'Mov {dia_relatorio} de {mes_certo}.xlsx'
            caminho_arquivo = fr'C:\Users\Anderson\PycharmProjects\robo_boby\Mov {dia_relatorio} de {mes_certo}.xlsx'

            attachment = open(caminho_arquivo, 'rb')

            part = MIMEBase('application', "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment', filename=Header(nome_arquivo, 'utf-8').encode())
            msg.attach(part)

            text = msg.as_string()
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(email_user, password)

            server.sendmail(email_user, to, text)
            attachment.close()

            server.quit()

            print("email enviado com anexo")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def envia_email_sem_anexo(self):
        try:
            saudacao, msg_final, email_user, to, password = self.mensagem_email()

            data_hoje, data_relatorio, dia_relatorio, mes_certo, ano_relatorio, data_str = self.datas_relatorio()

            subject = f'Mov - Sem Movimentação no dia {data_str}'

            msg = MIMEMultipart()
            msg['From'] = email_user
            msg['Subject'] = subject

            body = f"{saudacao}\n\nNo dia {data_str} não teve movimentação no estoque.\n\n\n\n" \
                   f"{msg_final}"

            msg.attach(MIMEText(body, 'plain'))

            text = msg.as_string()
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(email_user, password)

            server.sendmail(email_user, to, text)
            server.quit()

            self.salvar_envio()

            print("email enviado sem anexo")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def envia_email_sem_saldo(self, produtos_sem_saldo):
        try:
            to_certo = ['<maquinas@unisold.com.br>']

            msg_itens = ""
            dados = produtos_sem_saldo
            for didis in dados:
                dat_s, cod_s, des_s, ref_s, um_s, ent_s, sai_s, saldo_s, op_s, cfo_s, loc_s, sol_s, obs_s = didis

                msg_itens = msg_itens + f"- Data: {dat_s}\n" \
                                        f"- Código: {cod_s}\n" \
                                        f"- Descrição: {des_s}\n" \
                                        f"- Referência.: {ref_s}\n" \
                                        f"- UM: {um_s} Saldo: {saldo_s}\n\n"

            saudacao, msg_final, email_user, to, password = self.mensagem_email()

            data_hoje, data_relatorio, dia_relatorio, mes_certo, ano_relatorio, data_str = self.datas_relatorio()

            subject = f'Mov - Problemas na Movimentação do dia {data_str}'

            msg = MIMEMultipart()
            msg['From'] = email_user
            msg['Subject'] = subject

            body = f"{saudacao}\n\nNão foi possivel enviar a movimentação do dia {data_str}, pois alguns materiais\n" \
                   f"ficaram com saldo negativo.\n\n" \
                   f"Segue abaixo materiais:\n\n" \
                   f"{msg_itens}\n\n\n" \
                   f"{msg_final}"

            msg.attach(MIMEText(body, 'plain'))

            text = msg.as_string()
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(email_user, password)

            server.sendmail(email_user, to_certo, text)
            server.quit()

            data_hoje, data_relatorio, dia_relatorio, mes_certo, ano_relatorio, data_str = self.datas_relatorio()

            current_time = (datetime.now())
            horario = current_time.strftime('%H:%M:%S')

            cursor = conecta.cursor()
            cursor.execute(f"Insert into ENVIA_MOV (ID, DATA_MOV, DATA_ENVIA, HORA_ENVIA, TIPO) "
                           f"values (GEN_ID(GEN_ENVIA_MOV_ID,1), '{data_relatorio}', '{data_hoje}', "
                           f"'{horario}', 2);")
            conecta.commit()

            print("salvo no banco sem saldo")

            print("email enviado problema saldo")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def salvar_envio(self):
        try:
            data_hoje, data_relatorio, dia_relatorio, mes_certo, ano_relatorio, data_str = self.datas_relatorio()

            current_time = (datetime.now())
            horario = current_time.strftime('%H:%M:%S')

            cursor = conecta.cursor()
            cursor.execute(f"Insert into ENVIA_MOV (ID, DATA_MOV, DATA_ENVIA, HORA_ENVIA, TIPO) "
                           f"values (GEN_ID(GEN_ENVIA_MOV_ID,1), '{data_relatorio}', '{data_hoje}', "
                           f"'{horario}', 1);")
            conecta.commit()

            print("salvo no banco")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excluir_arquivo(self):
        try:
            data_hoje, data_relatorio, dia_relatorio, mes_certo, ano_relatorio, data_str = self.datas_relatorio()

            nome_arquivos = fr'C:\Users\Anderson\PycharmProjects\robo_boby\Mov {dia_relatorio} de {mes_certo}.xlsx'
            attachment = open(nome_arquivos, 'rb')
            attachment.close()
            os.remove(nome_arquivos)
            self.salvar_envio()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def select_mistura_op(self, cod, num_op):
        try:
            dados_para_tabela = []
            campo_br = ""

            cur = conecta.cursor()
            cur.execute(f"SELECT id, descricao, id_versao FROM produto where codigo = {cod};")
            detalhes_produtos = cur.fetchall()

            id_prod, descricao, id_versao = detalhes_produtos[0]

            cursor = conecta.cursor()
            cursor.execute(f"SELECT estprod.id, prod.codigo, prod.descricao, "
                           f"COALESCE(prod.obs, ' ') as obs, prod.unidade, "
                           f"((SELECT quantidade FROM ordemservico where numero = {num_op}) * "
                           f"(estprod.quantidade)) AS Qtde, "
                           f"prod.localizacao, prod.quantidade "
                           f"FROM estrutura_produto as estprod "
                           f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                           f"where estprod.id_estrutura = {id_versao} ORDER BY prod.descricao;")
            select_estrut = cursor.fetchall()

            for dados_estrut in select_estrut:
                id_mat_e, cod_e, descr_e, ref_e, um_e, qtde_e, local_e, saldo_e = dados_estrut

                cursor = conecta.cursor()
                cursor.execute(f"SELECT max(estprod.id), max(prod.codigo), max(prod.descricao), "
                               f"sum(prodser.QTDE_ESTRUT_PROD)as total "
                               f"FROM estrutura_produto as estprod "
                               f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                               f"INNER JOIN produtoos as prodser ON estprod.id = prodser.id_estrut_prod "
                               f"where prodser.numero = {num_op} and estprod.id = {id_mat_e} "
                               f"group by prodser.id_estrut_prod;")
                select_os_resumo = cursor.fetchall()

                if not select_os_resumo:
                    dados0 = (id_mat_e, cod_e, descr_e, ref_e, um_e, qtde_e, local_e, saldo_e,
                              campo_br, campo_br, campo_br, campo_br, campo_br, campo_br)
                    dados_para_tabela.append(dados0)

                for dados_res in select_os_resumo:
                    id_mat_sum, cod_sum, descr_sum, qtde_sum = dados_res
                    sobras = qtde_e - qtde_sum
                    if sobras > 0:
                        dados1 = (id_mat_e, cod_e, descr_e, ref_e, um_e, sobras, local_e, saldo_e,
                                  campo_br, campo_br, campo_br, campo_br, campo_br, campo_br)
                        dados_para_tabela.append(dados1)

                    cursor = conecta.cursor()
                    cursor.execute(f"select prodser.id_estrut_prod, "
                                   f"COALESCE((extract(day from prodser.data)||'/'||"
                                   f"extract(month from prodser.data)||'/'||"
                                   f"extract(year from prodser.data)), '') AS DATA, prod.codigo, prod.descricao, "
                                   f"COALESCE(prod.obs, '') as obs, prod.unidade, "
                                   f"prodser.quantidade, prodser.QTDE_ESTRUT_PROD "
                                   f"from produtoos as prodser "
                                   f"INNER JOIN produto as prod ON prodser.produto = prod.id "
                                   f"where prodser.numero = {num_op} and prodser.id_estrut_prod = {id_mat_e};")
                    select_os = cursor.fetchall()

                    for dados_os in select_os:
                        id_mat_os, data_os, cod_os, descr_os, ref_os, um_os, qtde_os, qtde_mat_os = dados_os

                        dados2 = (id_mat_e, cod_e, descr_e, ref_e, um_e, qtde_mat_os, local_e, saldo_e,
                                  data_os, cod_os, descr_os, ref_os, um_os, qtde_os)
                        dados_para_tabela.append(dados2)

            tabela_estrutura = []
            tabela_consumo_os = []

            for itens in dados_para_tabela:
                id_mat, cod_est, descr_est, ref_est, um_est, qtde_est, local, saldo, \
                    data_os, cod_os, descr_os, ref_os, um_os, qtde_os = itens

                qtde_est_str = str(qtde_est)
                qtde_est_float = float(qtde_est_str)
                qtde_est_red = "%.3f" % qtde_est_float

                if saldo == "":
                    saldo_red = saldo
                else:
                    saldo_str = str(saldo)
                    saldo_float = float(saldo_str)
                    saldo_red = "%.3f" % saldo_float

                if qtde_os == "":
                    qtde_os_red = qtde_os
                else:
                    qtde_os_str = str(qtde_os)
                    qtde_os_float = float(qtde_os_str)
                    qtde_os_red = "%.3f" % qtde_os_float

                lista_est = (id_mat, cod_est, descr_est, ref_est, um_est, qtde_est_red, local, saldo_red)
                tabela_estrutura.append(lista_est)

                lista_os = (id_mat, data_os, cod_os, descr_os, ref_os, um_os, qtde_os_red)
                tabela_consumo_os.append(lista_os)

            return tabela_consumo_os, tabela_estrutura

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def sql_movimentacao_tipo(self, data_inicio, data_fim, num_tipo, nome_tipo):
        try:
            cursor = conecta.cursor()
            cursor.execute(f"SELECT COALESCE((extract(day from m.data)||'/'||"
                           f"extract(month from m.data)||'/'||extract(year from m.data)), '') AS DATA, "
                           f"produto.codigo, produto.descricao, "
                           f"COALESCE(produto.obs, ''), produto.unidade, "
                           f"COALESCE(CASE WHEN m.tipo < 200 THEN m.quantidade END, 0) AS Qtde_Entrada, "
                           f"COALESCE(CASE WHEN m.tipo > 200 THEN m.quantidade END, 0) AS Qtde_Saida, "
                           f"(select case when sum(quantidade) is null then 0 else sum(quantidade) end "
                           f"from movimentacao where produto=m.produto "
                           f"and tipo<200 and localestoque=m.localestoque)-"
                           f"(select case when sum(quantidade) is null then 0 else sum(quantidade) end "
                           f"from movimentacao where produto=m.produto "
                           f"and tipo>200 and localestoque=m.localestoque)+"
                           f"(case when ((select sum(m2.quantidade) from movimentacao m2 "
                           f"where m2.localestoque=m.localestoque and m2.produto=m.produto and "
                           f"(((m.tipo<200) and ((m2.data>m.data) or ((m2.data=m.data) "
                           f"and (m2.id>m.id)))) or(m.tipo>200 and m2.data>m.data)) "
                           f"and m2.tipo<200)*-1) is null then 0 else "
                           f"((select sum(m2.quantidade) from movimentacao m2 "
                           f"where m2.localestoque=m.localestoque and m2.produto=m.produto and "
                           f"(((m.tipo<200) and ((m2.data>m.data) or((m2.data=m.data) "
                           f"and (m2.id>m.id)))) or(m.tipo>200 and m2.data>m.data)) "
                           f"and m2.tipo<200)*-1) end) + "
                           f"(case when (select sum(m2.quantidade) from movimentacao m2 "
                           f"where m2.localestoque=m.localestoque and m2.produto=m.produto and "
                           f"((m2.data=m.data and (m2.id>m.id  or (m.tipo<200)) )or(m2.data>m.data)) "
                           f"and m2.tipo>200) is null then 0 else (select sum(m2.quantidade) "
                           f"from movimentacao m2 where m2.localestoque=m.localestoque "
                           f"and m2.produto=m.produto and ((m2.data=m.data "
                           f"and (m2.id>m.id or (m.tipo<200)) )or(m2.data>m.data)) and m2.tipo>200) end) "
                           f"as saldo, "
                           f"CASE WHEN m.tipo = 210 THEN ('OP '|| produtoos.numero) "
                           f"WHEN m.tipo = 110 THEN ('OP '|| ordemservico.numero) "
                           f"WHEN m.tipo = 130 THEN ('NF '|| entradaprod.nota) "
                           f"WHEN m.tipo = 140 THEN ('INVENTÁRIO') "
                           f"WHEN m.tipo = 240 THEN ('INVENTÁRIO') "
                           f"WHEN m.tipo = 230 THEN ('NF '|| saidaprod.numero) "
                           f"WHEN m.tipo = 250 THEN ('Devol. OS '|| produtoservico.numero) "
                           f"WHEN m.tipo = 112 THEN ('OS '|| produtoservico.numero) "
                           f"WHEN m.tipo = 220 THEN 'CI' "
                           f"END AS OS_NF_CI, "
                           f"COALESCE(natop.descricao, ''), localestoque.nome, "
                           f"CASE WHEN m.tipo = 210 THEN (funcionarios.funcionario) "
                           f"WHEN m.tipo = 110 THEN (funcionarios.funcionario) "
                           f"WHEN m.tipo = 130 THEN (fornecedores.razao) "
                           f"WHEN m.tipo = 140 THEN (funcionarios.funcionario) "
                           f"WHEN m.tipo = 230 THEN (clientes.razao) "
                           f"WHEN m.tipo = 250 THEN (funcionarios.funcionario) "
                           f"WHEN m.tipo = 112 THEN (funcionarios.funcionario) "
                           f"WHEN m.tipo = 220 THEN (funcionarios.funcionario) "
                           f"WHEN m.tipo = 240 THEN (funcionarios.funcionario) "
                           f"END AS teste, "
                           f"COALESCE(m.obs, '') "
                           f"FROM movimentacao m "
                           f"INNER JOIN produto ON (m.codigo = produto.codigo) "
                           f"INNER JOIN localestoque ON (m.localestoque = localestoque.id) "
                           f"LEFT JOIN funcionarios ON (m.funcionario = funcionarios.id) "
                           f"LEFT JOIN saidaprod ON (m.id = saidaprod.movimentacao) "
                           f"LEFT JOIN entradaprod ON (m.id = entradaprod.movimentacao) "
                           f"LEFT JOIN produtoservico ON (m.id = produtoservico.movimentacao) "
                           f"LEFT JOIN ordemservico ON (m.id = ordemservico.movimentacao) "
                           f"LEFT JOIN produtoos ON (m.id = produtoos.movimentacao) "
                           f"LEFT JOIN fornecedores ON (entradaprod.fornecedor = fornecedores.id) "
                           f"LEFT JOIN clientes ON (saidaprod.cliente = clientes.id) "
                           f"LEFT JOIN natop ON (( COALESCE( saidaprod.natureza, 0 ) + "
                           f"COALESCE( entradaprod.natureza, 0 ) ) = natop.ID) "
                           f"WHERE m.data >= '{data_inicio}' and m.data <= '{data_fim}' "
                           f"and localestoque.id IN (1, 2) and m.tipo = {num_tipo} "
                           f"order by m.data, {nome_tipo}, m.id;")
            results = cursor.fetchall()

            return results

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def tipos_movimentos(self):
        try:
            ent_nf = 130
            ent_nf_130 = "entradaprod.nota"

            ent_op = 110
            ent_op_110 = "ordemservico.numero"

            ent_os = 112
            ent_os_112 = "produtoservico.numero"

            ent_inv = 140
            ent_inv_140 = "produtoservico.numero"

            sai_nf = 230
            sai_nf_230 = "saidaprod.numero"

            sai_devolucao = 250
            sai_devolucao_250 = "produtoservico.numero"

            sai_op = 210
            sai_op_210 = "produtoos.numero"

            sai_ci = 220
            sai_ci_220 = "produtoos.numero"

            sai_inv = 240
            sai_inv_240 = "produtoos.numero"

            return ent_nf, ent_nf_130, ent_op, ent_op_110, ent_os, ent_os_112, sai_nf, \
                sai_nf_230, ent_inv, ent_inv_140, sai_devolucao, sai_devolucao_250, \
                sai_op, sai_op_210, sai_ci, sai_ci_220, sai_inv, sai_inv_240

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def juntando_mov_por_tipo(self, data_muda):
        try:
            ent_nf, ent_nf_130, ent_op, ent_op_110, ent_os, ent_os_112, sai_nf, sai_nf_230, \
                ent_inv, ent_inv_140, sai_devolucao, sai_devolucao_250, sai_op, sai_op_210, sai_ci, \
                sai_ci_220, sai_inv, sai_inv_240 = self.tipos_movimentos()

            mov_tipo = []
            ops_entradas = []

            total = self.sql_movimentacao_tipo(data_muda, data_muda, ent_nf, ent_nf_130)
            for dados in total:
                mov_tipo.append(dados)

            total = self.sql_movimentacao_tipo(data_muda, data_muda, ent_op, ent_op_110)
            for dados in total:
                ops_entradas.append(dados)
                mov_tipo.append(dados)

            total = self.sql_movimentacao_tipo(data_muda, data_muda, ent_os, ent_os_112)
            for dados in total:
                mov_tipo.append(dados)

            total = self.sql_movimentacao_tipo(data_muda, data_muda, ent_inv, ent_inv_140)
            for dados in total:
                mov_tipo.append(dados)

            total = self.sql_movimentacao_tipo(data_muda, data_muda, sai_nf, sai_nf_230)
            for dados in total:
                data, cod, descr, ref, um, ent, saida, saldo, os_nf_ci, cfop, local, solcic, obs = dados
                if not os_nf_ci:
                    pass
                else:
                    mov_tipo.append(dados)

            total = self.sql_movimentacao_tipo(data_muda, data_muda, sai_devolucao, sai_devolucao_250)
            for dados in total:
                mov_tipo.append(dados)

            total = self.sql_movimentacao_tipo(data_muda, data_muda, sai_op, sai_op_210)
            for dados in total:
                mov_tipo.append(dados)

            total = self.sql_movimentacao_tipo(data_muda, data_muda, sai_ci, sai_ci_220)
            for dados in total:
                mov_tipo.append(dados)

            total = self.sql_movimentacao_tipo(data_muda, data_muda, sai_inv, sai_inv_240)
            for dados in total:
                mov_tipo.append(dados)

            return mov_tipo, ops_entradas

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def sql_movimentacao(self, data_inicio, data_fim):
        try:
            cursor = conecta.cursor()
            cursor.execute(f"SELECT COALESCE((extract(day from m.data)||'/'||"
                           f"extract(month from m.data)||'/'||extract(year from m.data)), '') AS DATA, "
                           f"produto.codigo, produto.descricao, "
                           f"COALESCE(produto.obs, ''), produto.unidade, "
                           f"COALESCE(CASE WHEN m.tipo < 200 THEN m.quantidade END, 0) AS Qtde_Entrada, "
                           f"COALESCE(CASE WHEN m.tipo > 200 THEN m.quantidade END, 0) AS Qtde_Saida, "
                           f"(select case when sum(quantidade) is null then 0 else sum(quantidade) end "
                           f"from movimentacao where produto=m.produto "
                           f"and tipo<200 and localestoque=m.localestoque)-"
                           f"(select case when sum(quantidade) is null then 0 else sum(quantidade) end "
                           f"from movimentacao where produto=m.produto "
                           f"and tipo>200 and localestoque=m.localestoque)+"
                           f"(case when ((select sum(m2.quantidade) from movimentacao m2 "
                           f"where m2.localestoque=m.localestoque and m2.produto=m.produto and "
                           f"(((m.tipo<200) and ((m2.data>m.data) or ((m2.data=m.data) "
                           f"and (m2.id>m.id)))) or(m.tipo>200 and m2.data>m.data)) "
                           f"and m2.tipo<200)*-1) is null then 0 else "
                           f"((select sum(m2.quantidade) from movimentacao m2 "
                           f"where m2.localestoque=m.localestoque and m2.produto=m.produto and "
                           f"(((m.tipo<200) and ((m2.data>m.data) or((m2.data=m.data) "
                           f"and (m2.id>m.id)))) or(m.tipo>200 and m2.data>m.data)) "
                           f"and m2.tipo<200)*-1) end) + "
                           f"(case when (select sum(m2.quantidade) from movimentacao m2 "
                           f"where m2.localestoque=m.localestoque and m2.produto=m.produto and "
                           f"((m2.data=m.data and (m2.id>m.id  or (m.tipo<200)) )or(m2.data>m.data)) "
                           f"and m2.tipo>200) is null then 0 else (select sum(m2.quantidade) "
                           f"from movimentacao m2 where m2.localestoque=m.localestoque "
                           f"and m2.produto=m.produto and ((m2.data=m.data "
                           f"and (m2.id>m.id or (m.tipo<200)) )or(m2.data>m.data)) and m2.tipo>200) end) "
                           f"as saldo, "
                           f"CASE WHEN m.tipo = 210 THEN ('OP '|| produtoos.numero) "
                           f"WHEN m.tipo = 110 THEN ('OP '|| ordemservico.numero) "
                           f"WHEN m.tipo = 130 THEN ('NF '|| entradaprod.nota) "
                           f"WHEN m.tipo = 230 THEN ('NF '|| saidaprod.numero) "
                           f"WHEN m.tipo = 250 THEN ('Devol. OS '|| produtoservico.numero) "
                           f"WHEN m.tipo = 112 THEN ('OS '|| produtoservico.numero) "
                           f"WHEN m.tipo = 220 THEN 'CI' "
                           f"END AS OS_NF_CI, "
                           f"COALESCE(natop.descricao, ''), localestoque.nome, "
                           f"COALESCE(funcionarios.funcionario, ''), COALESCE(m.obs, '') "
                           f"FROM movimentacao m "
                           f"INNER JOIN produto ON (m.codigo = produto.codigo) "
                           f"INNER JOIN localestoque ON (m.localestoque = localestoque.id) "
                           f"LEFT JOIN funcionarios ON (m.funcionario = funcionarios.id) "
                           f"LEFT JOIN saidaprod ON (m.id = saidaprod.movimentacao) "
                           f"LEFT JOIN entradaprod ON (m.id = entradaprod.movimentacao) "
                           f"LEFT JOIN produtoservico ON (m.id = produtoservico.movimentacao) "
                           f"LEFT JOIN ordemservico ON (m.id = ordemservico.movimentacao) "
                           f"LEFT JOIN produtoos ON (m.id = produtoos.movimentacao) "
                           f"LEFT JOIN natop ON (( COALESCE( saidaprod.natureza, 0 ) + "
                           f"COALESCE( entradaprod.natureza, 0 ) ) = natop.ID) "
                           f"WHERE m.data >= '{data_inicio}' and m.data <= '{data_fim}' "
                           f"and localestoque.id IN (1, 2) "
                           f"order by m.data, Qtde_Entrada, m.id;")
            select_movimentacao = cursor.fetchall()

            return select_movimentacao

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excel_op(self, data, op_int, cod, descr, ref, um, entrada, obs, consumo_os, estrutura, caminho, aba_sheet):
        try:
            print("entrei op")
            book = load_workbook(caminho)

            ws = book["Sheet2"]

            mp_copy = book.copy_worksheet(ws)
            mp_copy.title = aba_sheet

            img = drawing.image.Image(r'C:\Users\Anderson\PycharmProjects\robo_boby\logo.jpg')
            mp_copy.add_image(img, 'C3')

            def lanca_dados_coluna(celula, informacao, tam_fonte, negrito):
                celula_sup_esq = mp_copy[celula]
                cel = mp_copy[celula]
                cel.alignment = Alignment(horizontal='center',
                                          vertical='center',
                                          text_rotation=0,
                                          wrap_text=False,
                                          shrink_to_fit=False,
                                          indent=0)
                cel.font = Font(size=tam_fonte, bold=negrito)

                celula_sup_esq.value = informacao

            def lanca_dados_mesclado(mesclado, celula, informacao, tam_fonte, negrito):
                mp_copy.merge_cells(mesclado)
                celula_sup_esq = mp_copy[celula]
                cel = mp_copy[celula]
                cel.alignment = Alignment(horizontal='center', vertical='center', text_rotation=0,
                                          wrap_text=False, shrink_to_fit=False, indent=0)
                cel.font = Font(size=tam_fonte, bold=negrito)
                celula_sup_esq.value = informacao

            cod_op_int = int(cod)

            dados_estrut = []
            dados_os_a = []
            dados_os_b = []
            dados_os_c = []
            total_qtde_mov = 0.00

            for tabi in estrutura:
                id_mat_est, cod_est, descr_est, ref_est, um_est, qtde_est, local, saldo = tabi
                dados = (cod_est, descr_est, ref_est, um_est, qtde_est)
                dados_estrut.append(dados)

            df = pd.DataFrame(dados_estrut, columns=['Cód.', 'Descrição', 'Referência', 'UM', 'Qtde'])

            codigo_int = {'Cód.': int}
            df = df.astype(codigo_int)
            qtde_float = {'Qtde': float}
            df = df.astype(qtde_float)

            itens_sem_id = 0

            for tabi2 in consumo_os:
                id_mat_os, data_os, cod_os, descr_os, ref_os, um_os, qtde_os = tabi2
                if qtde_os == "":
                    print(f"id da materia prima não está vinculado na OP {op_int}")
                    itens_sem_id = itens_sem_id + 1
                else:
                    qtde_os_float = float(qtde_os)

                    total_qtde_mov += qtde_os_float

                    dados1 = (data, cod_os)
                    dados2 = descr_os
                    dados3 = (ref_os, um_os, qtde_os)
                    dados_os_a.append(dados1)
                    dados_os_b.append(dados2)
                    dados_os_c.append(dados3)

            if itens_sem_id == 0:

                df1 = pd.DataFrame(dados_os_a, columns=['Data', 'Cód.'])
                df2 = pd.DataFrame(dados_os_b, columns=['Descrição'])
                df3 = pd.DataFrame(dados_os_c, columns=['Referência', 'UM', 'Qtde'])

                codigo_int1 = {'Cód.': int}
                df1 = df1.astype(codigo_int1)

                qtde_float3 = {'Qtde': float}
                df3 = df3.astype(qtde_float3)

                writer = pd.ExcelWriter(caminho, engine='openpyxl')

                # incluo a formatação no writer
                writer.book = book

                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                linhas_frame = df.shape[0]
                colunas_frame = df.shape[1]

                linhas_frame1 = df1.shape[0]
                colunas_frame1 = df1.shape[1]

                linhas_frame2 = df2.shape[0]
                colunas_frame2 = df2.shape[1]

                linhas_frame3 = df3.shape[0]
                colunas_frame3 = df3.shape[1]

                inicia = 14
                rows = range(inicia, inicia + linhas_frame)
                columns = range(2, colunas_frame + 2)

                for row in rows:
                    for col in columns:
                        mp_copy.cell(row, col).alignment = Alignment(horizontal='center', vertical='center',
                                                                     wrap_text=True)
                        mp_copy.cell(row, col).border = Border(left=Side(border_style='thin', color='00000000'),
                                                               right=Side(border_style='thin', color='00000000'),
                                                               top=Side(border_style='thin', color='00000000'),
                                                               bottom=Side(border_style='thin', color='00000000'),
                                                               diagonal=Side(border_style='thick', color='00000000'),
                                                               diagonal_direction=0,
                                                               outline=Side(border_style='thin', color='00000000'),
                                                               vertical=Side(border_style='thin', color='00000000'),
                                                               horizontal=Side(border_style='thin', color='00000000'))

                inicia1 = 14
                rows1 = range(inicia1, inicia1 + linhas_frame1)
                columns1 = range(8, colunas_frame1 + 8)

                for row1 in rows1:
                    for col1 in columns1:
                        mp_copy.cell(row1, col1).alignment = Alignment(horizontal='center', vertical='center',
                                                                       wrap_text=True)
                        mp_copy.cell(row1, col1).border = Border(left=Side(border_style='thin', color='00000000'),
                                                                 right=Side(border_style='thin', color='00000000'),
                                                                 top=Side(border_style='thin', color='00000000'),
                                                                 bottom=Side(border_style='thin', color='00000000'),
                                                                 diagonal=Side(border_style='thick', color='00000000'),
                                                                 diagonal_direction=0,
                                                                 outline=Side(border_style='thin', color='00000000'),
                                                                 vertical=Side(border_style='thin', color='00000000'),
                                                                 horizontal=Side(border_style='thin', color='00000000'))

                inicia2 = 14
                rows2 = range(inicia2, inicia2 + linhas_frame2)
                columns2 = range(10, colunas_frame2 + 11)

                for row2 in rows2:
                    mp_copy.merge_cells(f"J{row2}:K{row2}")
                    for col2 in columns2:
                        mp_copy.cell(row2, col2).alignment = Alignment(horizontal='center', vertical='center',
                                                                       wrap_text=True)
                        mp_copy.cell(row2, col2).border = Border(left=Side(border_style='thin', color='00000000'),
                                                                 right=Side(border_style='thin', color='00000000'),
                                                                 top=Side(border_style='thin', color='00000000'),
                                                                 bottom=Side(border_style='thin', color='00000000'),
                                                                 diagonal=Side(border_style='thick', color='00000000'),
                                                                 diagonal_direction=0,
                                                                 outline=Side(border_style='thin', color='00000000'),
                                                                 vertical=Side(border_style='thin', color='00000000'),
                                                                 horizontal=Side(border_style='thin', color='00000000'))

                inicia3 = 14
                rows3 = range(inicia3, inicia3 + linhas_frame3)
                columns3 = range(12, colunas_frame3 + 12)

                linhas_certas3 = linhas_frame3 + 14

                for row3 in rows3:
                    for col3 in columns3:
                        mp_copy.cell(row3, col3).alignment = Alignment(horizontal='center', vertical='center',
                                                                       wrap_text=True)
                        mp_copy.cell(row3, col3).border = Border(left=Side(border_style='thin', color='00000000'),
                                                                 right=Side(border_style='thin', color='00000000'),
                                                                 top=Side(border_style='thin', color='00000000'),
                                                                 bottom=Side(border_style='thin', color='00000000'),
                                                                 diagonal=Side(border_style='thick', color='00000000'),
                                                                 diagonal_direction=0,
                                                                 outline=Side(border_style='thin', color='00000000'),
                                                                 vertical=Side(border_style='thin', color='00000000'),
                                                                 horizontal=Side(border_style='thin', color='00000000'))

                lanca_dados_coluna("D6", data, 16, True)
                lanca_dados_mesclado('M4:N4', 'M4', op_int, 18, True)
                lanca_dados_coluna("B9", cod_op_int, 12, False)
                lanca_dados_mesclado('C9:D9', 'C9', descr, 12, False)
                lanca_dados_mesclado('E9:H9', 'E9', ref, 12, False)
                lanca_dados_coluna("I9", um, 12, False)
                lanca_dados_coluna("J9", entrada, 12, False)
                if obs:
                    lanca_dados_mesclado('K9:N9', 'K9', obs, 12, False)
                else:
                    lanca_dados_mesclado('K9:N9', 'K9', " ", 12, False)

                lanca_dados_coluna(f'N{linhas_certas3}', total_qtde_mov, 12, True)

                lanca_dados_mesclado(f'L{linhas_certas3}:M{linhas_certas3}', f'l{linhas_certas3}', "Total Mov.", 12,
                                     True)

                # para escrever só os valores em um lugar específico:
                df.to_excel(writer, sheet_name=aba_sheet, startrow=13, startcol=1, header=False, index=False)
                df1.to_excel(writer, sheet_name=aba_sheet, startrow=13, startcol=7, header=False, index=False)
                df2.to_excel(writer, sheet_name=aba_sheet, startrow=13, startcol=9, header=False, index=False)
                df3.to_excel(writer, sheet_name=aba_sheet, startrow=13, startcol=11, header=False, index=False)

                writer.save()
                print("final op")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def remove_modelo_op(self, caminho):
        try:
            wb = load_workbook(caminho)
            if 'Sheet2' in wb.sheetnames:
                wb.remove(wb['Sheet2'])
            wb.save(caminho)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excel_mov(self, dados_tabela, arquivo_modelo, caminho, aba_sheet):
        try:
            df = pd.DataFrame(dados_tabela, columns=['Data', 'Cód.', 'Descrição', 'Referência', 'UM',
                                                     'Entrada', 'Saída', 'Saldo', 'OS/NF/CI', 'CFOP',
                                                     'Local', 'Solicitante', 'OBS'])

            codigo_int = {'Cód.': int}
            df = df.astype(codigo_int)
            entrada_float = {'Entrada': float}
            df = df.astype(entrada_float)
            saida_float = {'Saída': float}
            df = df.astype(saida_float)
            saldo_float = {'Saldo': float}
            df = df.astype(saldo_float)

            book = load_workbook(arquivo_modelo)

            ws = book["Sheet1"]
            ws.title = aba_sheet

            writer = pd.ExcelWriter(caminho, engine='openpyxl')

            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

            linhas_frame = df.shape[0]
            colunas_frame = df.shape[1]

            ws = book.active

            inicia = 6
            rows = range(inicia, inicia + linhas_frame)
            columns = range(1, colunas_frame + 1)

            for row in rows:
                for col in columns:
                    ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center',
                                                            wrap_text=True)
                    ws.cell(row, col).border = Border(left=Side(border_style='thin', color='00000000'),
                                                      right=Side(border_style='thin', color='00000000'),
                                                      top=Side(border_style='thin', color='00000000'),
                                                      bottom=Side(border_style='thin', color='00000000'),
                                                      diagonal=Side(border_style='thick', color='00000000'),
                                                      diagonal_direction=0,
                                                      outline=Side(border_style='thin', color='00000000'),
                                                      vertical=Side(border_style='thin', color='00000000'),
                                                      horizontal=Side(border_style='thin', color='00000000'))

            # para escrever só os valores em um lugar específico:
            df.to_excel(writer, sheet_name=aba_sheet, startrow=5, startcol=0, header=False, index=False)

            writer.save()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def final(self):
        try:
            data_hoje, data_relatorio, dia_relatorio, mes_certo, ano_relatorio, data_str = self.datas_relatorio()

            select_movimentacao = self.sql_movimentacao(data_relatorio, data_relatorio)

            saldo_negativo = 0
            prod_sem_saldo = []

            if select_movimentacao:
                mov_tipo, ops_entradas = self.juntando_mov_por_tipo(data_relatorio)

                if not mov_tipo:
                    self.envia_email_sem_anexo()
                else:
                    for dados in mov_tipo:
                        dat_s, cod_s, des_s, ref_s, um_s, ent_s, sai_s, saldo_s, op_s, cfo_s, loc_s, sol_s, \
                            obs_s = dados
                        if saldo_s < 0:
                            dad = (dat_s, cod_s, des_s, ref_s, um_s, ent_s, sai_s, saldo_s, op_s, cfo_s, loc_s,
                                   sol_s, obs_s)
                            prod_sem_saldo.append(dad)
                            saldo_negativo = saldo_negativo + 1

                    if saldo_negativo > 0:
                        msg_tipo2 = self.muda_msg_sem_saldo()
                        if msg_tipo2 == 'FOI':
                            self.envia_email_sem_saldo(prod_sem_saldo)
                    else:
                        nomis = fr'C:\Users\Anderson\PycharmProjects\robo_boby\Mov {dia_relatorio} de {mes_certo}.xlsx'

                        modelo = r'C:\Users\Anderson\PycharmProjects\robo_boby\agora.xlsx'
                        aba = 'Movimentação'

                        self.excel_mov(mov_tipo, modelo, nomis, aba)

                        if ops_entradas:
                            for dados_op in ops_entradas:
                                data, cod, descr, ref, um, entrada, saida, saldo, op, cfop, local, solicitante, \
                                    obs = dados_op

                                posicao = op.find("OP ")
                                inicio = posicao + 3
                                escolha = op[inicio:]

                                op_int = int(escolha)

                                cur = conecta.cursor()
                                cur.execute(f"SELECT id, numero, datainicial, status, produto, quantidade, obs "
                                            f"FROM ordemservico where numero = {op_int};")
                                extrair_dados = cur.fetchall()
                                id_os, numero_os, data_emissao, status_os, produto_os, qtde_os, obs = extrair_dados[0]

                                tabela_consumo_os, tabela_estrutura = self.select_mistura_op(cod, op_int)

                                self.excel_op(data, op_int, cod, descr, ref, um, entrada, obs,
                                              tabela_consumo_os, tabela_estrutura, nomis, op)

                        self.remove_modelo_op(nomis)
                        print("gerado arquivo")

                        self.envia_email_com_anexo()
                        self.excluir_arquivo()

            else:
                self.envia_email_sem_anexo()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def recebaa(self):
        try:
            self.final()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def btn_function(self):
        try:
            Thread(target=self.recebaa).start()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_envio(self):
        try:
            data_hoje, data_relatorio, dia_relatorio, mes_certo, ano_relatorio, data_str = self.datas_relatorio()

            if data_relatorio < data_hoje:
                cursor = conecta.cursor()
                cursor.execute(f"SELECT * FROM envia_mov where data_mov = '{data_relatorio}' and tipo = 1;")
                select_envio = cursor.fetchall()

                if not select_envio:
                    self.recebaa()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_enviooooo(self):
        try:
            msg_teste = ''
            data_hoje, data_relatorio, dia_relatorio, mes_certo, ano_relatorio, data_str = self.datas_relatorio()

            if data_relatorio < data_hoje:
                cursor = conecta.cursor()
                cursor.execute(f"SELECT * FROM envia_mov where data_mov = '{data_relatorio}' and tipo = 1;")
                select_envio = cursor.fetchall()
                print(data_relatorio)
                print(select_envio)

                if not select_envio:
                    msg_teste = 'FOI'
                else:
                    msg_teste = 'NÃO FOI'

            return msg_teste

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def muda_msg_sem_saldo(self):
        try:
            msg_teste = ''
            data_hoje, data_relatorio, dia_relatorio, mes_certo, ano_relatorio, data_str = self.datas_relatorio()

            if data_relatorio < data_hoje:
                cursor = conecta.cursor()
                cursor.execute(f"SELECT * FROM envia_mov where data_mov = '{data_relatorio}' and tipo = 2;")
                select_envio = cursor.fetchall()

                if not select_envio:
                    msg_teste = 'FOI'
                else:
                    msg_teste = 'NÃO FOI'

            return msg_teste

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)


contador_iteracoes = 0

chama_classe = EnviaMovimentacao()

while contador_iteracoes < 5:
    msg_tipo1 = chama_classe.verifica_enviooooo()

    if msg_tipo1 == 'NÃO FOI':
        break
    elif msg_tipo1 == 'FOI':
        chama_classe.verifica_envio()
        time.sleep(2)
        contador_iteracoes += 1
    else:
        break
