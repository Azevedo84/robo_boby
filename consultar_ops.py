from core.banco import conecta
from core.erros import trata_excecao
from openpyxl import Workbook
from openpyxl.styles import Font
from pathlib import Path

class ConsultarOPS:
    def __init__(self):
        self.manipula_ops()

    def manipula_ops(self):
        try:
            cursor = conecta.cursor()
            cursor.execute(f"SELECT ordser.datainicial, ordser.dataprevisao, ordser.numero, prod.codigo, "
                           f"prod.descricao, COALESCE(prod.obs, '') as obs, prod.unidade, "
                           f"ordser.quantidade, ordser.status, ordser.ID_ESTRUTURA "
                           f"FROM ordemservico AS ordser "
                           f"INNER JOIN produto prod ON ordser.produto = prod.id "
                           f"WHERE ordser.status = 'A' "
                           f"ORDER BY ordser.numero;")
            op_abertas = cursor.fetchall()

            self.conversao_manipula_dados(op_abertas)

        except Exception as e:
            trata_excecao(e)
            raise

    def conversao_manipula_dados(self, op_abertas):
        try:
            if op_abertas:
                op_ab_editado = []
                for dados_op in op_abertas:
                    emissao, previsao, op, cod, descr, ref, um, qtde, status, id_estrut = dados_op

                    if id_estrut:
                        data_em_texto = '{}/{}/{}'.format(emissao.day, emissao.month, emissao.year)

                        if previsao:
                            data_prev = '{}/{}/{}'.format(previsao.day, previsao.month, previsao.year)
                        else:
                            data_prev = ''

                        cursor = conecta.cursor()
                        cursor.execute(f"SELECT id, codigo FROM produto where codigo = {cod};")
                        select_prod = cursor.fetchall()

                        idez, cod = select_prod[0]

                        total_estrut = 0
                        total_consumo = 0

                        cursor = conecta.cursor()
                        cursor.execute(f"SELECT estprod.id, "
                                       f"((SELECT quantidade FROM ordemservico where numero = {op}) * "
                                       f"(estprod.quantidade)) AS Qtde "
                                       f"FROM estrutura_produto as estprod "
                                       f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                                       f"where estprod.id_estrutura = {id_estrut};")
                        itens_estrutura = cursor.fetchall()

                        for dads in itens_estrutura:
                            ides, quantidade = dads
                            total_estrut += 1

                            cursor = conecta.cursor()
                            cursor.execute(f"SELECT max(prodser.ID_ESTRUT_PROD), "
                                           f"sum(prodser.QTDE_ESTRUT_PROD) as total "
                                           f"FROM estrutura_produto as estprod "
                                           f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                                           f"INNER JOIN produtoos as prodser ON estprod.id = prodser.ID_ESTRUT_PROD "
                                           f"where prodser.numero = {op} and estprod.id = {ides} "
                                           f"group by prodser.ID_ESTRUT_PROD;")
                            itens_consumo = cursor.fetchall()
                            for duds in itens_consumo:
                                id_mats, qtde_mats = duds
                                if ides == id_mats and quantidade == qtde_mats:
                                    total_consumo += 1

                        dados = (data_em_texto, data_prev, op, cod, descr, ref, um, qtde, total_estrut, total_consumo)
                        op_ab_editado.append(dados)

                if op_ab_editado:
                    self.gerar_excel(op_ab_editado)

        except Exception as e:
            trata_excecao(e)
            raise

    def gerar_excel(self, op_ab_editado):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "OPs Abertas"

            cabecalho = [
                "Emissão",
                "Previsão",
                "OP",
                "Código",
                "Descrição",
                "Referência",
                "UM",
                "Quantidade",
                "Estrutura",
                "Consumido"
            ]

            # Cabeçalho
            for coluna, titulo in enumerate(cabecalho, start=1):
                celula = ws.cell(row=1, column=coluna)
                celula.value = titulo
                celula.font = Font(bold=True)

            # Dados
            for linha, dados in enumerate(op_ab_editado, start=2):
                for coluna, valor in enumerate(dados, start=1):
                    ws.cell(row=linha, column=coluna).value = valor

            # Ajusta largura das colunas
            for coluna in ws.columns:
                tamanho = max(len(str(c.value)) if c.value is not None else 0 for c in coluna)
                ws.column_dimensions[coluna[0].column_letter].width = tamanho + 3

            # Área de Trabalho
            caminho = Path.home() / "Desktop" / "OPs_Abertas.xlsx"

            wb.save(caminho)

            print(f"Excel salvo em:\n{caminho}")

        except Exception as e:
            trata_excecao(e)
            raise

if __name__ == "__main__":
    ConsultarOPS()