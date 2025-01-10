import sys
from banco_dados.conexao import conecta
from banco_dados.controle_erros import grava_erro_banco
from comandos.conversores import valores_para_float
import inspect
import os
import traceback

from pathlib import Path
from comandos.excel import (edita_fonte, criar_workbook, letra_coluna, edita_preenchimento, edita_alinhamento,
                            edita_bordas)

from openpyxl.utils import get_column_letter


class PcpPrevisao:
    def __init__(self):
        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

        self.calculo_1_dados_previsao()

    def trata_excecao(self, nome_funcao, mensagem, arquivo, excecao):
        try:
            tb = traceback.extract_tb(excecao)
            num_linha_erro = tb[-1][1]

            traceback.print_exc()
            print(f'Houve um problema no arquivo: {arquivo} na função: "{nome_funcao}"\n{mensagem} {num_linha_erro}')
            print(f'Houve um problema no arquivo:\n\n{arquivo}\n\n'
                  f'Comunique o desenvolvedor sobre o problema descrito abaixo:\n\n'
                  f'{nome_funcao}: {mensagem}')

            grava_erro_banco(nome_funcao, mensagem, arquivo, num_linha_erro)

        except Exception as e:
            nome_funcao_trat = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            tb = traceback.extract_tb(exc_traceback)
            num_linha_erro = tb[-1][1]
            print(f'Houve um problema no arquivo: {self.nome_arquivo} na função: "{nome_funcao_trat}"\n'
                  f'{e} {num_linha_erro}')
            grava_erro_banco(nome_funcao_trat, e, self.nome_arquivo, num_linha_erro)

    def calculo_1_dados_previsao(self):
        try:

            tabela_final = []

            dados_tabela = self.manipula_dados_pi()

            if dados_tabela:
                saldos = []

                for i in dados_tabela:
                    num_pi, id_prod, codigo, descr, ref, um, qtde, previsao, nivi, cliente, solicitante, conj = i

                    pcte_p_estrutura = [1, num_pi, codigo, qtde, saldos]

                    estrutura = self.calculo_3_verifica_estrutura(pcte_p_estrutura)

                    qtde_itens_estrut = len(estrutura)

                    sem_op = 0

                    if qtde_itens_estrut:
                        cursor = conecta.cursor()
                        cursor.execute(f"select * from ordemservico "
                                       f"where status = 'A' AND produto = {id_prod};")
                        op_abertas = cursor.fetchall()
                        if op_abertas:
                            sem_op = 1

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT produto, qtde_itens FROM RESUMO_ESTRUTURA where produto = {id_prod};")
                    detalhes_resumo = cursor.fetchall()
                    if detalhes_resumo:
                        total_itens = int(detalhes_resumo[0][1])
                    else:
                        total_itens = 0

                    if total_itens:
                        porcentagem = ((total_itens - qtde_itens_estrut) / total_itens) * 100
                    else:
                        porcentagem = 0

                    porc_int = int(porcentagem)

                    if porc_int == 100:
                        status = "CONCLUÍDO"
                    else:
                        if sem_op:
                            status = "PRODUÇÃO"
                        else:
                            if conj == 10:
                                status = "PROJETO"
                            else:
                                status = "COMPRAS"

                    prev = previsao.strftime('%d/%m/%Y')

                    if status != "CONCLUÍDO":
                        dados = (num_pi, cliente, codigo, descr, ref, um, qtde, prev, solicitante, status, porc_int)
                        tabela_final.append(dados)

            self.excel(tabela_final)


        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_dados_pi(self):
        try:
            dados_p_tabela = []

            cursor = conecta.cursor()
            cursor.execute(f"SELECT prodint.id_pedidointerno, prod.id, prod.codigo, prod.descricao, "
                           f"COALESCE(prod.obs, '') as obs, "
                           f"prod.unidade, prodint.qtde, prodint.data_previsao, cli.razao, ped.solicitante, "
                           f"prod.conjunto "
                           f"FROM PRODUTOPEDIDOINTERNO as prodint "
                           f"INNER JOIN produto as prod ON prodint.id_produto = prod.id "
                           f"INNER JOIN pedidointerno as ped ON prodint.id_pedidointerno = ped.id "
                           f"INNER JOIN clientes as cli ON ped.id_cliente = cli.id "
                           f"where prodint.status = 'A' order by prodint.data_previsao;")
            dados_interno = cursor.fetchall()
            if dados_interno:
                for i in dados_interno:
                    num_pi, id_prod, cod, descr, ref, um, qtde, entrega, cliente, solicitante, conj = i

                    dados = (num_pi, id_prod, cod, descr, ref, um, qtde, entrega, "", cliente, solicitante, conj)

                    dados_p_tabela.append(dados)

            return dados_p_tabela

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def calculo_3_verifica_estrutura(self, dados_total):
        try:
            nivel, num_pi, codigo, qtde, lista_saldos = dados_total

            cursor = conecta.cursor()
            cursor.execute(f"SELECT prod.id, prod.codigo, prod.descricao, COALESCE(prod.obs, ''), "
                           f"prod.unidade, tip.tipomaterial, prod.id_versao "
                           f"FROM produto as prod "
                           f"LEFT JOIN tipomaterial as tip ON prod.tipomaterial = tip.id "
                           f"WHERE prod.codigo = '{codigo}';")
            detalhes_pai = cursor.fetchall()

            if detalhes_pai:
                id_pai, cod_pai, descr_pai, ref_pai, um_pai, tipo, id_estrut = detalhes_pai[0]

                cursor.execute(f"SELECT produto_id, saldo FROM SALDO_ESTOQUE "
                               f"WHERE produto_id = '{id_pai}' and local_estoque = 1;")
                detalhes_saldo = cursor.fetchall()

                saldo = detalhes_saldo[0][1] if detalhes_saldo else 0

                qtde_float = valores_para_float(qtde)
                saldo_float = valores_para_float(saldo)

                nova_qtde = 0

                prod_saldo_encontrado = False
                for cod_sal_e, saldo_e in lista_saldos:
                    if cod_sal_e == cod_pai:
                        prod_saldo_encontrado = True
                        break

                if prod_saldo_encontrado:
                    for i_ee, (cod_ee, saldo_ee) in enumerate(lista_saldos):
                        if cod_ee == cod_pai:
                            novo_saldo = saldo_ee - qtde_float
                            lista_saldos[i_ee] = (cod_pai, novo_saldo)

                            nova_qtde = qtde_float - saldo_ee
                            break
                else:
                    novo_saldo = saldo_float - qtde_float
                    lista_saldos.append((cod_pai, novo_saldo))
                    nova_qtde = qtde_float - saldo_float

                filhos = []

                if nova_qtde > 0:
                    dadoss = (nivel, num_pi, cod_pai, descr_pai, ref_pai, um_pai, nova_qtde)
                    filhos.append(dadoss)

                    nivel_plus = nivel + 1

                    if id_estrut:
                        cursor.execute(
                            f"SELECT prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, prod.unidade, "
                            f"(estprod.quantidade * {nova_qtde}) as qtde "
                            f"FROM estrutura_produto as estprod "
                            f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                            f"WHERE estprod.id_estrutura = {id_estrut};")
                        dados_estrutura = cursor.fetchall()

                        if dados_estrutura:
                            for prod in dados_estrutura:
                                cod_f, descr_f, ref_f, um_f, qtde_f = prod
                                pcte_filho = [nivel_plus, num_pi, cod_f, qtde_f, lista_saldos]
                                filhos_recursivos = self.calculo_3_verifica_estrutura(pcte_filho)
                                if filhos_recursivos:
                                    filhos.extend(filhos_recursivos)

                return filhos

            # Caso não haja detalhes_pai ou outra condição não seja satisfeita
            return []

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)
            return []  # Retorne uma lista vazia em caso de exceção

    def excel(self, tabela_final):
        try:
            if tabela_final:
                desktop = Path.home() / "Desktop"
                desk_str = str(desktop)
                nome_req = '/Status Produção.xlsx'
                caminho = (desk_str + nome_req)

                workbook = criar_workbook()
                sheet = workbook.active
                sheet.title = "Lista Completa"

                headers = ["Nº PI", "Cliente", "Código", "Descrição", "Referência", "UM", "Qtde", "Entrega",
                           "Solicitante", "Status", "% Conclusão"]
                sheet.append(headers)

                # Estilo para os cabeçalhos
                header_row = sheet[1]
                for cell in header_row:
                    edita_fonte(cell, negrito=True)
                    edita_preenchimento(cell)
                    edita_alinhamento(cell)

                # Adiciona os dados à planilha
                for d_ex in tabela_final:
                    num_pi, cliente, codigo, descr, ref, um, qtde, entreg, solicitante, status, porc = d_ex

                    porcen = int(porc)
                    codigu = int(codigo)
                    pi_int = int(num_pi)

                    qtde_e = float(qtde) if qtde else 0.00

                    sheet.append([pi_int, cliente, codigu, descr, ref, um, qtde_e, entreg, solicitante,
                                  status, porcen])

                # Estilo para as células
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        edita_bordas(cell)
                        edita_alinhamento(cell)

                # Ajustar largura das colunas
                max_width_limit = 50  # Define um limite máximo para largura de coluna
                for column in sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        if cell.value:
                            cell_value_str = str(cell.value)
                            max_length = max(max_length, len(cell_value_str))
                    # Aplica um ajuste e limita a largura máxima
                    adjusted_width = min(max_length * 1.2, max_width_limit)
                    sheet.column_dimensions[column_letter].width = adjusted_width

                # Formatação numérica
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=7, max_col=9):
                    for cell in row:
                        cell.number_format = '0.00'

                for linha in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=9, max_col=9):
                    for cell in linha:
                        cell.number_format = '0'

                workbook.save(caminho)

                print("Excel salvo com sucesso!")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)


chama_classe = PcpPrevisao()