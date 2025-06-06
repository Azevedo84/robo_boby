import sys
from banco_dados.conexao import conecta
from banco_dados.controle_erros import grava_erro_banco
from comandos.conversores import valores_para_float
from datetime import timedelta, date, datetime
import inspect
import os
import math
import traceback


class TelaPcpPrevisaoV2:
    def __init__(self):
        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

        self.line_SemanCompra = "1"

        self.dados_pi = []
        self.dados_previsao = []
        self.data_inicio = None
        self.funcionario = "5"
        self.horas_dia = "8"

        self.definir_data_inicio()
        self.manipula_dados_pi()
        self.calculo_1_chamar_funcao()
        self.excel()

    def trata_excecao(self, nome_funcao, mensagem, arquivo, excecao):
        try:
            tb = traceback.extract_tb(excecao)
            num_linha_erro = tb[-1][1]

            traceback.print_exc()
            print(f'Houve um problema no arquivo: {arquivo} na função: "{nome_funcao}"\n{mensagem} {num_linha_erro}')
            print(f'Houve um problema no arquivo:\n\n{arquivo}\n\n'
                                 f'Comunique o desenvolvedor sobre o problema descrito abaixo:\n\n'
                                 f'{nome_funcao}: {mensagem}')

            grava_erro_banco(nome_funcao, mensagem, arquivo, num_linha_erro)

        except Exception as e:
            nome_funcao_trat = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            tb = traceback.extract_tb(exc_traceback)
            num_linha_erro = tb[-1][1]
            print(f'Houve um problema no arquivo: {self.nome_arquivo} na função: "{nome_funcao_trat}"\n'
                  f'{e} {num_linha_erro}')
            grava_erro_banco(nome_funcao_trat, e, self.nome_arquivo, num_linha_erro)

    def retorna_calculo_meses(self, dados_tabela):
        try:
            print("retorna_calculo_meses")

            tab_prev = []
            for datas in dados_tabela:
                cod_prev = datas[1]
                dt_previsao = datas[6]
                data_obj = datetime.strptime(dt_previsao, "%d/%m/%Y").date()
                cc = (cod_prev, data_obj)
                tab_prev.append(cc)

            tab_ordenada = sorted(tab_prev, key=lambda x: x[1])

            data_mais_alta = max(tab_ordenada, key=lambda x: x[1])

            tab_meses = []
            for item in tab_ordenada:
                cod_isso = item[0]
                data = item[1]
                diferenca = (data - data_mais_alta[1]).days

                if diferenca < 0:
                    diferenca1 = diferenca * -1
                else:
                    diferenca1 = 0

                meses = diferenca1 / 30 if diferenca1 != 0 else 0

                meses_arredondados = math.ceil(meses)

                dd = (cod_isso, meses_arredondados)
                tab_meses.append(dd)

            return tab_meses

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)
            return None

    def retorna_oc_abertas(self, cod_prod):
        try:
            qtdes_oc = 0

            if cod_prod:
                cursor = conecta.cursor()
                cursor.execute(f"SELECT COALESCE(prodreq.mestre, ''), req.dataemissao, prodreq.quantidade "
                               f"FROM produtoordemsolicitacao as prodreq "
                               f"INNER JOIN produto as prod ON prodreq.produto = prod.ID "
                               f"INNER JOIN ordemsolicitacao as req ON prodreq.mestre = req.idsolicitacao "
                               f"LEFT JOIN produtoordemrequisicao as preq ON prodreq.id = preq.id_prod_sol "
                               f"WHERE prodreq.status = 'A' "
                               f"and prod.codigo = {cod_prod} "
                               f"AND preq.id_prod_sol IS NULL "
                               f"ORDER BY prodreq.mestre;")
                dados_sol = cursor.fetchall()

                if dados_sol:
                    for i_sol in dados_sol:
                        num_sol, emissao_sol, qtde_sol = i_sol
                        qtdes_oc += float(qtde_sol)

                cursor = conecta.cursor()
                cursor.execute(f"SELECT sol.idsolicitacao, prodreq.quantidade, req.data, prodreq.numero, "
                               f"prodreq.destino, prodreq.id_prod_sol "
                               f"FROM produtoordemrequisicao as prodreq "
                               f"INNER JOIN produto as prod ON prodreq.produto = prod.ID "
                               f"INNER JOIN ordemrequisicao as req ON prodreq.mestre = req.id "
                               f"INNER JOIN produtoordemsolicitacao as prodsol ON prodreq.id_prod_sol = prodsol.id "
                               f"INNER JOIN ordemsolicitacao as sol ON prodsol.mestre = sol.idsolicitacao "
                               f"where prodreq.status = 'A' "
                               f"and prod.codigo = {cod_prod};")
                dados_req = cursor.fetchall()

                if dados_req:
                    for i_req in dados_req:
                        num_sol_req, qtde_req, emissao_req, num_req, destino, id_prod_sol = i_req
                        qtdes_oc += float(qtde_req)

                cursor = conecta.cursor()
                cursor.execute(
                    f"SELECT sol.idsolicitacao, prodreq.numero, oc.data, oc.numero, forn.razao, "
                    f"prodoc.quantidade, prodoc.produzido, prodoc.dataentrega "
                    f"FROM ordemcompra as oc "
                    f"INNER JOIN produtoordemcompra as prodoc ON oc.id = prodoc.mestre "
                    f"INNER JOIN produtoordemrequisicao as prodreq ON prodoc.id_prod_req = prodreq.id "
                    f"INNER JOIN produto as prod ON prodoc.produto = prod.id "
                    f"INNER JOIN fornecedores as forn ON oc.fornecedor = forn.id "
                    f"INNER JOIN produtoordemsolicitacao as prodsol ON prodreq.id_prod_sol = prodsol.id "
                    f"INNER JOIN ordemsolicitacao as sol ON prodsol.mestre = sol.idsolicitacao "
                    f"where oc.entradasaida = 'E' "
                    f"AND oc.STATUS = 'A' "
                    f"AND prodoc.produzido < prodoc.quantidade "
                    f"and prod.codigo = {cod_prod}"
                    f"ORDER BY oc.numero;")
                dados_oc = cursor.fetchall()

                if dados_oc:
                    for i_oc in dados_oc:
                        num_sol_oc, id_req_oc, emissao_oc, num_oc, forncec_oc, qtde_oc, prod_oc, entrega_oc = i_oc
                        qtdes_oc += float(qtde_oc)

            return qtdes_oc

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)
            return None

    def retorna_ops_saldo_ops_abertas(self, cod_pai, cod_filho):
        try:
            qtde_ops = 0

            if cod_pai:
                cursor = conecta.cursor()
                cursor.execute(f"SELECT id, codigo, id_versao FROM produto where codigo = {cod_pai};")
                select_prod = cursor.fetchall()
                id_pai, cod, id_versao = select_prod[0]

                if id_versao:
                    cursor = conecta.cursor()
                    cursor.execute(f"select ordser.datainicial, ordser.dataprevisao, ordser.numero, prod.id, "
                                   f"prod.descricao, "
                                   f"COALESCE(prod.obs, '') as obs, prod.unidade, "
                                   f"ordser.quantidade "
                                   f"from ordemservico as ordser "
                                   f"INNER JOIN produto prod ON ordser.produto = prod.id "
                                   f"where ordser.status = 'A' AND prod.codigo = {cod_pai} "
                                   f"order by ordser.numero;")
                    op_abertas = cursor.fetchall()

                    if op_abertas:
                        for i in op_abertas:
                            num_op = i[2]

                            cursor = conecta.cursor()
                            cursor.execute(f"SELECT estprod.id, prod.codigo, prod.descricao, "
                                           f"COALESCE(prod.obs, '') as obs, prod.unidade, "
                                           f"((SELECT quantidade FROM ordemservico where numero = {num_op}) * "
                                           f"(estprod.quantidade)) AS Qtde, "
                                           f"COALESCE(prod.localizacao, ''), prod.quantidade "
                                           f"FROM estrutura_produto as estprod "
                                           f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                                           f"where estprod.id_estrutura = {id_versao} and prod.codigo = {cod_filho} "
                                           f"ORDER BY prod.descricao;")
                            select_estrut = cursor.fetchall()
                            if select_estrut:
                                for dados_estrut in select_estrut:
                                    id_mat_e, cod_e, descr_e, ref_e, um_e, qtde_e, local_e, saldo_e = dados_estrut

                                    cursor = conecta.cursor()
                                    cursor.execute(f"SELECT max(estprod.id), max(prod.codigo), max(prod.descricao), "
                                                   f"sum(prodser.QTDE_ESTRUT_PROD)as total "
                                                   f"FROM estrutura_produto as estprod "
                                                   f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                                                   f"INNER JOIN produtoos as prodser ON "
                                                   f"estprod.id = prodser.id_estrut_prod "
                                                   f"where prodser.numero = {num_op} and estprod.id = {id_mat_e} "
                                                   f"group by prodser.id_estrut_prod;")
                                    select_os_resumo = cursor.fetchall()

                                    if select_os_resumo:
                                        for dados_res in select_os_resumo:
                                            id_mat_sum, cod_sum, descr_sum, qtde_sum = dados_res

                                            qtde_sum_float = valores_para_float(qtde_sum)

                                            qtde_ops += qtde_sum_float

            return qtde_ops

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)
            return None

    def manipula_ordens_compra(self, dados_total):
        try:
            lista_ocs, cod_pai, qtdei_float = dados_total

            qtde_ocs = self.retorna_oc_abertas(cod_pai)

            prod_ocs_encontrado = False
            for cod_oc_p, saldo_p in lista_ocs:
                if cod_oc_p == cod_pai:
                    prod_ocs_encontrado = True
                    break

            if not prod_ocs_encontrado:
                if qtde_ocs > 0:
                    lanca_saldo = (cod_pai, qtde_ocs)
                    lista_ocs.append(lanca_saldo)

                    qtde_float_com_oc = qtdei_float - qtde_ocs
                else:
                    qtde_float_com_oc = qtdei_float
            else:
                qtde_float_com_oc = qtdei_float

            return lista_ocs, qtde_float_com_oc

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)
            return None

    def manipula_ordens_producao(self, dados_total):
        try:
            lista_ops, cod_origem, cod_pai, qtdei_float = dados_total

            qtde_ops = self.retorna_ops_saldo_ops_abertas(cod_origem, cod_pai)

            prod_ops_encontrado = False
            for cod_op_p, saldo_p in lista_ops:
                if cod_op_p == cod_pai:
                    prod_ops_encontrado = True
                    break

            if not prod_ops_encontrado:
                if qtde_ops > 0:
                    lanca_saldo = (cod_pai, qtde_ops)
                    lista_ops.append(lanca_saldo)

                    qtde_float_op = qtdei_float - qtde_ops
                else:
                    qtde_float_op = qtdei_float
            else:
                if qtde_ops > 0:
                    qtde_float_op = qtdei_float - qtde_ops
                else:
                    qtde_float_op = qtdei_float

            return lista_ops, qtde_float_op

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)
            return None

    def definir_folgas(self, data_inicio, data_fim):
        try:
            contagem_sabados = 0
            contagem_domingos = 0

            data_atual = data_inicio
            while data_atual <= data_fim:
                if data_atual.weekday() == 5:
                    contagem_sabados += 1
                elif data_atual.weekday() == 6:
                    contagem_domingos += 1

                data_atual += timedelta(days=1)

            return contagem_sabados, contagem_domingos

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)
            return None

    def retorna_data_entrega(self, id_pais):
        try:
            tempos_de_entrega = []
            fornecedor = ''
            cursor = conecta.cursor()
            cursor.execute(f"SELECT oc.data, oc.numero, prodoc.produto, prodoc.quantidade, mov.data, forn.razao "
                           f"FROM produtoordemcompra as prodoc "
                           f"INNER JOIN entradaprod as ent ON prodoc.mestre = ent.ordemcompra "
                           f"INNER JOIN movimentacao as mov ON ent.movimentacao = mov.id "
                           f"INNER JOIN fornecedores as forn ON ent.fornecedor = forn.id "
                           f"INNER JOIN ordemcompra as oc ON prodoc.mestre = oc.id "
                           f"WHERE prodoc.produto = '{id_pais}' and oc.entradasaida = 'E';")
            extrair_prod = cursor.fetchall()

            if extrair_prod:
                for registro in extrair_prod:
                    data_emissao = registro[0]
                    data_entrega = registro[4]
                    fornecedor = registro[5]

                    tempo_entrega_dias = (data_entrega - data_emissao).days
                    tempos_de_entrega.append(tempo_entrega_dias)
            if tempos_de_entrega:
                media_entrega = sum(tempos_de_entrega) / len(tempos_de_entrega)
            else:
                media_entrega = 0

            entrega = int(media_entrega)

            return entrega, fornecedor

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)
            return None

    def manipula_dados_tabela_producao(self, cod_prod):
        try:
            dados_ops = []
            cursor = conecta.cursor()
            cursor.execute(f"select ordser.datainicial, ordser.dataprevisao, ordser.numero, prod.codigo, "
                           f"prod.descricao, "
                           f"COALESCE(prod.obs, '') as obs, prod.unidade, "
                           f"ordser.quantidade "
                           f"from ordemservico as ordser "
                           f"INNER JOIN produto prod ON ordser.produto = prod.id "
                           f"where ordser.status = 'A' and prod.codigo = {cod_prod} "
                           f"order by ordser.numero;")
            op_abertas = cursor.fetchall()
            if op_abertas:
                for dados_op in op_abertas:
                    emissao, previsao, op, cod, descr, ref, um, qtde = dados_op

                    dados = (op, qtde)
                    dados_ops.append(dados)

            return dados_ops

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)
            return None

    def manipula_dados_pi(self):
        try:
            print("manipula_dados_pi")

            dados_p_tabela = []
            tabela_final = []

            cursor = conecta.cursor()
            cursor.execute(f"SELECT prodint.id_pedidointerno, prod.codigo, prod.descricao, "
                           f"COALESCE(prod.obs, '') as obs, "
                           f"prod.unidade, prodint.qtde, prodint.data_previsao "
                           f"FROM PRODUTOPEDIDOINTERNO as prodint "
                           f"INNER JOIN produto as prod ON prodint.id_produto = prod.id "
                           f"INNER JOIN pedidointerno as ped ON prodint.id_pedidointerno = ped.id "
                           f"INNER JOIN clientes as cli ON ped.id_cliente = cli.id "
                           f"where prodint.status = 'A';")
            dados_interno = cursor.fetchall()
            if dados_interno:
                for i in dados_interno:
                    num_pi, cod, descr, ref, um, qtde, entrega = i

                    dados = (num_pi, cod, descr, ref, um, qtde, entrega, "", "")

                    dados_p_tabela.append(dados)

                tab_ordenada = sorted(dados_p_tabela, key=lambda x: x[6])

                for ii in tab_ordenada:
                    num_pis, cods, des, refs, um, qti, pr, niv, calc = ii

                    prev = pr.strftime('%d/%m/%Y')

                    coco = (num_pis, cods, des, refs, um, qti, prev, niv, calc)
                    tabela_final.append(coco)

            if tabela_final:
                self.dados_pi = tabela_final

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_data_inicio(self):
        try:
            data_inicio = date.today() + timedelta(1)
            dia_da_semana = data_inicio.weekday()
            if dia_da_semana == 5:
                data_ini = date.today() + timedelta(3)
            elif dia_da_semana == 6:
                data_ini = date.today() + timedelta(2)
            else:
                data_ini = date.today() + timedelta(1)

            self.data_inicio = data_ini

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def calculo_1_chamar_funcao(self):
        try:
            print("calculo_1_chamar_funcao")

            dados = self.dados_pi

            if not dados:
                print(f'A tabela "Pedidos Internos Pendentes" está vazia!')
            else:
                self.calculo_2_dados_previsao()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def calculo_2_dados_previsao(self):
        try:
            print("calculo_2_dados_previsao")

            tudo_tudo = []
            dados_niveis = []

            dados_tabela = self.dados_pi

            if dados_tabela:
                saldos = []
                ops = []
                ocs = []

                tab_meses = self.retorna_calculo_meses(dados_tabela)

                for i in dados_tabela:
                    num_pi, codigo, descr, ref, um, qtde, previsao, nivi, entreg = i
                    agrega_pt = 0
                    for cod_isso, meses_isso in tab_meses:
                        if cod_isso == codigo:
                            if meses_isso > 0:
                                agrega_pt = meses_isso * 2

                    cod_origem = ""
                    descr_origem = ""
                    pacote = ""

                    pcte_p_estrutura = [agrega_pt, 1, codigo, qtde, cod_origem, descr_origem, saldos, ops, ocs,
                                        codigo, num_pi, pacote]

                    estrutura = self.calculo_3_verifica_estrutura(pcte_p_estrutura)

                    if estrutura:
                        ultimo_nivel = max(item[1] for item in estrutura)
                        cece = (codigo, ultimo_nivel, previsao)
                        dados_niveis.append(cece)

                        for ii in estrutura:
                            tudo_tudo.append(ii)

            if tudo_tudo:
                self.calculo_4_final_lanca_tabelas(tudo_tudo)
            else:
                print(f'Este Plano de produção está concluído!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def calculo_3_verifica_estrutura(self, dados_total):
        try:
            pontos, nivel, codigos, qtdei, cod_or, descr_or, lista_saldos, lista_ops, lista_ocs, cod_fat, \
            num_pi, pacote = dados_total

            if codigos == "21542":
                print(pontos, nivel, codigos, qtdei, cod_or, descr_or)

            cursor = conecta.cursor()
            cursor.execute(f"SELECT prod.id, prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, "
                           f"prod.unidade, prod.quantidade, tip.tipomaterial, prod.id_versao "
                           f"FROM produto as prod "
                           f"LEFT JOIN tipomaterial as tip ON prod.tipomaterial = tip.id "
                           f"where prod.codigo = {codigos};")
            detalhes_pai = cursor.fetchall()
            id_pai, cod_pai, descr_pai, ref_pai, um_pai, saldo, tipo, id_estrut = detalhes_pai[0]

            if codigos == "21542":
                print(detalhes_pai[0])

            qtdei_float = valores_para_float(qtdei)
            saldo_float = valores_para_float(saldo)

            pcte_ops = [lista_ops, cod_or, cod_pai, qtdei_float]
            lista_ops_mex, qtde_float_com_op = self.manipula_ordens_producao(pcte_ops)

            pcte_ocs = [lista_ocs, cod_pai, qtde_float_com_op]
            lista_ocs_mex, qtde_flt_c_oc = self.manipula_ordens_compra(pcte_ocs)

            prod_saldo_encontrado = False
            for cod_sal_e, saldo_e in lista_saldos:
                if cod_sal_e == cod_pai:
                    prod_saldo_encontrado = True
                    break

            if prod_saldo_encontrado:
                for i_ee, (cod_ee, saldo_ee) in enumerate(lista_saldos):
                    if cod_ee == cod_pai:
                        novo_saldo = saldo_ee - qtde_flt_c_oc
                        lista_saldos[i_ee] = (cod_pai, novo_saldo)
                        break

            else:
                novo_saldo = saldo_float - qtde_flt_c_oc
                lanca_saldo = (cod_pai, novo_saldo)
                lista_saldos.append(lanca_saldo)

            if codigos == "21542":
                print(novo_saldo, qtde_flt_c_oc)

            if novo_saldo < 0 < qtde_flt_c_oc:
                if codigos == "21542":
                    print("entrou")
                coco = novo_saldo + qtde_flt_c_oc
                if coco > 0:
                    nova_qtde = novo_saldo * -1
                else:
                    nova_qtde = qtde_flt_c_oc

                dadoss = (pontos, nivel, cod_pai, descr_pai, ref_pai, um_pai, nova_qtde, cod_or, descr_or,
                          cod_fat, num_pi, pacote)
                nov_msg = f"{cod_pai}({qtde_flt_c_oc}), "
                pacote += nov_msg

                filhos = [dadoss]

                nivel_plus = nivel + 1
                pts_plus = pontos + 1

                if id_estrut:
                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, prod.unidade, "
                                   f"(estprod.quantidade * {nova_qtde}) as qtde "
                                   f"FROM estrutura_produto as estprod "
                                   f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                                   f"where estprod.id_estrutura = {id_estrut};")
                    dados_estrutura = cursor.fetchall()

                    if dados_estrutura:
                        for prod in dados_estrutura:
                            cod_f, descr_f, ref_f, um_f, qtde_f = prod

                            pcte_filho = [pts_plus, nivel_plus, cod_f, qtde_f, cod_pai, descr_pai, lista_saldos,
                                          lista_ops_mex, lista_ocs_mex, cod_fat, num_pi, pacote]
                            filhos.extend(self.calculo_3_verifica_estrutura(pcte_filho))

            else:
                filhos = []

            return filhos

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)
            return None

    def calculo_4_final_lanca_tabelas(self, estrutura_final):
        try:
            print("calculo_4_final_lanca_tabelas")

            lista_de_listas_ordenada = sorted(estrutura_final, key=lambda x: (-x[0], -x[1]))

            lista_com_tempos = []

            for est in lista_de_listas_ordenada:
                pts, nivel, cod, descr, ref, um, qtde, cod_pai, descr_pai, cod_fat, num_pi, pacote = est

                cursor = conecta.cursor()
                cursor.execute(f"SELECT prod.id, prod.codigo, COALESCE(prod.tempo, 0) as temps, tip.tipomaterial "
                               f"FROM produto as prod "
                               f"LEFT JOIN tipomaterial as tip ON prod.tipomaterial = tip.id "
                               f"where prod.codigo = {cod};")
                detalhes_tempo = cursor.fetchall()
                id_pai, cod_bc, tempo, tipo = detalhes_tempo[0]

                entrega, forn = self.retorna_data_entrega(id_pai)

                dedos = (pts, nivel, cod, tempo, forn, entrega, tipo, qtde, cod_pai, descr_pai, cod_fat, num_pi, pacote)
                lista_com_tempos.append(dedos)

            self.calculo_5_recebe_dados(lista_com_tempos)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def calculo_5_recebe_dados(self, estrutura_ord):
        try:
            print("calculo_5_recebe_dados")

            temp_acum = 0.00

            data_ini = self.data_inicio
            data_inicio = datetime.combine(data_ini, datetime.min.time())

            qtde_func = 10
            horas_dia = 8
            semana_compra = 2

            horas_por_dia = float(qtde_func) * float(horas_dia)

            tabela_pra_tabela = []
            tabela_pra_datas = []

            nivel_anterior = 0

            ultima_data = None

            for i in estrutura_ord:
                pontos_i, nivel_i, cod_i, temp_i, forn_i, entreg_i, tipo_i, qtde_i, cod_or_i, descr_or_i, \
                cod_fat, num_pi, pacote = i

                if not tipo_i:
                    tipo = ""
                else:
                    tipo = tipo_i

                if temp_i > 0:
                    tempo_i2 = float(temp_i)

                    total_horas = int(tempo_i2)
                    tempo_decimal = (tempo_i2 - total_horas)
                    minutos_convert = (float(tempo_decimal) * 0.5) / 0.3
                    hor_min_convertido = float(total_horas) + minutos_convert
                    soma_horas_com_qtde = float(qtde_i) * hor_min_convertido

                    temp_pc = "%.2f" % hor_min_convertido

                    temp_acum += soma_horas_com_qtde

                    dias_de_producao = temp_acum / horas_por_dia

                    data_producao = data_inicio + timedelta(dias_de_producao)

                else:
                    if pontos_i < nivel_anterior or nivel_anterior == 0:
                        soma_horas_com_qtde = (int(semana_compra) * 7) * horas_por_dia

                        temp_acum += soma_horas_com_qtde
                        dias_de_producao = temp_acum / horas_por_dia
                        data_producao = data_inicio + timedelta(dias_de_producao)

                        nivel_anterior = pontos_i
                    else:
                        data_producao = ultima_data

                    temp_pc = ""

                dia_da_semana = data_producao.weekday()
                if dia_da_semana == 5:
                    temp_acum += (horas_por_dia * 2)
                    dias_de_prod = temp_acum / horas_por_dia
                    data_final_prod = data_inicio + timedelta(dias_de_prod)

                elif dia_da_semana == 6:
                    temp_acum += horas_por_dia
                    dias_de_prod = temp_acum / horas_por_dia
                    data_final_prod = data_inicio + timedelta(dias_de_prod)

                else:
                    dias_de_prod = temp_acum / horas_por_dia
                    data_final_prod = data_inicio + timedelta(dias_de_prod)

                dt_prod = data_final_prod.strftime('%d/%m/%Y')
                ultima_data = data_final_prod

                qtde_c = "%.2f" % qtde_i
                acum_c = "%.2f" % temp_acum

                dedos1 = (pontos_i, nivel_i, cod_i, qtde_c, dt_prod, temp_pc, acum_c, tipo, forn_i,
                          cod_or_i, descr_or_i, cod_fat, num_pi, pacote)

                tabela_pra_tabela.append(dedos1)

                dedos3 = (cod_i, data_final_prod, dt_prod, nivel_i)
                tabela_pra_datas.append(dedos3)

            ultimo_data = max(item[1].date() for item in tabela_pra_datas)

            sabados, domingos = self.definir_folgas(data_inicio.date(), ultimo_data)
            dias_folga = sabados + domingos

            print("Folga", dias_folga)
            print("total itens", (str(len(tabela_pra_tabela))))
            print("data final", ultimo_data)

            self.calculo_6_manipula_final_tab(tabela_pra_tabela)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def calculo_6_manipula_final_tab(self, lista_final):
        try:
            print("calculo_6_manipula_final_tab")

            tabela_nova = []
            tabela_p_pi = []

            if lista_final:
                for i in lista_final:
                    pontos, nivel, codigo, qtde, entrega, horas, acumulado, tipo, forn, cod_pai, descr_pai, \
                    cod_fat, num_pi, pacote = i

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, "
                                   f"prod.unidade, prod.conjunto "
                                   f"FROM produto as prod "
                                   f"where prod.codigo = {codigo};")
                    detalhes_tempo = cursor.fetchall()
                    codis, descr, ref, um, conjunto = detalhes_tempo[0]

                    if conjunto == 10:
                        conj = "PRODUTO ACABADO"
                    else:
                        conj = "MATERIA-PRIMA"

                    if codis == "21542":
                        print(pontos, codis, descr, ref, um, qtde, entrega, conj, cod_pai, pacote, num_pi)

                    dados = (pontos, codis, descr, ref, um, qtde, entrega, conj, cod_pai, pacote, num_pi)
                    tabela_nova.append(dados)

                    dados1 = (nivel, codis, entrega, cod_fat, num_pi)
                    tabela_p_pi.append(dados1)

            if tabela_nova:
                self.calculo_7_manipula_previsao_pi(tabela_p_pi)

                self.dados_previsao = tabela_nova

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def calculo_7_manipula_previsao_pi(self, dados_previsao):
        try:
            print("calculo_7_manipula_previsao_pi")

            nova_lista = []

            dados_pi = self.dados_pi

            contagem_cod_fat_num_pi_fat = {}

            for nivel_pr, cod_pr, entrega, cod_fat, num_pi_fat in dados_previsao:
                chave = (cod_fat, num_pi_fat)
                if chave in contagem_cod_fat_num_pi_fat:
                    contagem_cod_fat_num_pi_fat[chave] += 1
                else:
                    contagem_cod_fat_num_pi_fat[chave] = 1

            for i in dados_pi:
                num_pi, cod_pi, descr_pi, ref_pi, um_pi, qtde_pi, limite_pi, previsao_pi, niveis_pi = i

                for chave, contagem in contagem_cod_fat_num_pi_fat.items():
                    cod_fat, num_pi_fat = chave
                    entrega = None

                    if num_pi == num_pi_fat and cod_pi == cod_fat:
                        for dados_previsao_item in dados_previsao:
                            if cod_pi == dados_previsao_item[1] and num_pi == dados_previsao_item[4]:
                                entrega = dados_previsao_item[2]
                                break

                        dados = (num_pi, cod_pi, descr_pi, ref_pi, um_pi, qtde_pi, limite_pi, entrega, contagem)
                        nova_lista.append(dados)

            if nova_lista:
                self.calculo_8_manipula_previsao_pi2(dados_pi, nova_lista)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def calculo_8_manipula_previsao_pi2(self, dados_pi, nova_lista):
        try:
            print("calculo_8_manipula_previsao_pi2")

            lista_nova_nova = []

            dados_pi_set = {(x[0], x[1]) for x in dados_pi}
            nova_lista_set = {(x[0], x[1]) for x in nova_lista}

            elementos_faltantes = dados_pi_set - nova_lista_set

            for elemento in elementos_faltantes:
                num_pi_enc, cod_enc = elemento

                for i in dados_pi:
                    num_pi, cod, descr_pi, ref_pi, um_pi, qtde_pi, limite_pi, entrega, contagem = i

                    if num_pi == num_pi_enc and cod == cod_enc:
                        dados = (num_pi, cod, descr_pi, ref_pi, um_pi, qtde_pi, limite_pi, "CONCLUÍDO", "0")
                        lista_nova_nova.append(dados)

            for ii in nova_lista:
                lista_nova_nova.append(ii)

            if lista_nova_nova:
                def converter_data(data_str):
                    if data_str.lower() == 'concluído':
                        return datetime.max
                    else:
                        return datetime.strptime(data_str, '%d/%m/%Y')

                if lista_nova_nova:
                    tab_ordenada = sorted(lista_nova_nova, key=lambda x: converter_data(x[7]))

                    self.dados_pi = tab_ordenada

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excel(self):
        try:
            dados_extraidos = self.dados_previsao
            dados_por_codigo = {}

            if dados_extraidos:
                for i in dados_extraidos:
                    codigo = i[1]
                    nivel = i[0]
                    quantidade = float(i[5])

                    if codigo not in dados_por_codigo:
                        dados_por_codigo[codigo] = list(i)
                        dados_por_codigo[codigo][5] = quantidade
                    else:
                        dados_por_codigo[codigo][5] += quantidade
                        if nivel > dados_por_codigo[codigo][0]:
                            nova_entrada = list(i)
                            nova_entrada[5] = dados_por_codigo[codigo][5]
                            dados_por_codigo[codigo] = nova_entrada

            resultado = []
            for item in dados_por_codigo.values():
                item[5] = f'{item[5]:.2f}'
                resultado.append(tuple(item))

            produtos_por_servico = {}

            for linha in resultado:
                pontos, codis, descr, ref, um, qtde, entrega, conj, cod_pai, pacote, num_pi = linha

                cursor = conecta.cursor()
                cursor.execute(f"""
                    SELECT prod.id, prod.codigo, prod.descricao, COALESCE(prod.obs, ''),
                           prod.unidade, prod.conjunto, tip.tipomaterial, COALESCE(serv.descricao, '')
                    FROM produto as prod
                    LEFT JOIN tipomaterial as tip ON prod.tipomaterial = tip.id
                    LEFT JOIN SERVICO_INTERNO as serv ON prod.ID_SERVICO_INTERNO = serv.id
                    WHERE prod.codigo = '{codis}';
                """)
                detalhes_produto = cursor.fetchall()
                if not detalhes_produto:
                    continue

                id_prod, codiss, descr, ref, um, conjunto, tipo, servico = detalhes_produto[0]

                cursor = conecta.cursor()
                cursor.execute(f"""
                    SELECT op.numero, op.codigo, op.id_estrutura, prod.descricao,
                           COALESCE(prod.obs, ''), prod.unidade,
                           COALESCE(prod.tipomaterial, ''), op.quantidade,
                           COALESCE(prod.id_servico_interno, '')
                    FROM ordemservico as op
                    INNER JOIN produto as prod ON op.produto = prod.id
                    WHERE op.status = 'A' AND prod.codigo = '{codis}';
                """)
                ops_abertas = cursor.fetchall()
                id_estrut = ops_abertas[0][2] if ops_abertas else ""
                num_op = ops_abertas[0][0] if ops_abertas else ""

                produto_info = {
                    "pontos": pontos,
                    "codigo": codis,
                    "descricao": descr,
                    "referencia": ref,
                    "unidade": um,
                    "tipo_material": tipo,
                    "quantidade": qtde,
                    "conjunto": conj,
                    "cod_pai": cod_pai,
                    "pacote": pacote,
                    "pi": num_pi,
                    "num_op": num_op,
                    "id_estrut": id_estrut,
                }

                if codis == "21542":
                    print(produto_info)

                if servico not in produtos_por_servico:
                    produtos_por_servico[servico] = []
                produtos_por_servico[servico].append(produto_info)

            codigos_existentes = set()
            for produtos in produtos_por_servico.values():
                for p in produtos:
                    codigos_existentes.add(p["codigo"])

            cursor = conecta.cursor()
            cursor.execute("""
                SELECT op.numero, prod.codigo, prod.descricao, COALESCE(prod.obs, ''), prod.unidade,
                       COALESCE(prod.tipomaterial, ''), op.quantidade
                FROM ordemservico as op
                INNER JOIN produto as prod ON op.produto = prod.id
                WHERE op.status = 'A';
            """)
            ops_abertas_gerais = cursor.fetchall()

            ops_sem_definicao = []
            for numero, codigo, descricao, referencia, unidade, tipo, quantidade in ops_abertas_gerais:
                if codigo not in codigos_existentes:
                    ops_sem_definicao.append({
                        "Num OP": numero,
                        "Código": codigo,
                        "Descrição": descricao,
                        "Referência": referencia,
                        "Unidade": unidade,
                        "Tipo": tipo,
                        "Quantidade": f'{quantidade:.2f}'
                    })

            if ops_sem_definicao:
                produtos_por_servico["OPs Sem Definição"] = ops_sem_definicao

            nova_lista_produtos, produtos_sem_op = self.separar_dados_select(produtos_por_servico)

            self.excel2(nova_lista_produtos)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def select_mistura(self, num_op, id_estrut):
        try:
            dados_tabela_estrut = []
            dados_tabela_consumo = []

            cursor = conecta.cursor()
            cursor.execute(f"SELECT estprod.id, prod.codigo, prod.descricao, "
                           f"COALESCE(prod.obs, '') as obs, prod.unidade, "
                           f"((SELECT quantidade FROM ordemservico where numero = {num_op}) * "
                           f"(estprod.quantidade)) AS Qtde, "
                           f"COALESCE(prod.localizacao, ''), prod.quantidade "
                           f"FROM estrutura_produto as estprod "
                           f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                           f"where estprod.id_estrutura = {id_estrut} ORDER BY prod.descricao;")
            select_estrut = cursor.fetchall()

            for dados_estrut in select_estrut:
                id_mat_e, cod_e, descr_e, ref_e, um_e, qtde_e, local_e, saldo_e = dados_estrut

                dados = (id_mat_e, cod_e, descr_e, ref_e, um_e, qtde_e, local_e, saldo_e)
                dados_tabela_estrut.append(dados)

                cursor = conecta.cursor()
                cursor.execute(f"SELECT max(estprod.id), max(prod.codigo), max(prod.descricao), "
                               f"sum(prodser.QTDE_ESTRUT_PROD) as total "
                               f"FROM estrutura_produto as estprod "
                               f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                               f"INNER JOIN produtoos as prodser ON estprod.id = prodser.id_estrut_prod "
                               f"where prodser.numero = {num_op} and estprod.id = {id_mat_e} "
                               f"group by prodser.id_estrut_prod;")
                select_os_resumo = cursor.fetchall()

                if select_os_resumo:
                    cursor = conecta.cursor()
                    cursor.execute(f"select prodser.id_estrut_prod, "
                                   f"COALESCE((extract(day from prodser.data)||'/'||"
                                   f"extract(month from prodser.data)||'/'||"
                                   f"extract(year from prodser.data)), '') AS DATA, prod.codigo, prod.descricao, "
                                   f"COALESCE(prod.obs, '') as obs, prod.unidade, "
                                   f"prodser.quantidade, prodser.qtde_estrut_prod "
                                   f"from produtoos as prodser "
                                   f"INNER JOIN produto as prod ON prodser.produto = prod.id "
                                   f"where prodser.numero = {num_op} and prodser.id_estrut_prod = {id_mat_e};")
                    select_os = cursor.fetchall()

                    for dados_os in select_os:
                        id_mat_os, data_os, cod_os, descr_os, ref_os, um_os, qtde_os, qtde_mat_os = dados_os

                        dados2 = (data_os, cod_os, descr_os, ref_os, um_os, qtde_os)
                        dados_tabela_consumo.append(dados2)

            return dados_tabela_estrut, dados_tabela_consumo

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)
            return None

    def separar_dados_select(self, produtos_por_servico):
        try:
            produtos_filtrados = {}
            produtos_sem_op = []

            for servico, lista_produtos in produtos_por_servico.items():
                if servico == "OPs Sem Definição":
                    produtos_filtrados[servico] = lista_produtos
                    continue

                for produto in lista_produtos:
                    num_op = produto.get("num_op")
                    id_estrut = produto.get("id_estrut")

                    if num_op:
                        tabela_estrutura, tabela_consumo_os = self.select_mistura(num_op, id_estrut)

                        qtde_itens_estrut = len(tabela_estrutura)
                        qtde_itens_op = len(tabela_consumo_os)

                        if qtde_itens_estrut and qtde_itens_estrut == qtde_itens_op:
                            if servico not in produtos_filtrados:
                                produtos_filtrados[servico] = []
                            produtos_filtrados[servico].append(produto)

                    else:
                        produtos_sem_op.append(produto)

            return produtos_filtrados, produtos_sem_op

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)
            return None

    def excel2(self, produtos_por_servico):
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Border, Side
            from openpyxl.utils import get_column_letter
            from pathlib import Path

            borda_padrao = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )

            planilhas_por_servico = {}

            for servico, produtos in produtos_por_servico.items():
                for p in produtos:
                    if servico == "OPs Sem Definição":
                        codigo = p.get("Código", "")
                    else:
                        try:
                            codigo = int(p["codigo"])
                        except (ValueError, TypeError, KeyError):
                            codigo = p.get("codigo", "")

                    if servico == "OPs Sem Definição":
                        linha = {
                            "Num OP": p.get("Num OP", ""),
                            "Código": codigo,
                            "Descrição": p.get("Descrição", ""),
                            "Referência": p.get("Referência", ""),
                            "Unidade": p.get("Unidade", ""),
                            "Tipo": p.get("Tipo", ""),
                            "Quantidade": p.get("Quantidade", "")
                        }
                        nome_aba = "OPs Sem Definição"
                    elif not servico.strip():
                        linha = {
                            "Nível": p.get("pontos", ""),
                            "Código": codigo,
                            "Descrição": p.get("descricao", ""),
                            "Referência": p.get("referencia", ""),
                            "Unidade": p.get("unidade", ""),
                            "Tipo": p.get("tipo_material", "")
                        }
                        nome_aba = "Itens Comprados"
                    else:
                        linha = {
                            "Num OP": p.get("num_op", ""),
                            "Código": codigo,
                            "Descrição": p.get("descricao", ""),
                            "Referência": p.get("referencia", ""),
                            "Unidade": p.get("unidade", ""),
                            "Nível": p.get("pontos", "")
                        }
                        nome_aba = servico.strip()[:31] or "Sem_Serviço"

                    if nome_aba not in planilhas_por_servico:
                        planilhas_por_servico[nome_aba] = []

                    planilhas_por_servico[nome_aba].append(linha)

            desktop = Path.home() / "Desktop"
            caminho_arquivo = desktop / "Produto por Serviço.xlsx"

            with pd.ExcelWriter(caminho_arquivo, engine="openpyxl") as writer:
                for aba, dados in planilhas_por_servico.items():
                    df = pd.DataFrame(dados)
                    if "Nível" in df.columns:
                        df.sort_values(by="Nível", ascending=False, inplace=True)
                    df.to_excel(writer, sheet_name=aba, index=False)

            wb = load_workbook(caminho_arquivo)

            for aba in wb.worksheets:
                for row in aba.iter_rows(min_row=1, max_row=aba.max_row, min_col=1, max_col=aba.max_column):
                    for cell in row:
                        cell.border = borda_padrao

                for col in aba.columns:
                    max_length = 0
                    column = get_column_letter(col[0].column)
                    for cell in col:
                        valor = str(cell.value) if cell.value is not None else ""
                        if len(valor) > max_length:
                            max_length = len(valor)
                    aba.column_dimensions[column].width = max_length + 2

            wb.save(caminho_arquivo)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)


tela = TelaPcpPrevisaoV2()
