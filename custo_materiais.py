import sys
from banco_dados.conexao import conecta
from banco_dados.controle_erros import grava_erro_banco
import os
import traceback
import inspect
from comandos.conversores import valores_para_float


class ExecutaCustoMateriais:
    def __init__(self):
        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

        self.inicio_de_tudo()

    def trata_excecao(self, nome_funcao, mensagem, arquivo, excecao):
        try:
            tb = traceback.extract_tb(excecao)
            num_linha_erro = tb[-1][1]

            traceback.print_exc()
            print(f'Houve um problema no arquivo: {arquivo} na função: "{nome_funcao}"\n{mensagem} {num_linha_erro}')

            grava_erro_banco(nome_funcao, mensagem, arquivo, num_linha_erro)

        except Exception as e:
            nome_funcao_trat = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            tb = traceback.extract_tb(exc_traceback)
            num_linha_erro = tb[-1][1]
            print(f'Houve um problema no arquivo: {self.nome_arquivo} na função: "{nome_funcao_trat}"\n'
                  f'{e} {num_linha_erro}')
            grava_erro_banco(nome_funcao_trat, e, self.nome_arquivo, num_linha_erro)

    def inicio_de_tudo(self):
        try:
            nova_lista = []

            codigo = 19907
            qtde = 1

            estrutura = self.calculo_verifica_estrutura(codigo, qtde)

            if estrutura:
                print(len(estrutura))
                for ii in estrutura:
                    codigo, descr, ref, um, qtde, tipo, custo = ii

                    qtde_float = valores_para_float(qtde)
                    custo_float = valores_para_float(custo)

                    total_prod = round((qtde_float * custo_float), 2)

                    dados = (codigo, descr, ref, um, qtde, custo, total_prod, tipo)
                    nova_lista.append(dados)

            if nova_lista:
                self.gerar_excel(nova_lista)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def calculo_verifica_estrutura(self, codigo, qtde):
        try:
            cursor = conecta.cursor()
            cursor.execute(f"SELECT prod.id_versao, prod.descricao, COALESCE(prod.obs, ''), prod.unidade, "
                           f"prod.localizacao, "
                           f"prod.conjunto, prod.CUSTOUNITARIO, prod.TIPOMATERIAL, prod.TERCEIRIZADO, tip.tipomaterial "
                           f"FROM produto as prod "
                           f"LEFT JOIN tipomaterial tip ON prod.tipomaterial = tip.id "
                           f"where prod.codigo = '{codigo}';")
            dados_prod = cursor.fetchall()

            id_estrut, descr, ref, um, local, conj, custo, tipo, custo_terc, nome_tipo = dados_prod[0]

            filhos = []

            if conj != 10:
                dados = (codigo, descr, ref, um, qtde, nome_tipo, custo)
                filhos = [dados]

            else:
                if tipo == 119:
                    dados = (codigo, descr, ref, um, qtde, nome_tipo, custo_terc)
                    filhos = [dados]

            if id_estrut:
                cursor = conecta.cursor()
                cursor.execute(f"SELECT prod.codigo, prod.descricao, prod.obs, prod.unidade,"
                               f"(estprod.quantidade * {qtde}) as qtde, prod.localizacao, prod.quantidade "
                               f"from estrutura_produto as estprod "
                               f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                               f"INNER JOIN conjuntos conj ON prod.conjunto = conj.id "
                               f"where estprod.id_estrutura = {id_estrut} "
                               f"order by conj.conjunto DESC, prod.descricao ASC;")
                estrutura_filho = cursor.fetchall()

                for dados_f in estrutura_filho:
                    cod_f = dados_f[0]
                    qtde_f = dados_f[4]

                    filhos.extend(self.calculo_verifica_estrutura(cod_f, qtde_f))

            return filhos

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)
            return None

    def gerar_excel(self, nova_lista):
        try:
            import os
            import sys
            import pandas as pd
            from datetime import datetime
            from openpyxl import load_workbook
            from openpyxl.styles import numbers

            colunas = [
                "Código",
                "Descrição",
                "Referência",
                "UM",
                "Quantidade",
                "Custo Unitário",
                "Total",
                "Tipo Material"
            ]

            # Cria DataFrame
            df = pd.DataFrame(nova_lista, columns=colunas)

            # Converte tipos corretamente
            df["Código"] = pd.to_numeric(df["Código"], errors="coerce")
            df["Quantidade"] = pd.to_numeric(df["Quantidade"], errors="coerce")
            df["Custo Unitário"] = pd.to_numeric(df["Custo Unitário"], errors="coerce")
            df["Total"] = pd.to_numeric(df["Total"], errors="coerce")

            # Caminho da Área de Trabalho
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            caminho_arquivo = os.path.join(
                desktop, f"estrutura_produto.xlsx"
            )

            # Gera Excel
            df.to_excel(caminho_arquivo, index=False, engine="openpyxl")

            # Abre o arquivo para formatar
            wb = load_workbook(caminho_arquivo)
            ws = wb.active

            # Formatos
            formato_moeda = 'R$ #,##0.00'
            formato_numero = '#,##0.00'
            formato_inteiro = '0'

            # Aplica formatação
            for linha in ws.iter_rows(min_row=2):
                linha[0].number_format = formato_inteiro  # Código
                linha[4].number_format = formato_numero  # Quantidade
                linha[5].number_format = formato_moeda  # Custo Unitário
                linha[6].number_format = formato_moeda  # Total

            # Ajusta largura das colunas
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 40
            ws.column_dimensions["C"].width = 18
            ws.column_dimensions["D"].width = 8
            ws.column_dimensions["E"].width = 14
            ws.column_dimensions["F"].width = 18
            ws.column_dimensions["G"].width = 18

            wb.save(caminho_arquivo)

            print(f"Excel gerado com sucesso:\n{caminho_arquivo}")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)


chama_classe = ExecutaCustoMateriais()