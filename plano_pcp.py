import sys
from banco_dados.conexao import conecta, conecta_robo
from banco_dados.controle_erros import grava_erro_banco
import os
import traceback
import inspect
from comandos.conversores import valores_para_float
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.header import Header
from email import encoders
from dados_email import email_user, password
from collections import defaultdict
from pdf2image import convert_from_path
from PIL import ImageFont, ImageDraw
from reportlab.lib.pagesizes import landscape, A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
import time
from datetime import datetime, date, timedelta
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Side, Alignment, Border, Font
import shutil


class ExecutaPlanoPcp:
    def __init__(self):
        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

        self.caminho_poppler = r'C:\Program Files\poppler-24.08.0\Library\bin'

        self.inicio_de_tudo_pi_abertas()
        self.inicio_de_tudo_ops_abertas()

    def trata_excecao(self, nome_funcao, mensagem, arquivo, excecao):
        try:
            tb = traceback.extract_tb(excecao)
            num_linha_erro = tb[-1][1]

            traceback.print_exc()
            print(f'Houve um problema no arquivo: {arquivo} na função: "{nome_funcao}"\n{mensagem} {num_linha_erro}')

            grava_erro_banco(nome_funcao, mensagem, arquivo, num_linha_erro)

        except Exception as e:
            nome_funcao_trat = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            tb = traceback.extract_tb(exc_traceback)
            num_linha_erro = tb[-1][1]
            print(f'Houve um problema no arquivo: {self.nome_arquivo} na função: "{nome_funcao_trat}"\n'
                  f'{e} {num_linha_erro}')
            grava_erro_banco(nome_funcao_trat, e, self.nome_arquivo, num_linha_erro)

    def dados_email(self):
        try:
            to = ['<maquinas@unisold.com.br>']

            current_time = (datetime.now())
            horario = current_time.strftime('%H')
            hora_int = int(horario)
            saudacao = ""
            if 4 < hora_int < 13:
                saudacao = "Bom Dia!"
            elif 12 < hora_int < 19:
                saudacao = "Boa Tarde!"
            elif hora_int > 18:
                saudacao = "Boa Noite!"
            elif hora_int < 5:
                saudacao = "Boa Noite!"

            msg_final = ""

            msg_final += f"Att,\n"
            msg_final += f"Suzuki Máquinas Ltda\n"
            msg_final += f"Fone (51) 3561.2583/(51) 3170.0965\n\n"
            msg_final += f"🟦 Mensagem gerada automaticamente pelo sistema de Planejamento e Controle da Produção (PCP) do ERP Suzuki.\n"
            msg_final += "🔸Por favor, não responda este e-mail diretamente."


            return saudacao, msg_final, to

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)
            return None

    def estrutura_prod_qtde_op(self, num_op, id_estrut):
        try:
            cursor = conecta.cursor()
            cursor.execute(f"SELECT estprod.id, prod.codigo, prod.descricao, "
                           f"COALESCE(prod.obs, '') as obs, prod.unidade, "
                           f"((SELECT quantidade FROM ordemservico where numero = {num_op}) * "
                           f"(estprod.quantidade)) AS Qtde, prod.quantidade, tip.tipomaterial "
                           f"FROM estrutura_produto as estprod "
                           f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                           f"LEFT JOIN tipomaterial tip ON prod.tipomaterial = tip.id "
                           f"where estprod.id_estrutura = {id_estrut} ORDER BY prod.descricao;")
            sel_estrutura = cursor.fetchall()

            return sel_estrutura

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)
            return None

    def teste(self, id_mat_e, lista_substitutos, num_op):
        try:
            saldo_substituto = 0

            cursor = conecta.cursor()
            cursor.execute(f"SELECT estprod.id, prod.codigo, prod.descricao "
                           f"FROM estrutura_produto as estprod "
                           f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                           f"where estprod.id = {id_mat_e};")
            sel_estrutura = cursor.fetchall()

            cod_original = sel_estrutura[0][1]

            cursor = conecta.cursor()
            cursor.execute(f"SELECT subs.cod_subs, prod.descricao, COALESCE(prod.obs, '') as obs, "
                           f"conj.conjunto, prod.unidade, prod.localizacao, prod.quantidade "
                           f"FROM SUBSTITUTO_MATERIAPRIMA as subs "
                           f"INNER JOIN produto as prod ON subs.cod_subs = prod.codigo "
                           f"INNER JOIN conjuntos conj ON prod.conjunto = conj.id "
                           f"WHERE subs.id_mat = {id_mat_e} "
                           f"and prod.quantidade > 0 "
                           f"AND subs.num_op = {num_op};")
            dados_mat_com_op = cursor.fetchall()
            if dados_mat_com_op:
                cod_subs, descr, ref, conj, um, local, saldo = dados_mat_com_op[0]

                dados = (cod_subs, descr, ref, um, local, saldo, cod_original)
                lista_substitutos.append(dados)

                saldo_substituto = float(saldo)

            else:
                cursor = conecta.cursor()
                cursor.execute(f"SELECT subs.cod_subs, prod.descricao, COALESCE(prod.obs, '') as obs, "
                               f"conj.conjunto, prod.unidade, prod.localizacao, prod.quantidade "
                               f"FROM SUBSTITUTO_MATERIAPRIMA as subs "
                               f"INNER JOIN produto as prod ON subs.cod_subs = prod.codigo "
                               f"INNER JOIN conjuntos conj ON prod.conjunto = conj.id "
                               f"WHERE subs.id_mat = {id_mat_e} "
                               f"and prod.quantidade > 0 "
                               f"AND subs.num_op is NULL;")
                dados_mat_sem_op = cursor.fetchall()
                if dados_mat_sem_op:
                    cod_subs, descr, ref, conj, um, local, saldo = dados_mat_sem_op[0]

                    dados = (cod_subs, descr, ref, um, local, saldo, cod_original)
                    lista_substitutos.append(dados)

                    saldo_substituto = float(saldo)

            return lista_substitutos, saldo_substituto

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)
            return None

    def consumo_op_por_id(self, num_op, id_materia_prima):
        try:
            cursor = conecta.cursor()
            cursor.execute(f"SELECT SUM(prodser.QTDE_ESTRUT_PROD) AS total_quantidade "
                           f"FROM produtoos AS prodser "
                           f"WHERE prodser.numero = {num_op} AND prodser.id_estrut_prod = {id_materia_prima};")
            total_quantidade = cursor.fetchone()[0]

            total_quantidade = total_quantidade if total_quantidade is not None else 0

            return total_quantidade

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)
            return None

    def verifica_ops_concluidas(self, num_op, id_estrut):
        try:
            lista_substitutos = []

            select_estrut = self.estrutura_prod_qtde_op(num_op, id_estrut)

            falta_material = []

            for dados_estrut in select_estrut:
                saldo_final = 0

                id_mat_e, cod_e, descr_e, ref_e, um_e, qtde_e, saldo_prod_e, tipo_e = dados_estrut

                lista_substitutos, saldo_subs = self.teste(id_mat_e, lista_substitutos, num_op)

                qtde_e_float = valores_para_float(qtde_e)
                saldo_final += saldo_subs + valores_para_float(saldo_prod_e)

                qtde_total_item_op = self.consumo_op_por_id(num_op, id_mat_e)

                qtde_total_item_op_float = valores_para_float(qtde_total_item_op)

                if qtde_total_item_op_float < qtde_e_float:
                    sobras = qtde_e_float - saldo_final - qtde_total_item_op_float
                    if sobras > 0:
                        dados = (cod_e, descr_e, ref_e, um_e, sobras)
                        falta_material.append(dados)

            return falta_material

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)
            return None

    def ops_abertas(self):
        try:
            tudo_tudo = []

            cursor = conecta.cursor()
            cursor.execute(f"SELECT op.numero, op.codigo, op.id_estrutura, prod.descricao, "
                           f"COALESCE(prod.obs, ''), "
                           f"prod.unidade, COALESCE(prod.tipomaterial, ''), op.quantidade "
                           f"FROM ordemservico as op "
                           f"INNER JOIN produto as prod ON op.produto = prod.id "
                           f"where op.status = 'A' order by op.numero;")
            ops_abertas = cursor.fetchall()

            if ops_abertas:
                for i in ops_abertas:
                    num_op, cod, id_estrut, descr, ref, um, tipo, qtde = i

                    print("agrupando ops abertas: ", i)

                    material_faltando = self.verifica_ops_concluidas(num_op, id_estrut)

                    if material_faltando:
                        for ii in material_faltando:
                            tudo_tudo.append(ii)

            # === AQUI entra o agrupamento por código ===
            agrupado = {}
            for cod, descr, ref, um, sobras in tudo_tudo:
                sobras_float = valores_para_float(sobras)

                if cod in agrupado:
                    agrupado[cod][4] += sobras_float  # soma as quantidades corretamente
                else:
                    agrupado[cod] = [cod, descr, ref, um, sobras_float]

            # Converte de volta para lista de tuplas
            tudo_tudo = [tuple(v) for v in agrupado.values()]

            return tudo_tudo

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def pedidos_abertos(self):
        try:
            tabela_nova = []

            cursor = conecta.cursor()
            cursor.execute(
                "SELECT prod.codigo, prod.descricao, "
                "COALESCE(prod.obs, ''), "
                "prod.unidade, "
                "SUM(prodint.qtde) as total_qtde "
                "FROM PRODUTOPEDIDOINTERNO AS prodint "
                "INNER JOIN produto AS prod ON prodint.id_produto = prod.id "
                "WHERE prodint.status = 'A' "
                "GROUP BY prod.codigo, prod.descricao, prod.obs, prod.unidade "
                "ORDER BY prod.codigo;"
            )
            dados_agrupados = cursor.fetchall()

            if dados_agrupados:
                for i in dados_agrupados:
                    cod, descr, ref, um, qtde = i

                    dados = (cod, descr, ref, um, qtde)
                    tabela_nova.append(dados)

            cursor = conecta.cursor()
            cursor.execute(f"SELECT prod.codigo, prod.descricao, "
                           f"COALESCE(prod.obs, ''), prod.unidade, (prodoc.quantidade - prodoc.produzido) "
                           f"FROM PRODUTOORDEMCOMPRA as prodoc "
                           f"INNER JOIN produto as prod ON prodoc.produto = prod.id "
                           f"INNER JOIN ordemcompra as oc ON prodoc.mestre = oc.id "
                           f"where prodoc.quantidade > prodoc.produzido "
                           f"and oc.status = 'A' "
                           f"and oc.entradasaida = 'S';")
            dados_interno = cursor.fetchall()
            if dados_interno:
                for i in dados_interno:
                    cod, descr, ref, um, qtde = i

                    dados = (cod, descr, ref, um, qtde)
                    tabela_nova.append(dados)

            return tabela_nova


        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def inicio_de_tudo_pi_abertas(self):
        try:
            lista_final = []

            dados_pi_abertas = self.pedidos_abertos()

            if dados_pi_abertas:
                for i in dados_pi_abertas:
                    qtde_nec = 0

                    qtde_nec_f = 0

                    cod, descr, ref, um, qtde_pi = i

                    print("pelas pis abertos: ", cod, descr, ref, um, qtde_pi)

                    qtde_pi_float = valores_para_float(qtde_pi)

                    qtde_nec += qtde_pi_float

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT prod.conjunto, prod.tipomaterial, tip.tipomaterial, conj.conjunto, "
                                   f"prod.quantidade "
                                   f"FROM produto as prod "
                                   f"LEFT JOIN conjuntos conj ON prod.conjunto = conj.id "
                                   f"LEFT JOIN tipomaterial tip ON prod.tipomaterial = tip.id "
                                   f"where prod.codigo = {cod};")
                    detalhes_produto = cursor.fetchall()
                    num_conj, num_tipo, tipo, conjunto, saldo = detalhes_produto[0]

                    qtde_nec, prod_sem_estrut, prod_sem_des, duplicado = self.verifica_situacao_produto(cod, ref, saldo, num_conj, qtde_nec)

                    if duplicado:
                        self.envia_email_desenho_duplicado(cod, ref)
                    else:
                        if prod_sem_des:
                            self.envia_email_nao_acha_desenho(ref, cod)
                        else:
                            if prod_sem_estrut:
                                s = re.sub(r"[^\d.]", "", ref)  # remove tudo que não é número ou ponto
                                s = re.sub(r"\.+$", "", s) # saída: 47.00.014.07

                                caminho_pdf = rf"\\Publico\C\OP\Projetos\{s}.pdf"
                                arquivo_pdf = f"{s}.pdf"

                                if os.path.exists(caminho_pdf):
                                    self.envia_email_sem_estrutura(caminho_pdf, arquivo_pdf, i)
                                else:
                                    self.envia_email_nao_acha_desenho(arquivo_pdf, cod)
                            else:
                                if qtde_nec > 0:
                                    qtde_nec_f += qtde_nec

                                    if num_tipo == 119:
                                        estrutura = self.manipula_dados_tabela_estrutura(cod)

                                        for ii in estrutura:
                                            cod_f, descr_f, ref_f, um_f, qtde_f, saldo_f, num_conj_f, tipo_f, num_tipo_f = ii

                                            qtde_nec_f, prod_sem_estr_f, prod_sem_des_f, duplicado_f = self.verifica_situacao_produto(cod_f, ref_f, saldo_f,
                                                                                                                     num_conj_f,
                                                                                                                     qtde_nec_f)
                                            if duplicado_f:
                                                self.envia_email_desenho_duplicado(cod, ref)
                                            else:
                                                if prod_sem_des_f:
                                                    self.envia_email_nao_acha_desenho(ref_f, cod_f)
                                                else:
                                                    if prod_sem_estr_f:
                                                        s_f = re.sub(r"[^\d.]", "", ref_f)
                                                        s_f = re.sub(r"\.+$", "", s_f)

                                                        caminho_pdf_f = rf"\\Publico\C\OP\Projetos\{s_f}.pdf"
                                                        arquivo_pdf_f = f"{s_f}.pdf"

                                                        if os.path.exists(caminho_pdf_f):
                                                            self.envia_email_sem_estrutura(caminho_pdf_f, arquivo_pdf_f, ii)
                                                        else:
                                                            self.envia_email_nao_acha_desenho(arquivo_pdf_f, cod_f)
                                                    else:
                                                        if qtde_nec_f > 0:
                                                            if num_tipo_f != 80:
                                                                dados = (cod_f, descr_f, ref_f, um_f, num_conj_f, num_tipo_f, tipo_f, qtde_nec_f)
                                                                lista_final.append(dados)
                                    else:
                                        if num_tipo != 80:
                                            dados = (cod, descr, ref, um, num_conj, num_tipo, tipo, qtde_nec)
                                            lista_final.append(dados)

            if lista_final:
                self.executar_tarefa(lista_final)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def inicio_de_tudo_ops_abertas(self):
        try:
            print("entrei nas ops abertas")
            lista_final = []

            dados_ops_abertas = self.ops_abertas()

            if dados_ops_abertas:
                for i in dados_ops_abertas:
                    qtde_nec = 0

                    qtde_nec_f = 0

                    cod, descr, ref, um, qtde_pi = i

                    print("pelas ops abertas: ", cod, descr, ref, um, qtde_pi)

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT prod.conjunto, prod.tipomaterial, tip.tipomaterial, conj.conjunto, "
                                   f"prod.quantidade "
                                   f"FROM produto as prod "
                                   f"LEFT JOIN conjuntos conj ON prod.conjunto = conj.id "
                                   f"LEFT JOIN tipomaterial tip ON prod.tipomaterial = tip.id "
                                   f"where prod.codigo = {cod};")
                    detalhes_produto = cursor.fetchall()
                    num_conj, num_tipo, tipo, conjunto, saldo = detalhes_produto[0]

                    qtde_nec, prod_sem_estrut, prod_sem_des, duplicado = self.verifica_situacao_produto(cod, ref, saldo, num_conj, qtde_nec)

                    if duplicado:
                        self.envia_email_desenho_duplicado(cod, ref)
                    else:
                        if prod_sem_des:
                            self.envia_email_nao_acha_desenho(ref, cod)
                        else:
                            if prod_sem_estrut:
                                s = re.sub(r"[^\d.]", "", ref)  # remove tudo que não é número ou ponto
                                s = re.sub(r"\.+$", "", s) # saída: 47.00.014.07

                                caminho_pdf = rf"\\Publico\C\OP\Projetos\{s}.pdf"
                                arquivo_pdf = f"{s}.pdf"

                                if os.path.exists(caminho_pdf):
                                    self.envia_email_sem_estrutura(caminho_pdf, arquivo_pdf, i)
                                else:
                                    self.envia_email_nao_acha_desenho(arquivo_pdf, cod)
                            else:
                                if qtde_nec > 0:
                                    qtde_nec_f += qtde_nec

                                    if num_tipo == 119:
                                        estrutura = self.manipula_dados_tabela_estrutura(cod)

                                        for ii in estrutura:
                                            cod_f, descr_f, ref_f, um_f, qtde_f, saldo_f, num_conj_f, tipo_f, num_tipo_f = ii

                                            qtde_nec_f, prod_sem_estr_f, prod_sem_des_f, duplicado_f = self.verifica_situacao_produto(cod_f, ref_f, saldo_f,
                                                                                                                     num_conj_f,
                                                                                                                     qtde_nec_f)
                                            if duplicado_f:
                                                self.envia_email_desenho_duplicado(cod, ref)
                                            else:
                                                if prod_sem_des_f:
                                                    self.envia_email_nao_acha_desenho(ref_f, cod_f)
                                                else:
                                                    if prod_sem_estr_f:
                                                        s_f = re.sub(r"[^\d.]", "", ref_f)
                                                        s_f = re.sub(r"\.+$", "", s_f)

                                                        caminho_pdf_f = rf"\\Publico\C\OP\Projetos\{s_f}.pdf"
                                                        arquivo_pdf_f = f"{s_f}.pdf"

                                                        if os.path.exists(caminho_pdf_f):
                                                            self.envia_email_sem_estrutura(caminho_pdf_f, arquivo_pdf_f, ii)
                                                        else:
                                                            self.envia_email_nao_acha_desenho(arquivo_pdf_f, cod_f)
                                                    else:
                                                        if qtde_nec_f > 0:
                                                            if num_tipo_f != 80:
                                                                dados = (cod_f, descr_f, ref_f, um_f, num_conj_f, num_tipo_f, tipo_f, qtde_nec_f)
                                                                lista_final.append(dados)
                                    else:
                                        if num_tipo != 80:
                                            dados = (cod, descr, ref, um, num_conj, num_tipo, tipo, qtde_nec)
                                            lista_final.append(dados)

                if lista_final:
                    self.executar_tarefa(lista_final)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def executar_tarefa(self, dados_produtos):
        try:
            print("executar tarefa")
            # --- Agrupamento e soma de quantidades ---
            produtos_unicos = defaultdict(lambda: list())

            for item in dados_produtos:
                cod, descr, ref, um, conjunto, num_tipo, tipo,  qtde = item

                if cod not in produtos_unicos:
                    produtos_unicos[cod] = [cod, descr, ref, um, conjunto, num_tipo, tipo, qtde]
                else:
                    produtos_unicos[cod][6] += qtde  # soma a quantidade

            tem_conjunto_10 = any(i[4] == 10 for i in produtos_unicos.values())

            if tem_conjunto_10:
                for i in produtos_unicos.values():
                    cod, descr, ref, um, conjunto, num_tipo, tipo, qtde = i

                    print("executar tarefa acabados: ", i)

                    qtde = round(qtde, 3)

                    if conjunto == 10:
                        if tipo != 119:
                            cursor = conecta.cursor()
                            cursor.execute(f"select id, numero from ordemservico "
                                           f"where numero = (select max(numero) from ordemservico);")
                            select_numero = cursor.fetchall()
                            idez, num = select_numero[0]
                            num_op = int(num) + 1

                            emissao = date.today()
                            emissao_br = emissao.strftime('%d/%m/%Y')

                            situacao = self.criar_op(num_op, cod, qtde, emissao)

                            if situacao:
                                if num_tipo == 87:
                                    self.cria_pdf_envia_email_conjunto(num_op, ref, qtde, emissao_br)
                                else:
                                    self.cria_pdf_envia_email_usinagem(num_op, ref, qtde, emissao_br)
            else:
                print("Nenhum item com conjunto == 10")

                agrupado = defaultdict(list)

                for i in produtos_unicos.values():
                    cod, descr, ref, um, conjunto, num_tipo, tipo, qtde = i

                    print("executar tarefa comprados: ", i)

                    qtde = round(qtde, 3)

                    ops_destino = self.manipula_dados_tabela_consumo_final(cod)
                    vendas_destino = self.manipula_dados_tabela_venda_final(cod)

                    if conjunto != 10 and cod != "60390":
                        agrupado[num_tipo, tipo].append((cod, descr, ref, um, conjunto, num_tipo, tipo, qtde, ops_destino, vendas_destino))

                # --- Impressão agrupada ---
                for (num_tipo, tipo), itens  in agrupado.items():
                    if not num_tipo:
                        self.envia_email_sem_tipo(itens)
                    else:
                        cursor = conecta.cursor()
                        cursor.execute(f"select id, orcamento, desenho "
                                       f"from tipomaterial "
                                       f"where id = {num_tipo};")
                        select_numero = cursor.fetchall()
                        precisa_orcamento = select_numero[0][1]
                        precisa_desenho = select_numero[0][2]

                        if precisa_orcamento == "S":
                            if precisa_desenho == "S":
                                todos_tem_desenho = True
                                desenhos_faltando = []

                                lista_nao_enviado_orcamento = []

                                for iiitens in itens:
                                    cod, descr, ref, um, conjunto, num_tipo, tipo, qtde, ops_dest, vnd_dest = iiitens

                                    tem_orcamento = self.verifica_orcamento_enviado_banco(cod)

                                    if not tem_orcamento:
                                        dados = (cod, descr, ref, um, conjunto, num_tipo, tipo, qtde, ops_dest, vnd_dest)
                                        lista_nao_enviado_orcamento.append(dados)

                                        caminho, arq = self.verifica_se_tem_pdf_desenho(ref)

                                        if not caminho:
                                            todos_tem_desenho = False
                                            dadinhos = (ref, cod)
                                            desenhos_faltando.append(dadinhos)
                                            break
                                if lista_nao_enviado_orcamento:
                                    if todos_tem_desenho:
                                        self.orcamento_com_desenho(num_tipo, tipo, lista_nao_enviado_orcamento)
                                    else:
                                        for titi in desenhos_faltando:
                                            ref, cod = titi

                                            self.envia_email_nao_acha_desenho(ref, cod)
                            else:
                                lista_nao_enviado_orcamento = []

                                for iiitens in itens:
                                    cod, descr, ref, um, conjunto, num_tipo, tipo, qtde, ops_dest, vnd_dest = iiitens

                                    tem_orcamento = self.verifica_orcamento_enviado_banco(cod)

                                    if not tem_orcamento:
                                        dados = (cod, descr, ref, um, conjunto, num_tipo, tipo, qtde, ops_dest,
                                                 vnd_dest)
                                        lista_nao_enviado_orcamento.append(dados)

                                if lista_nao_enviado_orcamento:
                                    self.orcamento_sem_desenho(num_tipo, tipo, lista_nao_enviado_orcamento)

                        else:
                            if precisa_desenho == "S":
                                if num_tipo == 84 or num_tipo == 85 or num_tipo == 116 or num_tipo == 125:
                                    todos_tem_desenho = True
                                    desenhos_faltando = []

                                    for iiitens in itens:
                                        cod, descr, ref, um, conjunto, num_tipo, tipo, qtde, ops_dest, vnd_dest = iiitens

                                        caminho_pdf, arq_pdf, caminho_dwg, arq_dwg = self.verifica_se_tem_pdf_desenho_chapa(ref)

                                        if not caminho_pdf or not caminho_dwg:
                                            todos_tem_desenho = False
                                            dadinhos = (ref, cod)
                                            desenhos_faltando.append(dadinhos)
                                            break
                                    if todos_tem_desenho:
                                        self.solicitacao_com_desenho_chapa(tipo, itens)
                                    else:
                                        for titi in desenhos_faltando:
                                            ref, cod = titi

                                            self.envia_email_nao_acha_desenho(ref, cod)
                                else:
                                    todos_tem_desenho = True
                                    desenhos_faltando = []

                                    for iiitens in itens:
                                        cod, descr, ref, um, conjunto, num_tipo, tipo, qtde, ops_dest, vnd_dest = iiitens

                                        caminho, arq = self.verifica_se_tem_pdf_desenho(ref)

                                        if not caminho:
                                            todos_tem_desenho = False
                                            dadinhos = (ref, cod)
                                            desenhos_faltando.append(dadinhos)
                                            break
                                    if todos_tem_desenho:
                                        self.solicitacao_com_desenho(tipo, itens)
                                    else:
                                        for titi in desenhos_faltando:
                                            ref, cod = titi

                                            self.envia_email_nao_acha_desenho(ref, cod)
                            else:
                                self.solicitacao_sem_desenho(tipo, itens)
        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_se_tem_pdf_desenho_chapa(self, ref):
        try:
            s = re.sub(r"[^\d.]", "", ref)  # remove tudo que não é número ou ponto
            s = re.sub(r"\.+$", "", s)

            caminho_pdf = rf"\\Publico\C\OP\Projetos\{s}.pdf"
            arquivo_pdf = f"{s}.pdf"
            caminho_dwg = rf"\\Publico\C\OP\Projetos\{s}.dwg"
            arquivo_dwg = f"{s}.dwg"

            if not os.path.exists(caminho_pdf):
                caminho_pdf = ""
                arquivo_pdf = ""

            if not os.path.exists(caminho_dwg):
                caminho_dwg = ""
                arquivo_dwg = ""

            return caminho_pdf, arquivo_pdf, caminho_dwg, arquivo_dwg

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_se_tem_pdf_desenho(self, ref):
        try:
            s = re.sub(r"[^\d.]", "", ref)  # remove tudo que não é número ou ponto
            s = re.sub(r"\.+$", "", s)  # saída: 47.00.014.07

            caminho_pdf = rf"\\Publico\C\OP\Projetos\{s}.pdf"
            arquivo_pdf = f"{s}.pdf"

            if not os.path.exists(caminho_pdf):
                caminho_pdf = ""
                arquivo_pdf = ""

            return caminho_pdf, arquivo_pdf

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_orcamento_enviado_banco(self, codigo):
        try:
            tem_orcamento = False

            cur = conecta.cursor()
            cur.execute(f"SELECT id, descricao, COALESCE(obs, '') as obs, unidade, id_versao "
                        f"FROM produto where codigo = {codigo};")
            detalhes_produto = cur.fetchall()
            id_prod, descricao_id, referencia_id, unidade_id, id_versao = detalhes_produto[0]

            cur = conecta_robo.cursor()
            cur.execute(f"SELECT id, id_produto "
                        f"FROM PRODUTO_ORCAMENTO where id_produto = {id_prod};")
            detalhes_orcamento = cur.fetchall()

            if detalhes_orcamento:
                tem_orcamento = True

            return tem_orcamento

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def envia_email_orcamento_com_desenho(self, num_sol, tipo, dados_itens, arquivo_excel):
        try:
            saudacao, msg_final, to = self.dados_email()

            subject = f"Solicitação de Orçamento – Compra Nº {num_sol} | Grupo {tipo}"

            msg = MIMEMultipart()
            msg['From'] = email_user
            msg['Subject'] = subject

            body = ""

            body += f"{saudacao}\n\n"
            body += f"Segue em anexo a solicitação de orçamento de Compra Nº {num_sol} do grupo de Fornecedores {tipo}, solicitada por Suzuki Máquinas.\n\n"
            body += f"Agradecemos pela atenção e pedimos que, em caso de dúvidas ou dificuldades com o arquivo, entre em contato pelo e-mail maquinas@unisold.com.br\n\n"
            body += f"{msg_final}"

            msg.attach(MIMEText(body, 'plain'))

            caminho_arquivo = fr'C:\Users\Anderson\PycharmProjects\robo_boby\{arquivo_excel}'

            attachment = open(caminho_arquivo, 'rb')

            part = MIMEBase('application', "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment',
                            filename=Header(arquivo_excel, 'utf-8').encode())
            msg.attach(part)

            for i in dados_itens:
                cod, descr, ref, um, conjunto, num_tipo, tipo, qtde, ops_destino, vendas_destino = i

                caminho, arq = self.verifica_se_tem_pdf_desenho(ref)

                with open(caminho, 'rb') as attachment:
                    part = MIMEBase('application', "octet-stream")
                    part.set_payload(attachment.read())
                    encoders.encode_base64(part)
                    part.add_header('Content-Disposition', 'attachment',
                                    filename=Header(arq, 'utf-8').encode())
                    msg.attach(part)

            text = msg.as_string()
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(email_user, password)

            server.sendmail(email_user, to, text)
            attachment.close()

            server.quit()

            print(f'Orçamento {num_sol} enviado com sucesso!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def envia_email_orcamento_sem_desenho(self, num_sol, tipo, arquivo_excel):
        try:
            saudacao, msg_final, to = self.dados_email()

            subject = f"Solicitação de Orçamento – Compra Nº {num_sol} | Grupo {tipo}"

            msg = MIMEMultipart()
            msg['From'] = email_user
            msg['Subject'] = subject

            body = ""

            body += f"{saudacao}\n\n"
            body += f"Segue em anexo a solicitação de orçamento de Compra Nº {num_sol} do grupo de Fornecedores {tipo}, solicitada por Suzuki Máquinas.\n\n"
            body += f"Agradecemos pela atenção e pedimos que, em caso de dúvidas ou dificuldades com o arquivo, entre em contato pelo e-mail maquinas@unisold.com.br\n\n"
            body += f"{msg_final}"

            msg.attach(MIMEText(body, 'plain'))

            caminho_arquivo = fr'C:\Users\Anderson\PycharmProjects\robo_boby\{arquivo_excel}'

            attachment = open(caminho_arquivo, 'rb')

            part = MIMEBase('application', "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment',
                            filename=Header(arquivo_excel, 'utf-8').encode())
            msg.attach(part)

            text = msg.as_string()
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(email_user, password)

            server.sendmail(email_user, to, text)
            attachment.close()

            server.quit()

            print(f'Orçamento {num_sol} enviado com sucesso!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def solicitacao_com_desenho_chapa(self, tipo, dados_itens):
        try:
            observacao_certa = f"SOLICITAÇÃO GERADO AUTOMATICAMENTE - {tipo}"

            nome_computador = "BOBY DE AZEVEDO"

            data_hoje = date.today()

            cursor = conecta.cursor()
            cursor.execute("select GEN_ID(GEN_ORDEMSOLICITACAO_ID,0) from rdb$database;")
            ultimo_req0 = cursor.fetchall()
            ultimo_req1 = ultimo_req0[0]
            ultimo_req = int(ultimo_req1[0]) + 1

            cursor = conecta.cursor()
            cursor.execute("""
                INSERT INTO ordemsolicitacao (IDSOLICITACAO, DATAEMISSAO, STATUS, OBS, NOME_PC)
                VALUES (GEN_ID(GEN_ORDEMSOLICITACAO_ID,1), ?, ?, ?, ?)
            """, (data_hoje, 'A', observacao_certa, nome_computador))

            for indice, itens in enumerate(dados_itens, start=1):
                cod, descr, ref, um, conjunto, num_tipo, tipo, qtde, ops_dest, vnd_dest = itens

                msg = ""

                if ops_dest:
                    for opss in ops_dest:
                        num_op, cod_op, descr_op = opss

                        if msg:
                            msg += f"\n- {cod_op} - {descr_op}"
                        else:
                            msg += f"- {cod_op} - {descr_op}"

                if vnd_dest:
                    for vendass in vnd_dest:
                        num_pi, num_ov, cliente = vendass

                        if msg:
                            if num_pi:
                                numero = f"\nPI {num_pi}"
                            else:
                                numero = f"\nPI {num_ov}"
                        else:
                            if num_pi:
                                numero = f"PI {num_pi}"
                            else:
                                numero = f"PI {num_ov}"

                        msg += f"- {numero}\n"

                cursor = conecta.cursor()
                cursor.execute(f"SELECT id, embalagem "
                               f"FROM produto where codigo = '{cod}';")
                dados_prod = cursor.fetchall()

                id_prod = dados_prod[0][0]

                cursor = conecta.cursor()
                cursor.execute(f"Insert into produtoordemsolicitacao (ID, MESTRE, ITEM, PRODUTO, QUANTIDADE, "
                               f"DATA, STATUS, DESTINO) "
                               f"values (GEN_ID(GEN_PRODUTOORDEMSOLICITACAO_ID,1), {ultimo_req}, {indice}, "
                               f"{id_prod}, '{qtde}', '{data_hoje}', "
                               f"'A', '{msg}');")

            self.grava_anexo_solicitacao_chapa(ultimo_req, dados_itens)
            conecta.commit()

            print(f"Solicitação Nº {ultimo_req} salva com sucesso!")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def solicitacao_com_desenho(self, tipo, dados_itens):
        try:
            observacao_certa = f"SOLICITAÇÃO GERADO AUTOMATICAMENTE - {tipo}"

            nome_computador = "BOBY DE AZEVEDO"

            data_hoje = date.today()

            cursor = conecta.cursor()
            cursor.execute("select GEN_ID(GEN_ORDEMSOLICITACAO_ID,0) from rdb$database;")
            ultimo_req0 = cursor.fetchall()
            ultimo_req1 = ultimo_req0[0]
            ultimo_req = int(ultimo_req1[0]) + 1

            cursor = conecta.cursor()
            cursor.execute(f"Insert into ordemsolicitacao (IDSOLICITACAO, DATAEMISSAO, STATUS, OBS, NOME_PC) "
                           f"values (GEN_ID(GEN_ORDEMSOLICITACAO_ID,1), "
                           f"'{data_hoje}', 'A', '{observacao_certa}', '{nome_computador}');")

            for indice, itens in enumerate(dados_itens, start=1):
                cod, descr, ref, um, conjunto, num_tipo, tipo, qtde, ops_dest, vnd_dest = itens

                msg = ""

                if ops_dest:
                    for opss in ops_dest:
                        num_op, cod_op, descr_op = opss

                        if msg:
                            msg += f"\n- {cod_op} - {descr_op}"
                        else:
                            msg += f"- {cod_op} - {descr_op}"

                if vnd_dest:
                    for vendass in vnd_dest:
                        num_pi, num_ov, cliente = vendass

                        if msg:
                            if num_pi:
                                numero = f"\nPI {num_pi}"
                            else:
                                numero = f"\nPI {num_ov}"
                        else:
                            if num_pi:
                                numero = f"PI {num_pi}"
                            else:
                                numero = f"PI {num_ov}"

                        msg += f"- {numero}\n"

                cursor = conecta.cursor()
                cursor.execute(f"SELECT id, embalagem "
                               f"FROM produto where codigo = '{cod}';")
                dados_prod = cursor.fetchall()

                id_prod = dados_prod[0][0]

                cursor = conecta.cursor()
                cursor.execute(f"Insert into produtoordemsolicitacao (ID, MESTRE, ITEM, PRODUTO, QUANTIDADE, "
                               f"DATA, STATUS, DESTINO) "
                               f"values (GEN_ID(GEN_PRODUTOORDEMSOLICITACAO_ID,1), {ultimo_req}, {indice}, "
                               f"{id_prod}, '{qtde}', '{data_hoje}', "
                               f"'A', '{msg}');")

            self.grava_anexo_solicitacao(ultimo_req, dados_itens)
            conecta.commit()

            print(f"Solicitação Nº {ultimo_req} salva com sucesso!")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def solicitacao_sem_desenho(self, tipo, dados_itens):
        try:
            observacao_certa = f"SOLICITAÇÃO GERADO AUTOMATICAMENTE - {tipo}"

            nome_computador = "BOBY DE AZEVEDO"

            data_hoje = date.today()

            cursor = conecta.cursor()
            cursor.execute("select GEN_ID(GEN_ORDEMSOLICITACAO_ID,0) from rdb$database;")
            ultimo_req0 = cursor.fetchall()
            ultimo_req1 = ultimo_req0[0]
            ultimo_req = int(ultimo_req1[0]) + 1

            cursor = conecta.cursor()
            cursor.execute(f"Insert into ordemsolicitacao (IDSOLICITACAO, DATAEMISSAO, STATUS, OBS, NOME_PC) "
                           f"values (GEN_ID(GEN_ORDEMSOLICITACAO_ID,1), "
                           f"'{data_hoje}', 'A', '{observacao_certa}', '{nome_computador}');")

            for indice, itens in enumerate(dados_itens, start=1):
                cod, descr, ref, um, conjunto, num_tipo, tipo, qtde, ops_dest, vnd_dest = itens

                msg = ""

                if ops_dest:
                    for opss in ops_dest:
                        num_op, cod_op, descr_op = opss

                        if msg:
                            msg += f"\n- {cod_op} - {descr_op}"
                        else:
                            msg += f"- {cod_op} - {descr_op}"

                if vnd_dest:
                    for vendass in vnd_dest:
                        num_pi, num_ov, cliente = vendass

                        if msg:
                            if num_pi:
                                numero = f"\nPI {num_pi}"
                            else:
                                numero = f"\nPI {num_ov}"
                        else:
                            if num_pi:
                                numero = f"PI {num_pi}"
                            else:
                                numero = f"PI {num_ov}"

                        msg += f"- {numero}\n"

                cursor = conecta.cursor()
                cursor.execute(f"SELECT id, embalagem "
                               f"FROM produto where codigo = '{cod}';")
                dados_prod = cursor.fetchall()

                id_prod = dados_prod[0][0]

                msg = msg.replace("\n", " ")[:200]

                cursor = conecta.cursor()
                cursor.execute(f"Insert into produtoordemsolicitacao (ID, MESTRE, ITEM, PRODUTO, QUANTIDADE, "
                               f"DATA, STATUS, DESTINO) "
                               f"values (GEN_ID(GEN_PRODUTOORDEMSOLICITACAO_ID,1), {ultimo_req}, {indice}, "
                               f"{id_prod}, '{qtde}', '{data_hoje}', "
                               f"'A', '{msg}');")

            conecta.commit()

            print(f"Solicitação Nº {ultimo_req} salva com sucesso!")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def grava_anexo_solicitacao_chapa(self, num_sol, dados_itens):
        try:
            for i in dados_itens:
                cod, descr, ref, um, conjunto, num_tipo, tipo, qtde, ops_dest, vnd_dest = i

                caminho_pdf, arq_pdf, caminho_dwg, arq_dwg = self.verifica_se_tem_pdf_desenho_chapa(ref)

                nome_arquivo_usuario_pdf = os.path.basename(caminho_pdf)
                nome_arquivo_final_pdf = f'{num_sol} - {nome_arquivo_usuario_pdf}'

                destin_path_pdf = os.path.join(r'\\PUBLICO\Python\0 - Versões Antigas\anexos', nome_arquivo_final_pdf)

                shutil.copy2(caminho_pdf, destin_path_pdf)

                cursor = conecta.cursor()
                cursor.execute(f"Insert into SOLICITACAO_ANEXO (ID, CAMINHO, ID_SOLICITACAO) "
                               f"values (GEN_ID(GEN_SOLICITACAO_ANEXO_ID,1), '{destin_path_pdf}', {num_sol});")

                nome_arquivo_usuario_dwg = os.path.basename(caminho_dwg)
                nome_arquivo_final_dwg = f'{num_sol} - {nome_arquivo_usuario_dwg}'

                destin_path_dwg = os.path.join(r'\\PUBLICO\Python\0 - Versões Antigas\anexos', nome_arquivo_final_dwg)

                shutil.copy2(caminho_dwg, destin_path_dwg)

                cursor = conecta.cursor()
                cursor.execute(f"Insert into SOLICITACAO_ANEXO (ID, CAMINHO, ID_SOLICITACAO) "
                               f"values (GEN_ID(GEN_SOLICITACAO_ANEXO_ID,1), '{destin_path_dwg}', {num_sol});")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def grava_anexo_solicitacao(self, num_sol, dados_itens):
        try:
            for i in dados_itens:
                cod, descr, ref, um, conjunto, num_tipo, tipo, qtde, ops_dest, vnd_dest = i

                caminho, arq = self.verifica_se_tem_pdf_desenho(ref)

                nome_arquivo_usuario = os.path.basename(caminho)
                nome_arquivo_final = f'{num_sol} - {nome_arquivo_usuario}'

                destination_path = os.path.join(r'\\PUBLICO\Python\0 - Versões Antigas\anexos', nome_arquivo_final)

                shutil.copy2(caminho, destination_path)

                cursor = conecta.cursor()
                cursor.execute(f"Insert into SOLICITACAO_ANEXO (ID, CAMINHO, ID_SOLICITACAO) "
                               f"values (GEN_ID(GEN_SOLICITACAO_ANEXO_ID,1), '{destination_path}', {num_sol});")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def orcamento_com_desenho(self, num_tipo, tipo, dados_itens):
        try:
            cursor = conecta_robo.cursor()
            cursor.execute(f"select GEN_ID(GEN_ENVIA_ORCAMENTO_ID,0) from rdb$database;")
            dados0 = cursor.fetchall()
            dados1 = dados0[0]
            dados2 = int(dados1[0]) + 1
            dados3 = str(dados2)
            num_sol = dados3

            arquivo_excel = self.gera_excel_orcamento(num_sol, dados_itens)

            self.envia_email_orcamento_com_desenho(num_sol, tipo, dados_itens, arquivo_excel)

            self.excluir_arquivo(arquivo_excel)

            self.inserir_banco_orcamento(num_tipo, dados_itens)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def orcamento_sem_desenho(self, num_tipo, tipo, dados_itens):
        try:
            cursor = conecta_robo.cursor()
            cursor.execute(f"select GEN_ID(GEN_ENVIA_ORCAMENTO_ID,0) from rdb$database;")
            dados0 = cursor.fetchall()
            dados1 = dados0[0]
            dados2 = int(dados1[0]) + 1
            dados3 = str(dados2)
            num_sol = dados3

            arquivo_excel = self.gera_excel_orcamento(num_sol, dados_itens)

            self.envia_email_orcamento_sem_desenho(num_sol, tipo, arquivo_excel)

            self.excluir_arquivo(arquivo_excel)

            self.inserir_banco_orcamento(num_tipo, dados_itens)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def inserir_banco_orcamento(self, num_tipo, dados_itens):
        try:
            cursor = conecta_robo.cursor()
            cursor.execute("select GEN_ID(GEN_ENVIA_ORCAMENTO_ID,0) from rdb$database;")
            ultimo_req0 = cursor.fetchall()
            ultimo_req1 = ultimo_req0[0]
            ultimo_req = int(ultimo_req1[0]) + 1

            cursor = conecta_robo.cursor()
            cursor.execute(f"Insert into ENVIA_ORCAMENTO (ID, ID_TIPO) "
                           f"values (GEN_ID(GEN_ENVIA_ORCAMENTO_ID,1), {num_tipo});")

            for i in dados_itens:
                cod, descr, ref, um, conjunto, num_tipo, tipo, qtde, ops_dest, vnd_dest = i

                cur = conecta.cursor()
                cur.execute(f"SELECT id, descricao, COALESCE(obs, '') as obs, unidade, id_versao "
                            f"FROM produto where codigo = {cod};")
                detalhes_produto = cur.fetchall()
                id_prod, descricao_id, referencia_id, unidade_id, id_versao = detalhes_produto[0]

                cursor = conecta_robo.cursor()
                cursor.execute(f"Insert into PRODUTO_ORCAMENTO (ID, ID_ORCAMENTO, ID_PRODUTO, QTDE) "
                               f"values (GEN_ID(GEN_PRODUTO_ORCAMENTO_ID,1), {ultimo_req}, {id_prod}, "
                               f"'{qtde}');")

            conecta_robo.commit()
            print(f"Nº Orçemento {ultimo_req} inserido no banco com sucesso!")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def gera_excel_orcamento(self, num_sol, dados_itens):
        try:
            obs_solicitacao = "ORÇAMENTO GERADO AUTOMATICAMENTE"

            data_hoje = date.today()
            data_certa = data_hoje.strftime("%d/%m/%Y")

            d_um = []

            for i in dados_itens:
                cod, descr, ref, um, conjunto, num_tipo, tipo, qtde, ops_destino, vendas_destino = i

                msg = ""

                if ops_destino:
                    for opss in ops_destino:
                        num_op, cod_op, descr_op = opss

                        if msg:
                            msg += f"\n- OP {num_op} - {cod_op} - {descr_op}"
                        else:
                            msg += f"- OP {num_op} - {cod_op} - {descr_op}"

                if vendas_destino:
                    for vendass in vendas_destino:
                        num_pi, num_ov, cliente = vendass

                        if msg:
                            if num_pi:
                                numero = f"\nPI {num_pi}"
                            else:
                                numero = f"\nPI {num_ov}"
                        else:
                            if num_pi:
                                numero = f"PI {num_pi}"
                            else:
                                numero = f"PI {num_ov}"

                        msg += f"- {numero} - {cliente}\n"

                qtde_float = valores_para_float(qtde)

                dados = (cod, descr, ref, um,qtde_float, msg)
                d_um.append(dados)

            df = pd.DataFrame(d_um, columns=['Código', 'Descrição', 'Referência', 'UM', 'Qtde', 'Destino'])

            codigo_int = {'Código': int}
            df = df.astype(codigo_int)
            qtde_float = {'Qtde': float}
            df = df.astype(qtde_float)

            caminho_arquivo_modelo= f'Mod_orcamento.xlsx'
            nome_arquivo = f'Orçamento {num_sol}.xlsx'

            book = load_workbook(caminho_arquivo_modelo)

            writer = pd.ExcelWriter(nome_arquivo, engine='openpyxl')

            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

            linhas_frame = df.shape[0]
            colunas_frame = df.shape[1]

            linhas_certas = linhas_frame + 2 + 9
            colunas_certas = colunas_frame + 1

            ws = book.active

            inicia = 11
            rows = range(inicia, inicia + linhas_frame)
            columns = range(1, colunas_certas)

            ws.row_dimensions[linhas_certas + 2].height = 30
            ws.row_dimensions[linhas_certas + 4].height = 30

            for row in rows:
                for col in columns:
                    ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center',
                                                            wrap_text=True)
                    ws.cell(row, col).border = Border(left=Side(border_style='thin', color='00000000'),
                                                      right=Side(border_style='thin', color='00000000'),
                                                      top=Side(border_style='thin', color='00000000'),
                                                      bottom=Side(border_style='thin', color='00000000'),
                                                      diagonal=Side(border_style='thick', color='00000000'),
                                                      diagonal_direction=0,
                                                      outline=Side(border_style='thin', color='00000000'),
                                                      vertical=Side(border_style='thin', color='00000000'),
                                                      horizontal=Side(border_style='thin', color='00000000'))

            ws.merge_cells(f'A8:D8')
            top_left_cell = ws[f'A8']
            c = ws[f'A8']
            c.alignment = Alignment(horizontal='center',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=14, bold=True)
            top_left_cell.value = 'Orçamento Nº  ' + num_sol

            ws.merge_cells(f'E8:F8')
            top_left_cell = ws[f'E8']
            c = ws[f'E8']
            c.alignment = Alignment(horizontal='center',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=14, bold=True)
            top_left_cell.value = 'Emissão:  ' + data_certa

            ws.merge_cells(f'B{linhas_certas + 2}:B{linhas_certas + 2}')
            top_left_cell = ws[f'B{linhas_certas + 2}']
            c = ws[f'B{linhas_certas + 2}']
            c.alignment = Alignment(horizontal='right',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=True)
            top_left_cell.value = "Observação:  "

            ws.merge_cells(f'C{linhas_certas + 2}:F{linhas_certas + 2}')
            top_left_cell = ws[f'C{linhas_certas + 2}']
            c = ws[f'C{linhas_certas + 2}']
            c.alignment = Alignment(horizontal='left',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=False)
            top_left_cell.value = obs_solicitacao

            df.to_excel(writer, 'Sheet1', startrow=10, startcol=0, header=False, index=False)

            writer.save()

            print("Excel Salvo")

            return nome_arquivo

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def envia_email_sem_tipo(self, dados):
        try:
            saudacao, msg_final, to = self.dados_email()

            subject = f'SEM TIPO PLANO PCP - PRODUTOS SEM TIPO DEFINIDOS (PI)!'

            msg = MIMEMultipart()
            msg['From'] = email_user
            msg['Subject'] = subject

            body = f"{saudacao}\n\nAbaixo a lista de produtos sem tipo definido:\n\n"

            for i in dados:
                body += f"- {i}\n\n"

            body += f"\n{msg_final}"

            msg.attach(MIMEText(body, 'plain'))

            text = msg.as_string()
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(email_user, password)

            server.sendmail(email_user, to, text)
            server.quit()

            print("email enviado")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excluir_arquivo(self, caminho_arquivo):
        try:
            if os.path.exists(caminho_arquivo):
                os.remove(caminho_arquivo)
            else:
                print("O arquivo não existe no caminho especificado.")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def criar_op(self, num_op, cod_prod, qtde_produto, emissao):
        try:
            situacao = False

            previsao = emissao + timedelta(weeks=4)

            cod_barras = "SUZ000" + str(num_op)

            cur = conecta.cursor()
            cur.execute(f"SELECT id, descricao, COALESCE(obs, '') as obs, unidade, id_versao "
                        f"FROM produto where codigo = {cod_prod};")
            detalhes_produto = cur.fetchall()
            id_prod, descricao_id, referencia_id, unidade_id, id_versao = detalhes_produto[0]

            id_prod_int = int(id_prod)

            if id_versao:
                obs_certo = "OP CRIADA PELO SERVIDOR"

                cursor = conecta.cursor()
                cursor.execute(f"Insert into ordemservico "
                               f"(id, produto, numero, quantidade, datainicial, obs, codbarras, status, codigo, "
                               f"dataprevisao, id_estrutura, etapa) "
                               f"values (GEN_ID(GEN_ORDEMSERVICO_ID,1), {id_prod_int}, {num_op}, "
                               f"'{qtde_produto}', '{emissao}', '{obs_certo}', '{cod_barras}', 'A', "
                               f"'{cod_prod}', '{previsao}', {id_versao}, 'ABERTA');")

                conecta.commit()

                situacao = True

                print(f'A Ordem de Produção Nº {num_op} foi criado com sucesso!')

            return situacao

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)
            return None

    def cria_pdf_envia_email_usinagem(self, num_op, ref, qtde, emissao_br):
        try:
            s = re.sub(r"[^\d.]", "", ref)
            s = re.sub(r"\.+$", "", s)

            caminho_original = rf"\\Publico\C\OP\Projetos\{s}.pdf"
            arquivo_pdf = f"{s}"

            arquivo_imagem = self.cria_imagem_do_desenho_usinagem(num_op, qtde, emissao_br, caminho_original, arquivo_pdf)

            diretorio_destino = r'\\publico\C\OP\Usinagem para Corte/'
            arquivo_pdf_final = f'OP {num_op} - {arquivo_pdf}.pdf'
            caminho_completo = os.path.join(diretorio_destino, arquivo_pdf_final)

            self.converte_png_para_pdf(arquivo_imagem, caminho_completo)

            self.excluir_arquivo(arquivo_imagem)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def cria_imagem_do_desenho_usinagem(self, num_op, qtde_produto, emissao_br, caminho_original, num_desenho_arq):
        try:
            num_op_str = f"OP {num_op}"
            if qtde_produto > 1:
                qtde_ordem = f"{qtde_produto} PÇS"
            else:
                qtde_ordem = f"{qtde_produto} PÇ"

            data_emissao = f"Emissão: {emissao_br}"

            images = convert_from_path(caminho_original, 500, poppler_path=self.caminho_poppler)

            imgs = images[0]

            draw = ImageDraw.Draw(imgs)
            font = ImageFont.truetype("tahoma.ttf", 150)
            font1 = ImageFont.truetype("tahoma.ttf", 70)

            def criar_texto(pos_horizontal, pos_vertical, texto, cor, fonte, largura_tra):
                draw.text((pos_horizontal, pos_vertical), texto, fill=cor, font=fonte, stroke_width=largura_tra)

            criar_texto(4500, 2900, num_op_str, (0, 0, 0), font, 4)
            criar_texto(5160, 2900, qtde_ordem, (0, 0, 0), font, 4)
            criar_texto(5000, 3150, data_emissao, (0, 0, 0), font1, 0)

            arquivo_final = f"{num_desenho_arq}.png"
            imgs.save(arquivo_final)

            time.sleep(1)

            return arquivo_final

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)
            return None

    def cria_pdf_envia_email_conjunto(self, num_op, ref, qtde, emissao_br):
        try:
            s = re.sub(r"[^\d.]", "", ref)
            s = re.sub(r"\.+$", "", s)

            caminho_original = rf"\\Publico\C\OP\Projetos\{s}.pdf"
            arquivo_pdf = f"{s}"

            arquivo_imagem = self.cria_imagem_do_desenho_conjunto(num_op, qtde, emissao_br, caminho_original, arquivo_pdf)

            diretorio_destino = r'\\publico\C\OP\Aguardando Material/'
            arquivo_pdf_final = f'OP {num_op} - {arquivo_pdf}.pdf'
            caminho_completo = os.path.join(diretorio_destino, arquivo_pdf_final)

            self.converte_png_para_pdf(arquivo_imagem, caminho_completo)
            self.excluir_arquivo(arquivo_imagem)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def cria_imagem_do_desenho_conjunto(self, num_op, qtde_produto, emissao_br, caminho_original, num_desenho_arq):
        try:
            num_op_str = f"OP {num_op}"
            if qtde_produto > 1:
                qtde_ordem = f"{qtde_produto} PÇS"
            else:
                qtde_ordem = f"{qtde_produto} PÇ"

            data_emissao = f"Emissão: {emissao_br}"

            images = convert_from_path(caminho_original, 500, poppler_path=self.caminho_poppler)

            imgs = images[0]

            draw = ImageDraw.Draw(imgs)
            font = ImageFont.truetype("tahoma.ttf", 150)
            font1 = ImageFont.truetype("tahoma.ttf", 70)

            def criar_texto(pos_horizontal, pos_vertical, texto, cor, fonte, largura_tra):
                draw.text((pos_horizontal, pos_vertical), texto, fill=cor, font=fonte, stroke_width=largura_tra)

            criar_texto(4500, 2830, num_op_str, (0, 0, 0), font, 4)
            criar_texto(5160, 2830, qtde_ordem, (0, 0, 0), font, 4)
            criar_texto(5000, 3080, data_emissao, (0, 0, 0), font1, 0)

            arquivo_final = f"{num_desenho_arq}.png"
            imgs.save(arquivo_final)

            time.sleep(1)

            return arquivo_final

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

            return None

    def converte_png_para_pdf(self, input_png, output_pdf):
        try:
            c = canvas.Canvas(output_pdf, pagesize=landscape(A4))
            img = ImageReader(input_png)
            width, height = img.getSize()

            aspect_ratio = width / height
            target_width = 800
            target_height = target_width / aspect_ratio

            x = (A4[1] - target_width) / 2
            y = (A4[0] - target_height) / 2

            c.drawImage(img, x, y, width=target_width, height=target_height)
            c.showPage()
            c.save()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_situacao_produto(self, cod, ref, saldo, num_conj, qtde_necessidade):
        try:
            produto_sem_estrutura = False
            acabado_sem_desenho = False
            prod_duplicado = False

            saldo_float = valores_para_float(saldo)
            qtde_necessidade -= saldo_float

            qtde_compra = self.manipula_dados_tabela_compra(cod)
            qtde_necessidade -= qtde_compra
            qtde_necessidade = round(qtde_necessidade, 3)

            if num_conj == 10:
                s = re.sub(r"[^\d.]", "", ref)
                s = re.sub(r"\.+$", "", s)

                cursor = conecta.cursor()
                cursor.execute(
                    f"SELECT DISTINCT codigo, descricao, COALESCE(obs, ''), unidade, COALESCE(tipomaterial, ''), "
                    f"COALESCE(localizacao, ''), id_versao "
                    f"FROM produto "
                    f"WHERE obs = '{ref}';")
                detalhes_produto = cursor.fetchall()

                if detalhes_produto:
                    qtde_itens = len(detalhes_produto)
                    if qtde_itens > 1:
                        prod_duplicado = True

                caminho_pdf = rf"\\Publico\C\OP\Projetos\{s}.pdf"

                if not os.path.exists(caminho_pdf):
                    acabado_sem_desenho = True

                estrutura = self.manipula_dados_tabela_estrutura(cod)

                if estrutura:
                    qtde_producao = self.manipula_dados_tabela_producao(cod)
                    qtde_necessidade -= qtde_producao
                    qtde_necessidade = round(qtde_necessidade, 3)

                else:
                    produto_sem_estrutura = True

            qtde_necessidade_ops = self.manipula_dados_tabela_consumo(cod)
            qtde_necessidade += qtde_necessidade_ops
            qtde_necessidade = round(qtde_necessidade, 3)

            return qtde_necessidade, produto_sem_estrutura, acabado_sem_desenho, prod_duplicado


        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_dados_tabela_estrutura(self, cod_prod):
        try:
            cursor = conecta.cursor()
            cursor.execute(f"SELECT id, codigo, id_versao FROM produto where codigo = {cod_prod};")
            select_prod = cursor.fetchall()
            idez, cod, id_estrut = select_prod[0]

            if id_estrut:
                cursor = conecta.cursor()
                cursor.execute(f"SELECT prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, "
                               f"prod.unidade, (estprod.quantidade * 1) as qtde, prod.quantidade, prod.conjunto, "
                               f"tip.tipomaterial, prod.tipomaterial "
                               f"from estrutura_produto as estprod "
                               f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                               f"INNER JOIN conjuntos conj ON prod.conjunto = conj.id "
                               f"LEFT JOIN tipomaterial tip ON prod.tipomaterial = tip.id "
                               f"where estprod.id_estrutura = {id_estrut} "
                               f"order by conj.conjunto DESC, prod.descricao ASC;")
                tabela_estrutura = cursor.fetchall()

                return tabela_estrutura

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_dados_tabela_producao(self, cod_prod):
        try:
            op_ab_editado = 0

            cursor = conecta.cursor()
            cursor.execute(f"select ordser.datainicial, ordser.dataprevisao, ordser.numero, prod.codigo, "
                           f"prod.descricao, "
                           f"COALESCE(prod.obs, '') as obs, prod.unidade, "
                           f"ordser.quantidade, ordser.id_estrutura "
                           f"from ordemservico as ordser "
                           f"INNER JOIN produto prod ON ordser.produto = prod.id "
                           f"where ordser.status = 'A' and prod.codigo = {cod_prod} "
                           f"order by ordser.numero;")
            op_abertas = cursor.fetchall()
            if op_abertas:
                for dados_op in op_abertas:
                    emissao, previsao, op, cod, descr, ref, um, qtde, id_estrut = dados_op

                    qtde_float = valores_para_float(qtde)

                    op_ab_editado += qtde_float

            return op_ab_editado

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_dados_tabela_consumo(self, cod_prod):
        try:
            qtde_necessidade = 0

            cursor = conecta.cursor()
            cursor.execute(f"SELECT prod.id, estprod.id, estprod.id_estrutura "
                           f"from estrutura_produto as estprod "
                           f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                           f"where prod.codigo = {cod_prod};")
            dados_estrut = cursor.fetchall()
            for i_estrut in dados_estrut:
                prod_id, ides_mat, id_estrutura = i_estrut

                cursor = conecta.cursor()
                cursor.execute(f"select id, id_produto, num_versao, data_versao, obs, data_criacao "
                               f"from estrutura where id = {id_estrutura};")
                estrutura = cursor.fetchall()

                id_produto = estrutura[0][1]

                cursor = conecta.cursor()
                cursor.execute(f"select ordser.datainicial, ordser.numero, prod.codigo, prod.descricao "
                               f"from ordemservico as ordser "
                               f"INNER JOIN produto prod ON ordser.produto = prod.id "
                               f"where ordser.status = 'A' "
                               f"and prod.id = {id_produto} and prod.id_versao = {id_estrutura} "
                               f"order by ordser.numero;")
                op_abertas = cursor.fetchall()
                if op_abertas:
                    for ii in op_abertas:
                        emissao, num_op, cod_pai, descr_pai = ii

                        cursor = conecta.cursor()
                        cursor.execute(f"SELECT estprod.id, prod.codigo, "
                                       f"((SELECT quantidade FROM ordemservico where numero = {num_op}) * "
                                       f"(estprod.quantidade)) AS Qtde "
                                       f"FROM estrutura_produto as estprod "
                                       f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                                       f"where estprod.id = {ides_mat};")
                        select_estrut = cursor.fetchall()
                        if select_estrut:
                            id_mat, cod_estrut, qtde_total = select_estrut[0]

                            total_float = valores_para_float(qtde_total)

                            cursor = conecta.cursor()
                            cursor.execute(f"SELECT max(prod.codigo), max(prod.descricao), "
                                           f"sum(prodser.QTDE_ESTRUT_PROD) as total "
                                           f"FROM estrutura_produto as estprod "
                                           f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                                           f"INNER JOIN produtoos as prodser ON estprod.id = prodser.ID_ESTRUT_PROD "
                                           f"where estprod.id_estrutura = {id_estrutura} "
                                           f"and prodser.numero = {num_op} and estprod.id = {id_mat} "
                                           f"group by prodser.ID_ESTRUT_PROD;")
                            select_os_resumo = cursor.fetchall()
                            if select_os_resumo:
                                for os_cons in select_os_resumo:
                                    cod_cons, descr_cons, qtde_cons_total = os_cons

                                    cons_float = valores_para_float(qtde_cons_total)

                                    qtde_necessidade += total_float - cons_float
                            else:
                                qtde_necessidade += total_float

            if qtde_necessidade:
                arred = round(qtde_necessidade, 3)
            else:
                arred = 0

            return arred

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_dados_tabela_consumo_final(self, cod_prod):
        try:
            tabela_nova = []

            cursor = conecta.cursor()
            cursor.execute(f"SELECT prod.id, estprod.id, estprod.id_estrutura "
                           f"from estrutura_produto as estprod "
                           f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                           f"where prod.codigo = {cod_prod};")
            dados_estrut = cursor.fetchall()
            for i_estrut in dados_estrut:
                prod_id, ides_mat, id_estrutura = i_estrut

                cursor = conecta.cursor()
                cursor.execute(f"select id, id_produto, num_versao, data_versao, obs, data_criacao "
                               f"from estrutura where id = {id_estrutura};")
                estrutura = cursor.fetchall()

                id_produto = estrutura[0][1]

                cursor = conecta.cursor()
                cursor.execute(f"select ordser.datainicial, ordser.numero, prod.codigo, prod.descricao "
                               f"from ordemservico as ordser "
                               f"INNER JOIN produto prod ON ordser.produto = prod.id "
                               f"where ordser.status = 'A' "
                               f"and prod.id = {id_produto} and prod.id_versao = {id_estrutura} "
                               f"order by ordser.numero;")
                op_abertas = cursor.fetchall()
                if op_abertas:
                    for ii in op_abertas:
                        emissao, num_op, cod_pai, descr_pai = ii

                        cursor = conecta.cursor()
                        cursor.execute(f"SELECT estprod.id, prod.codigo, "
                                       f"((SELECT quantidade FROM ordemservico where numero = {num_op}) * "
                                       f"(estprod.quantidade)) AS Qtde "
                                       f"FROM estrutura_produto as estprod "
                                       f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                                       f"where estprod.id = {ides_mat};")
                        select_estrut = cursor.fetchall()
                        if select_estrut:
                            id_mat, cod_estrut, qtde_total = select_estrut[0]

                            cursor = conecta.cursor()
                            cursor.execute(f"SELECT max(prod.codigo), max(prod.descricao), "
                                           f"sum(prodser.QTDE_ESTRUT_PROD) as total "
                                           f"FROM estrutura_produto as estprod "
                                           f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                                           f"INNER JOIN produtoos as prodser ON estprod.id = prodser.ID_ESTRUT_PROD "
                                           f"where estprod.id_estrutura = {id_estrutura} "
                                           f"and prodser.numero = {num_op} and estprod.id = {id_mat} "
                                           f"group by prodser.ID_ESTRUT_PROD;")
                            select_os_resumo = cursor.fetchall()
                            if select_os_resumo:
                                for os_cons in select_os_resumo:
                                    cod_cons, descr_cons, qtde_cons_total = os_cons

                                    if qtde_total != qtde_cons_total:
                                        dados = (num_op, cod_pai, descr_pai)
                                        tabela_nova.append(dados)
                            else:
                                dados = (num_op, cod_pai, descr_pai)
                                tabela_nova.append(dados)

            return tabela_nova

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_dados_tabela_venda_final(self, cod_prod):
        try:
            tabela_nova = []

            cursor = conecta.cursor()
            cursor.execute(f"SELECT ped.emissao, ped.id, cli.razao, prodint.qtde, "
                           f"prodint.data_previsao "
                           f"FROM PRODUTOPEDIDOINTERNO as prodint "
                           f"INNER JOIN produto as prod ON prodint.id_produto = prod.id "
                           f"INNER JOIN pedidointerno as ped ON prodint.id_pedidointerno = ped.id "
                           f"INNER JOIN clientes as cli ON ped.id_cliente = cli.id "
                           f"where prodint.status = 'A' and prod.codigo = {cod_prod};")
            dados_pi = cursor.fetchall()

            if dados_pi:
                for i_pi in dados_pi:
                    emissao_pi, num_pi, clie_pi, qtde_pi, entrega_pi = i_pi

                    dados_pi = (num_pi, "", clie_pi)
                    tabela_nova.append(dados_pi)

            cursor = conecta.cursor()
            cursor.execute(f"SELECT oc.data, oc.numero, cli.razao, prodoc.quantidade, prodoc.dataentrega, "
                           f"COALESCE(prodoc.id_pedido, ''), COALESCE(prodoc.id_expedicao, '') "
                           f"FROM PRODUTOORDEMCOMPRA as prodoc "
                           f"INNER JOIN produto as prod ON prodoc.produto = prod.id "
                           f"INNER JOIN ordemcompra as oc ON prodoc.mestre = oc.id "
                           f"INNER JOIN clientes as cli ON oc.cliente = cli.id "
                           f"LEFT JOIN pedidointerno as ped ON prodoc.id_pedido = ped.id "
                           f"where prodoc.quantidade > prodoc.produzido "
                           f"and oc.status = 'A' "
                           f"and oc.entradasaida = 'S' "
                           f"and prod.codigo = {cod_prod};")
            dados_ov = cursor.fetchall()
            if dados_ov:
                for i_ov in dados_ov:
                    emissao_ov, num_ov, clie_ov, qtde_ov, entrega_ov, num_pi_ov, num_exp = i_ov

                    dados = (num_pi_ov, num_ov, clie_ov)
                    tabela_nova.append(dados)

            return tabela_nova

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_dados_tabela_compra(self, cod_prod):
        try:
            qtde_compras = 0

            cursor = conecta.cursor()
            cursor.execute(f"SELECT COALESCE(prodreq.mestre, ''), req.dataemissao, prodreq.quantidade "
                           f"FROM produtoordemsolicitacao as prodreq "
                           f"INNER JOIN produto as prod ON prodreq.produto = prod.ID "
                           f"INNER JOIN ordemsolicitacao as req ON prodreq.mestre = req.idsolicitacao "
                           f"LEFT JOIN produtoordemrequisicao as preq ON prodreq.id = preq.id_prod_sol "
                           f"WHERE prodreq.status = 'A' "
                           f"and prod.codigo = {cod_prod} "
                           f"AND preq.id_prod_sol IS NULL "
                           f"ORDER BY prodreq.mestre;")
            dados_sol = cursor.fetchall()

            if dados_sol:
                for i_sol in dados_sol:
                    num_sol, emissao_sol, qtde_sol = i_sol

                    qtde_sol_float = valores_para_float(qtde_sol)

                    qtde_compras += qtde_sol_float

            cursor = conecta.cursor()
            cursor.execute(f"SELECT sol.idsolicitacao, prodreq.quantidade, req.data, prodreq.numero, "
                           f"prodreq.destino, prodreq.id_prod_sol "
                           f"FROM produtoordemrequisicao as prodreq "
                           f"INNER JOIN produto as prod ON prodreq.produto = prod.ID "
                           f"INNER JOIN ordemrequisicao as req ON prodreq.mestre = req.id "
                           f"LEFT JOIN produtoordemsolicitacao as prodsol ON prodreq.id_prod_sol = prodsol.id "
                           f"LEFT JOIN ordemsolicitacao as sol ON prodsol.mestre = sol.idsolicitacao "
                           f"where prodreq.status = 'A' "
                           f"and prod.codigo = {cod_prod};")
            dados_req = cursor.fetchall()

            if dados_req:
                for i_req in dados_req:
                    num_sol_req, qtde_req, emissao_req, num_req, destino, id_prod_sol = i_req

                    qtde_req_float = valores_para_float(qtde_req)

                    qtde_compras += qtde_req_float

            cursor = conecta.cursor()
            cursor.execute(
                f"SELECT sol.idsolicitacao, prodreq.numero, oc.data, oc.numero, forn.razao, "
                f"prodoc.quantidade, prodoc.produzido, prodoc.dataentrega "
                f"FROM ordemcompra as oc "
                f"INNER JOIN produtoordemcompra as prodoc ON oc.id = prodoc.mestre "
                f"LEFT JOIN produtoordemrequisicao as prodreq ON prodoc.id_prod_req = prodreq.id "
                f"INNER JOIN produto as prod ON prodoc.produto = prod.id "
                f"INNER JOIN fornecedores as forn ON oc.fornecedor = forn.id "
                f"LEFT JOIN produtoordemsolicitacao as prodsol ON prodreq.id_prod_sol = prodsol.id "
                f"LEFT JOIN ordemsolicitacao as sol ON prodsol.mestre = sol.idsolicitacao "
                f"where oc.entradasaida = 'E' "
                f"AND oc.STATUS = 'A' "
                f"AND prodoc.produzido < prodoc.quantidade "
                f"and prod.codigo = {cod_prod}"
                f"ORDER BY oc.numero;")
            dados_oc = cursor.fetchall()

            if dados_oc:
                for i_oc in dados_oc:
                    num_sol_oc, id_req_oc, emissao_oc, num_oc, forncec_oc, qtde_oc, prod_oc, entrega_oc = i_oc

                    qtde_oc_float = valores_para_float(qtde_oc)

                    qtde_compras += qtde_oc_float

            return qtde_compras

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def envia_email_sem_estrutura(self, caminho, arquivo, produto):
        try:
            saudacao, msg_final, to = self.dados_email()

            subject = f'PLANO PCP - Produto não possui estrutura!'

            msg = MIMEMultipart()
            msg['From'] = email_user
            msg['Subject'] = subject

            body = f"{saudacao}\n\n" \
                   f"Segue lista dos itens encontrados:\n\n"

            for i in produto:
                body += f"{i}.\n"

            body += f"\n\n{msg_final}"

            msg.attach(MIMEText(body, 'plain'))

            attachment = open(caminho, 'rb')

            part = MIMEBase('application', "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment',
                            filename=Header(arquivo, 'utf-8').encode())
            msg.attach(part)

            text = msg.as_string()
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(email_user, password)

            server.sendmail(email_user, to, text)
            attachment.close()

            server.quit()

            print(f'produto sem estrutura enviada com sucesso!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def envia_email_nao_acha_desenho(self, arquivo_pdf, cod):
        try:
            saudacao, msg_final, to = self.dados_email()

            subject = f'PLANO PCP - Não foi encontrado o desenho {arquivo_pdf}'

            msg = MIMEMultipart()
            msg['From'] = email_user
            msg['Subject'] = subject

            print(arquivo_pdf)

            body = f"{saudacao}\n\n" \
                   f"O desenho {arquivo_pdf} de código {cod} não foi encontrado no cadastro dos produtos.\n\n" \
                   f"{msg_final}"

            msg.attach(MIMEText(body, 'plain'))

            text = msg.as_string()
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(email_user, password)

            server.sendmail(email_user, to, text)
            server.quit()

            print("email enviado sem arquivo pdf desenho")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def envia_email_desenho_duplicado(self, cod, ref):
        try:
            saudacao, msg_final, to = self.dados_email()

            subject = f'PLANO PCP - Foi encontrado produtos com desenho duplicado no cadastro!'

            msg = MIMEMultipart()
            msg['From'] = email_user
            msg['Subject'] = subject

            body = f"{saudacao}\n\n" \
                   f"Segue lista dos itens encontrados:\n\n"

            body += f" - {cod} - {ref}.\n"

            body += f"\n\n{msg_final}"

            msg.attach(MIMEText(body, 'plain'))

            text = msg.as_string()
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(email_user, password)

            server.sendmail(email_user, to, text)

            server.quit()

            print(f'produto desenho duplicado enviada com sucesso!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)


chama_classe = ExecutaPlanoPcp()
