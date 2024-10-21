import sys
from banco_dados.controle_erros import grava_erro_banco
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl import Workbook, load_workbook, drawing
from openpyxl.utils import get_column_letter
import pandas as pd
import os
import inspect
import traceback

nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
nome_arquivo = os.path.basename(nome_arquivo_com_caminho)


def trata_excecao(nome_funcao, mensagem, arquivo, excecao):
    try:
        tb = traceback.extract_tb(excecao)
        num_linha_erro = tb[-1][1]

        traceback.print_exc()
        print(f'Houve um problema no arquivo: {arquivo} na função: "{nome_funcao}"\n{mensagem} {num_linha_erro}')

        grava_erro_banco(nome_funcao, mensagem, arquivo, num_linha_erro)

    except Exception as e:
        nome_funcao_trat = inspect.currentframe().f_code.co_name
        exc_traceback = sys.exc_info()[2]
        tb = traceback.extract_tb(exc_traceback)
        num_linha_erro = tb[-1][1]
        print(f'Houve um problema no arquivo: {nome_arquivo} na função: "{nome_funcao_trat}"\n'
              f'{e} {num_linha_erro}')
        grava_erro_banco(nome_funcao_trat, e, nome_arquivo, num_linha_erro)


def lanca_dados_mesclado(mp_copy, mesclado, celula, informacao, tam_fonte, negrito):
    try:
        mp_copy.merge_cells(mesclado)
        celula_sup_esq = mp_copy[celula]
        cel = mp_copy[celula]
        edita_alinhamento(cel)
        edita_fonte(cel, tamanho=tam_fonte, negrito=negrito)
        celula_sup_esq.value = informacao

    except Exception as e:
        nome_funcao = inspect.currentframe().f_code.co_name
        exc_traceback = sys.exc_info()[2]
        trata_excecao(nome_funcao, str(e), nome_arquivo, exc_traceback)


def lanca_dados_coluna(ws, celula, informacao, tam_fonte, negrito):
    try:
        celula_sup_esq = ws[celula]
        cel = ws[celula]
        edita_alinhamento(cel)
        edita_fonte(cel, tamanho=tam_fonte, negrito=negrito)
        celula_sup_esq.value = informacao

    except Exception as e:
        nome_funcao = inspect.currentframe().f_code.co_name
        exc_traceback = sys.exc_info()[2]
        trata_excecao(nome_funcao, str(e), nome_arquivo, exc_traceback)


def linhas_colunas_p_edicao(sheet, min_linha, max_linha, min_coluna, max_coluna):
    try:
        for row in sheet.iter_rows(min_row=min_linha,
                                   max_row=max_linha,
                                   min_col=min_coluna,
                                   max_col=max_coluna):
            for cell in row:
                yield cell

    except Exception as e:
        nome_funcao = inspect.currentframe().f_code.co_name
        exc_traceback = sys.exc_info()[2]
        trata_excecao(nome_funcao, str(e), nome_arquivo, exc_traceback)


def edita_alinhamento(cell, ali_horizontal='center', ali_vertical='center', rotacao=0, quebra_linha=False,
                      encolher=False, recuar=0):
    try:
        cell.alignment = Alignment(horizontal=ali_horizontal,
                                   vertical=ali_vertical,
                                   text_rotation=rotacao,
                                   wrap_text=quebra_linha,
                                   shrink_to_fit=encolher,
                                   indent=recuar)

    except Exception as e:
        nome_funcao = inspect.currentframe().f_code.co_name
        exc_traceback = sys.exc_info()[2]
        trata_excecao(nome_funcao, str(e), nome_arquivo, exc_traceback)


def edita_bordas(cell):
    try:
        cell.border = Border(left=Side(border_style='thin', color='00000000'),
                             right=Side(border_style='thin', color='00000000'),
                             top=Side(border_style='thin', color='00000000'),
                             bottom=Side(border_style='thin', color='00000000'),
                             diagonal=Side(border_style='thick', color='00000000'),
                             diagonal_direction=0,
                             outline=Side(border_style='thin', color='00000000'),
                             vertical=Side(border_style='thin', color='00000000'),
                             horizontal=Side(border_style='thin', color='00000000'))

    except Exception as e:
        nome_funcao = inspect.currentframe().f_code.co_name
        exc_traceback = sys.exc_info()[2]
        trata_excecao(nome_funcao, str(e), nome_arquivo, exc_traceback)


def edita_preenchimento(cell):
    try:
        cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

    except Exception as e:
        nome_funcao = inspect.currentframe().f_code.co_name
        exc_traceback = sys.exc_info()[2]
        trata_excecao(nome_funcao, str(e), nome_arquivo, exc_traceback)


def edita_fonte(cell, nome_fonte='Calibri', tamanho=10, negrito=False):
    try:
        cell.font = Font(name=nome_fonte, size=tamanho, bold=negrito)

    except Exception as e:
        nome_funcao = inspect.currentframe().f_code.co_name
        exc_traceback = sys.exc_info()[2]
        trata_excecao(nome_funcao, str(e), nome_arquivo, exc_traceback)


def carregar_workbook(caminho):
    try:
        book = load_workbook(caminho)

        return book

    except Exception as e:
        nome_funcao = inspect.currentframe().f_code.co_name
        exc_traceback = sys.exc_info()[2]
        trata_excecao(nome_funcao, str(e), nome_arquivo, exc_traceback)


def letra_coluna(coluna):
    try:
        column_letter = get_column_letter(coluna)

        return column_letter

    except Exception as e:
        nome_funcao = inspect.currentframe().f_code.co_name
        exc_traceback = sys.exc_info()[2]
        trata_excecao(nome_funcao, str(e), nome_arquivo, exc_traceback)


def ajusta_larg_coluna(planilha, coluna, largura_ajustada):
    try:
        planilha.column_dimensions[get_column_letter(coluna[0].column)].width = largura_ajustada

    except Exception as e:
        nome_funcao = inspect.currentframe().f_code.co_name
        exc_traceback = sys.exc_info()[2]
        trata_excecao(nome_funcao, str(e), nome_arquivo, exc_traceback)


def criar_workbook():
    try:
        book = Workbook()

        return book

    except Exception as e:
        nome_funcao = inspect.currentframe().f_code.co_name
        exc_traceback = sys.exc_info()[2]
        trata_excecao(nome_funcao, str(e), nome_arquivo, exc_traceback)


def adiciona_imagem(ws, caminho_arquivo, celula):
    try:
        img = drawing.image.Image(caminho_arquivo)
        ws.add_image(img, celula)

    except Exception as e:
        nome_funcao = inspect.currentframe().f_code.co_name
        exc_traceback = sys.exc_info()[2]
        trata_excecao(nome_funcao, str(e), nome_arquivo, exc_traceback)


def dataframe_pandas(dados_lista, colunas):
    try:
        df = pd.DataFrame(dados_lista, columns=colunas)

        return df

    except Exception as e:
        nome_funcao = inspect.currentframe().f_code.co_name
        exc_traceback = sys.exc_info()[2]
        trata_excecao(nome_funcao, str(e), nome_arquivo, exc_traceback)


def escritor_dataframe(caminho):
    try:
        writer = pd.ExcelWriter(caminho, engine='openpyxl')

        return writer

    except Exception as e:
        nome_funcao = inspect.currentframe().f_code.co_name
        exc_traceback = sys.exc_info()[2]
        trata_excecao(nome_funcao, str(e), nome_arquivo, exc_traceback)


def escritor_direto_dataframe(df, writer, nome_planilha, partida, ini_coluna, cabecalho, indice):
    try:
        df.to_excel(writer,
                    sheet_name=nome_planilha,
                    startrow=partida,
                    startcol=ini_coluna,
                    header=cabecalho,
                    index=indice)

    except Exception as e:
        nome_funcao = inspect.currentframe().f_code.co_name
        exc_traceback = sys.exc_info()[2]
        trata_excecao(nome_funcao, str(e), nome_arquivo, exc_traceback)
