import sys
from banco_dados.conexao import conecta, conecta_robo
from banco_dados.controle_erros import grava_erro_banco
from comandos.conversores import valores_para_float
import os
import inspect
import smtplib
import traceback
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.header import Header
from email import encoders
from datetime import datetime, date
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from dados_email import email_user, password
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Side, Alignment, Border, Font

from collections import defaultdict


class EnviaIndustrializacao:
    def __init__(self):
        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

        self.manipula_comeco()
        
    def trata_excecao(self, nome_funcao, mensagem, arquivo, excecao):
        try:
            tb = traceback.extract_tb(excecao)
            num_linha_erro = tb[-1][1]

            traceback.print_exc()
            print(f'Houve um problema no arquivo: {arquivo} na função: "{nome_funcao}"\n{mensagem} {num_linha_erro}')

            grava_erro_banco(nome_funcao, mensagem, arquivo, num_linha_erro)

        except Exception as e:
            nome_funcao_trat = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            tb = traceback.extract_tb(exc_traceback)
            num_linha_erro = tb[-1][1]
            print(f'Houve um problema no arquivo: {self.nome_arquivo} na função: "{nome_funcao_trat}"\n'
                  f'{e} {num_linha_erro}')
            grava_erro_banco(nome_funcao_trat, e, self.nome_arquivo, num_linha_erro)

    def dados_email(self):
        try:
            to = ['<maquinas@unisold.com.br>']

            current_time = (datetime.now())
            horario = current_time.strftime('%H')
            hora_int = int(horario)
            saudacao = ""
            if 4 < hora_int < 13:
                saudacao = "Bom Dia!"
            elif 12 < hora_int < 19:
                saudacao = "Boa Tarde!"
            elif hora_int > 18:
                saudacao = "Boa Noite!"
            elif hora_int < 5:
                saudacao = "Boa Noite!"

            msg_final = f"Att,\n" \
                        f"Suzuki Máquinas Ltda\n" \
                        f"Fone (51) 3561.2583/(51) 3170.0965\n\n" \
                        f"Mensagem enviada automaticamente, por favor não responda.\n\n" \
                        f"Se houver algum problema com o recebimento de emails ou conflitos com o arquivo excel, " \
                        f"favor entrar em contato pelo email maquinas@unisold.com.br.\n\n"

            return saudacao, msg_final, to

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def adicionar_tabelas_listagem(self, dados, cabecalho):
        try:
            elements = []

            style_lista = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.gray),
                                      ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                                      ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                      ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                                      ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                                      ('GRID', (0, 0), (-1, -1), 1, colors.black),
                                      ('FONTSIZE', (0, 0), (-1, 0), 10),
                                      ('FONTSIZE', (0, 1), (-1, -1), 8)])

            table = Table([cabecalho] + dados)
            table.setStyle(style_lista)
            elements.append(table)

            return elements

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def gerar_pdf_listagem_separar(self, caminho_listagem, lista):
        try:
            lista_final = []

            for i in lista:
                (cod_pai, descr_pai, ref_pai, um_pai, cod_filho, descr_filho, ref_filho, um_filho,
                 local_filho, saldo_filho, ops_destino_pai, vendas_destino_pai, id_mov) = i
                dados = (cod_filho, descr_filho, ref_filho, um_filho, local_filho, saldo_filho)
                lista_final.append(dados)

            margem_esquerda = 0
            margem_direita = 5
            margem_superior = 25
            margem_inferior = 5

            doc = SimpleDocTemplate(caminho_listagem, pagesize=A4,
                                    leftMargin=margem_esquerda,
                                    rightMargin=margem_direita,
                                    topMargin=margem_superior,
                                    bottomMargin=margem_inferior)

            titulo = ['INDÚSTRIALIZAÇÃO']
            style_lista = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.gray),
                                      ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                                      ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                      ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                                      ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                                      ('GRID', (0, 0), (-1, -1), 1, colors.black),
                                      ('FONTSIZE', (0, 0), (-1, 0), 10),
                                      ('FONTSIZE', (0, 1), (-1, -1), 8)])

            table = Table([titulo])
            table.setStyle(style_lista)
            elements = [table]

            cabecalho_lista = ['CÓDIGO', 'DESCRIÇÃO', 'REFERÊNCIA', 'UM', 'LOCALIZAÇÃO', 'SALDO']
            elem_lista = self.adicionar_tabelas_listagem(lista_final, cabecalho_lista)

            cabecalho_transp = ['', 'TRANSPORTE']
            dados_transp = [('PESO LÍQUIDO', ''), ('PESO BRUTO', ''), ('VOLUME', '')]
            elem_transp = self.adicionar_tabelas_listagem(dados_transp, cabecalho_transp)

            cabecalho_medida = ['MEDIDAS', '            ']
            dados_medida = [('ALTURA (MM)', ''), ('LARGURA (MM)', ''), ('COMPRIMENTO (MM)', '')]
            elem_medida = self.adicionar_tabelas_listagem(dados_medida, cabecalho_medida)

            cabecalho_motorista = ['DAMDFE', 'MOTORISTA']
            dados_motorista = [('PLACA', ''), ('NOME', ''), ('CPF', '')]
            elem_motorista = self.adicionar_tabelas_listagem(dados_motorista, cabecalho_motorista)

            espaco_em_branco = Table([[None]], style=[('SIZE', (0, 0), (0, 0), 20)])

            # Criar tabela para colocar medidas e motorista lado a lado
            tabela_medida_motorista = Table([[elem_transp, elem_medida, elem_motorista]],
                                            colWidths=[170, 170])  # Ajuste as larguras conforme necessário

            elementos = (elements + [espaco_em_branco] + elem_lista + [espaco_em_branco] +
                         [tabela_medida_motorista])  # Adiciona a tabela com medidas e motorista lado a lado

            doc.build(elementos)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def inserir_no_banco(self, lista_banco):
        try:
            for i in lista_banco:
                (cod_pai, descr_pai, ref_pai, um_pai, cod_filho, descr_filho, ref_filho, um_filho,
                 local_filho, saldo_filho, ops_destino_pai, vendas_destino_pai, id_mov) = i

                cursor = conecta_robo.cursor()
                cursor.execute(f"Insert into ENVIA_INDUSTRIALIZACAO (ID, id_envia_mov, cod_prod) "
                               f"values (GEN_ID(GEN_ENVIA_INDUSTRIALIZACAO_ID,1), {id_mov}, {cod_filho});")

                conecta_robo.commit()

                print(f"Salvo no banco com sucesso!")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def envia_email(self, caminho, arquivo):
        try:
            saudacao, msg_final, to = self.dados_email()

            to = ['<maquinas@unisold.com.br>']

            subject = f'IND - Separar Produtos para industrialização'

            msg = MIMEMultipart()
            msg['From'] = email_user
            msg['Subject'] = subject

            body = f"{saudacao}\n\n" \
                   f"Alguns produtos estão prontos para enviar para industrialização.\n\n" \
                   f"{msg_final}"

            msg.attach(MIMEText(body, 'plain'))

            diretorio_atual = os.path.dirname(os.path.abspath(__file__))
            caminho_arquivo = os.path.join(diretorio_atual, caminho)
            attachment = open(caminho_arquivo, 'rb')

            part = MIMEBase('application', "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment',
                            filename=Header(arquivo, 'utf-8').encode())
            msg.attach(part)

            text = msg.as_string()
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(email_user, password)

            server.sendmail(email_user, to, text)
            attachment.close()

            server.quit()

            print(f'Email enviado com sucesso!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excluir_arquivo(self, caminho_arquivo):
        try:
            if os.path.exists(caminho_arquivo):
                os.remove(caminho_arquivo)
            else:
                print("O arquivo não existe no caminho especificado.")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_comeco(self):
        try:
            lista_produtos = []

            agrupado = defaultdict(list)

            cursor = conecta.cursor()
            cursor.execute("""
                SELECT id, data_mov 
                FROM envia_mov 
                WHERE data_mov >= DATEADD(-1 MONTH TO CURRENT_DATE)
            """)
            dados_mov = cursor.fetchall()

            if dados_mov:
                for i in dados_mov:
                    id_mov, data_mov = i

                    print(i)

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT prod.codigo, prod.descricao, COALESCE(prod.obs, ''), "
                                   f"prod.unidade, prod.localizacao, prod.quantidade "
                                   f"FROM movimentacao AS mov "
                                   f"INNER JOIN produto prod ON mov.produto = prod.id "
                                   f"WHERE mov.data = '{data_mov}' and mov.tipo < 200;")
                    dados_mov = cursor.fetchall()

                    if dados_mov:
                        for ii in dados_mov:
                            cod_filho, descr_filho, ref_filho, um_filho, local_filho, saldo_filho = ii

                            prod_saldo_encontrado = False
                            for cod_sal_e, descr_e in lista_produtos:
                                if cod_sal_e == cod_filho:
                                    prod_saldo_encontrado = True
                                    break

                            if not prod_saldo_encontrado:
                                saldo_float_filho = valores_para_float(saldo_filho)

                                if saldo_float_filho > 0:
                                    dados_colhidos = self.manipula_dados_onde_usa(cod_filho)
                                    if dados_colhidos:
                                        dados = (cod_filho, descr_filho)

                                        lista_produtos.append(dados)

                                        cur = conecta_robo.cursor()
                                        cur.execute(f"SELECT * from ENVIA_INDUSTRIALIZACAO "
                                                    f"where id_envia_mov = {id_mov} and cod_prod = {cod_filho};")
                                        dados_salvos = cur.fetchall()

                                        if not dados_salvos:
                                            cod_pai, descr_pai, ref_pai, um_pai, qtde_pai = dados_colhidos[0]

                                            caminho, arq = self.verifica_se_tem_pdf_desenho(ref_pai)

                                            if not caminho:
                                                self.envia_email_nao_acha_desenho(ref_pai, cod_pai)
                                            else:
                                                cursor = conecta.cursor()
                                                cursor.execute(
                                                    f"SELECT prod.conjunto, prod.tipomaterial, tip.tipomaterial, conj.conjunto "
                                                    f"FROM produto as prod "
                                                    f"LEFT JOIN conjuntos conj ON prod.conjunto = conj.id "
                                                    f"LEFT JOIN tipomaterial tip ON prod.tipomaterial = tip.id "
                                                    f"where prod.codigo = {cod_pai};")
                                                detalhes_produto = cursor.fetchall()
                                                num_conj_pai, num_tipo_pai, tipo_pai, conj_pai = detalhes_produto[0]

                                                ops_destino_pai = self.manipula_dados_tabela_consumo_final(cod_pai)
                                                vendas_destino_pai = self.manipula_dados_tabela_venda_final(cod_pai)

                                                agrupado[num_tipo_pai, tipo_pai].append(
                                                    (cod_pai, descr_pai, ref_pai, um_pai, cod_filho, descr_filho,
                                                     ref_filho, um_filho, local_filho, saldo_filho, ops_destino_pai,
                                                     vendas_destino_pai, id_mov))

            if agrupado:
                # --- Impressão agrupada ---
                for (num_tipo, tipo), itens in agrupado.items():
                    if not num_tipo:
                        self.envia_email_sem_tipo(itens)
                    else:
                        todos_tem_desenho = True
                        desenhos_faltando = []
                        lista_nao_enviado_orcamento = []

                        for iiitens in itens:
                            (cod_pai, descr_pai, ref_pai, um_pai, cod_filho, descr_filho,
                             ref_filho, um_filho, local_filho, saldo_filho,
                             ops_destino_pai, vendas_destino_pai, id_mov) = iiitens

                            tem_orcamento = self.verifica_orcamento_enviado_banco(cod_pai)

                            if not tem_orcamento:
                                dados = (cod_pai, descr_pai, ref_pai, um_pai, cod_filho, descr_filho,
                                                     ref_filho, um_filho, local_filho, saldo_filho, ops_destino_pai,
                                                     vendas_destino_pai, id_mov)
                                lista_nao_enviado_orcamento.append(dados)

                                caminho, arq = self.verifica_se_tem_pdf_desenho(ref_pai)

                                if not caminho:
                                    todos_tem_desenho = False
                                    dadinhos = (ref_pai, cod_pai)
                                    desenhos_faltando.append(dadinhos)
                                    break
                        if lista_nao_enviado_orcamento:
                            if todos_tem_desenho:
                                arquivo = 'Listagem - Ind.pdf'
                                caminho = os.path.join(os.path.dirname(os.path.abspath(__file__)), arquivo)

                                self.gerar_pdf_listagem_separar(caminho, lista_nao_enviado_orcamento)
                                self.inserir_no_banco(lista_nao_enviado_orcamento)

                                self.orcamento_com_desenho(num_tipo, tipo, lista_nao_enviado_orcamento)

                                self.excluir_arquivo(arquivo)
                            else:
                                for titi in desenhos_faltando:
                                    ref, cod = titi

                                    self.envia_email_nao_acha_desenho(ref, cod)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_orcamento_enviado_banco(self, codigo):
        try:
            tem_orcamento = False

            cur = conecta.cursor()
            cur.execute(f"SELECT id, descricao, COALESCE(obs, '') as obs, unidade, id_versao "
                        f"FROM produto where codigo = {codigo};")
            detalhes_produto = cur.fetchall()
            id_prod, descricao_id, referencia_id, unidade_id, id_versao = detalhes_produto[0]

            cur = conecta_robo.cursor()
            cur.execute(f"SELECT id, id_produto "
                        f"FROM PRODUTO_ORCAMENTO where id_produto = {id_prod};")
            detalhes_orcamento = cur.fetchall()

            if detalhes_orcamento:
                tem_orcamento = True

            return tem_orcamento

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def envia_email_sem_tipo(self, dados):
        try:
            saudacao, msg_final, to = self.dados_email()

            subject = f'SEM TIPO PLANO PCP - PRODUTOS SEM TIPO DEFINIDOS (PI)!'

            msg = MIMEMultipart()
            msg['From'] = email_user
            msg['Subject'] = subject

            body = f"{saudacao}\n\nAbaixo a lista de produtos sem tipo definido:\n\n"

            for i in dados:
                body += f"- {i}\n\n"

            body += f"\n{msg_final}"

            msg.attach(MIMEText(body, 'plain'))

            text = msg.as_string()
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(email_user, password)

            server.sendmail(email_user, to, text)
            server.quit()

            print("email enviado")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def orcamento_com_desenho(self, num_tipo, tipo, dados_itens):
        try:
            cursor = conecta_robo.cursor()
            cursor.execute(f"select GEN_ID(GEN_ENVIA_ORCAMENTO_ID,0) from rdb$database;")
            dados0 = cursor.fetchall()
            dados1 = dados0[0]
            dados2 = int(dados1[0]) + 1
            dados3 = str(dados2)
            num_sol = dados3

            arquivo_excel = self.gera_excel_orcamento(num_sol, dados_itens)

            self.envia_email_orcamento_com_desenho(num_sol, tipo, dados_itens, arquivo_excel)

            self.excluir_arquivo(arquivo_excel)

            self.inserir_banco_orcamento(num_tipo, dados_itens)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def inserir_banco_orcamento(self, num_tipo, dados_itens):
        try:
            cursor = conecta_robo.cursor()
            cursor.execute("select GEN_ID(GEN_ENVIA_ORCAMENTO_ID,0) from rdb$database;")
            ultimo_req0 = cursor.fetchall()
            ultimo_req1 = ultimo_req0[0]
            ultimo_req = int(ultimo_req1[0]) + 1

            cursor = conecta_robo.cursor()
            cursor.execute(f"Insert into ENVIA_ORCAMENTO (ID, ID_TIPO) "
                           f"values (GEN_ID(GEN_ENVIA_ORCAMENTO_ID,1), {num_tipo});")

            for i in dados_itens:
                (cod_pai, descr_pai, ref_pai, um_pai, cod_filho, descr_filho, ref_filho, um_filho,
                 local_filho, saldo_filho, ops_destino_pai, vendas_destino_pai, id_mov) = i

                cur = conecta.cursor()
                cur.execute(f"SELECT id, descricao, COALESCE(obs, '') as obs, unidade, id_versao "
                            f"FROM produto where codigo = {cod_pai};")
                detalhes_produto = cur.fetchall()
                id_prod, descricao_id, referencia_id, unidade_id, id_versao = detalhes_produto[0]

                cursor = conecta_robo.cursor()
                cursor.execute(f"Insert into PRODUTO_ORCAMENTO (ID, ID_ORCAMENTO, ID_PRODUTO, QTDE) "
                               f"values (GEN_ID(GEN_PRODUTO_ORCAMENTO_ID,1), {ultimo_req}, {id_prod}, "
                               f"'{saldo_filho}');")

            conecta_robo.commit()
            print(f"Nº Orçemento {ultimo_req} inserido no banco com sucesso!")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def envia_email_orcamento_com_desenho(self, num_sol, tipo, dados_itens, arquivo_excel):
        try:
            saudacao, msg_final, to = self.dados_email()

            subject = f"Solicitação de Orçamento – Compra Nº {num_sol} | Grupo {tipo}"

            msg = MIMEMultipart()
            msg['From'] = email_user
            msg['Subject'] = subject

            body = ""

            body += f"{saudacao}\n\n"
            body += f"Segue em anexo a solicitação de orçamento de Compra Nº {num_sol} do grupo de Fornecedores {tipo}, solicitada por Suzuki Máquinas.\n\n"
            body += f"Agradecemos pela atenção e pedimos que, em caso de dúvidas ou dificuldades com o arquivo, entre em contato pelo e-mail maquinas@unisold.com.br\n\n"
            body += f"{msg_final}"

            msg.attach(MIMEText(body, 'plain'))

            arquivo_sepera_pdf = 'Listagem - Ind.pdf'
            caminho_sepera_pdf = os.path.join(os.path.dirname(os.path.abspath(__file__)), arquivo_sepera_pdf)
            diretorio_atual = os.path.dirname(os.path.abspath(__file__))
            caminho_arquivo_pdf = os.path.join(diretorio_atual, caminho_sepera_pdf)
            attachment = open(caminho_arquivo_pdf, 'rb')
            part = MIMEBase('application', "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment',
                            filename=Header(arquivo_sepera_pdf, 'utf-8').encode())
            msg.attach(part)

            caminho_arquivo = fr'C:\Users\Anderson\PycharmProjects\robo_boby\{arquivo_excel}'

            attachment = open(caminho_arquivo, 'rb')

            part = MIMEBase('application', "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment',
                            filename=Header(arquivo_excel, 'utf-8').encode())
            msg.attach(part)

            for i in dados_itens:
                (cod_pai, descr_pai, ref_pai, um_pai, cod_filho, descr_filho, ref_filho,
                 um_filho, local_filho, saldo_filho, ops_destino_pai, vendas_destino_pai, id_mov) = i

                caminho, arq = self.verifica_se_tem_pdf_desenho(ref_pai)

                with open(caminho, 'rb') as attachment:
                    part = MIMEBase('application', "octet-stream")
                    part.set_payload(attachment.read())
                    encoders.encode_base64(part)
                    part.add_header('Content-Disposition', 'attachment',
                                    filename=Header(arq, 'utf-8').encode())
                    msg.attach(part)

            text = msg.as_string()
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(email_user, password)

            server.sendmail(email_user, to, text)
            attachment.close()

            server.quit()

            print(f'Orçamento {num_sol} enviado com sucesso!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def gera_excel_orcamento(self, num_sol, dados_itens):
        try:
            obs_solicitacao = "ORÇAMENTO GERADO AUTOMATICAMENTE"

            data_hoje = date.today()
            data_certa = data_hoje.strftime("%d/%m/%Y")

            d_um = []

            for i in dados_itens:
                (cod_pai, descr_pai, ref_pai, um_pai, cod_filho, descr_filho, ref_filho, um_filho,
                 local_filho, saldo_filho, ops_destino, vendas_destino, id_mov) = i

                msg = ""

                if ops_destino:
                    for opss in ops_destino:
                        num_op, cod_op, descr_op = opss

                        if msg:
                            msg += f"\n- OP {num_op} - {cod_op} - {descr_op}"
                        else:
                            msg += f"- OP {num_op} - {cod_op} - {descr_op}"

                if vendas_destino:
                    for vendass in vendas_destino:
                        num_pi, num_ov, cliente = vendass

                        if msg:
                            if num_pi:
                                numero = f"\nPI {num_pi}"
                            else:
                                numero = f"\nPI {num_ov}"
                        else:
                            if num_pi:
                                numero = f"PI {num_pi}"
                            else:
                                numero = f"PI {num_ov}"

                        msg += f"- {numero} - {cliente}\n"

                qtde_float = valores_para_float(saldo_filho)

                dados = (cod_pai, descr_pai, ref_pai, um_pai, qtde_float, msg)
                d_um.append(dados)

            df = pd.DataFrame(d_um, columns=['Código', 'Descrição', 'Referência', 'UM', 'Qtde', 'Destino'])

            codigo_int = {'Código': int}
            df = df.astype(codigo_int)
            qtde_float = {'Qtde': float}
            df = df.astype(qtde_float)

            caminho_arquivo_modelo = f'Mod_orcamento.xlsx'
            nome_arquivo = f'Orçamento {num_sol}.xlsx'

            book = load_workbook(caminho_arquivo_modelo)

            writer = pd.ExcelWriter(nome_arquivo, engine='openpyxl')

            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

            linhas_frame = df.shape[0]
            colunas_frame = df.shape[1]

            linhas_certas = linhas_frame + 2 + 9
            colunas_certas = colunas_frame + 1

            ws = book.active

            inicia = 11
            rows = range(inicia, inicia + linhas_frame)
            columns = range(1, colunas_certas)

            ws.row_dimensions[linhas_certas + 2].height = 30
            ws.row_dimensions[linhas_certas + 4].height = 30

            for row in rows:
                for col in columns:
                    ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center',
                                                            wrap_text=True)
                    ws.cell(row, col).border = Border(left=Side(border_style='thin', color='00000000'),
                                                      right=Side(border_style='thin', color='00000000'),
                                                      top=Side(border_style='thin', color='00000000'),
                                                      bottom=Side(border_style='thin', color='00000000'),
                                                      diagonal=Side(border_style='thick', color='00000000'),
                                                      diagonal_direction=0,
                                                      outline=Side(border_style='thin', color='00000000'),
                                                      vertical=Side(border_style='thin', color='00000000'),
                                                      horizontal=Side(border_style='thin', color='00000000'))

            ws.merge_cells(f'A8:D8')
            top_left_cell = ws[f'A8']
            c = ws[f'A8']
            c.alignment = Alignment(horizontal='center',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=14, bold=True)
            top_left_cell.value = 'Orçamento Nº  ' + num_sol

            ws.merge_cells(f'E8:F8')
            top_left_cell = ws[f'E8']
            c = ws[f'E8']
            c.alignment = Alignment(horizontal='center',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=14, bold=True)
            top_left_cell.value = 'Emissão:  ' + data_certa

            ws.merge_cells(f'B{linhas_certas + 2}:B{linhas_certas + 2}')
            top_left_cell = ws[f'B{linhas_certas + 2}']
            c = ws[f'B{linhas_certas + 2}']
            c.alignment = Alignment(horizontal='right',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=True)
            top_left_cell.value = "Observação:  "

            ws.merge_cells(f'C{linhas_certas + 2}:F{linhas_certas + 2}')
            top_left_cell = ws[f'C{linhas_certas + 2}']
            c = ws[f'C{linhas_certas + 2}']
            c.alignment = Alignment(horizontal='left',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=False)
            top_left_cell.value = obs_solicitacao

            df.to_excel(writer, 'Sheet1', startrow=10, startcol=0, header=False, index=False)

            writer.save()

            print("Excel Salvo")

            return nome_arquivo

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_dados_tabela_consumo_final(self, cod_prod):
        try:
            tabela_nova = []

            cursor = conecta.cursor()
            cursor.execute(f"SELECT prod.id, estprod.id, estprod.id_estrutura "
                           f"from estrutura_produto as estprod "
                           f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                           f"where prod.codigo = {cod_prod};")
            dados_estrut = cursor.fetchall()
            for i_estrut in dados_estrut:
                prod_id, ides_mat, id_estrutura = i_estrut

                cursor = conecta.cursor()
                cursor.execute(f"select id, id_produto, num_versao, data_versao, obs, data_criacao "
                               f"from estrutura where id = {id_estrutura};")
                estrutura = cursor.fetchall()

                id_produto = estrutura[0][1]

                cursor = conecta.cursor()
                cursor.execute(f"select ordser.datainicial, ordser.numero, prod.codigo, prod.descricao "
                               f"from ordemservico as ordser "
                               f"INNER JOIN produto prod ON ordser.produto = prod.id "
                               f"where ordser.status = 'A' "
                               f"and prod.id = {id_produto} and prod.id_versao = {id_estrutura} "
                               f"order by ordser.numero;")
                op_abertas = cursor.fetchall()
                if op_abertas:
                    for ii in op_abertas:
                        emissao, num_op, cod_pai, descr_pai = ii

                        cursor = conecta.cursor()
                        cursor.execute(f"SELECT estprod.id, prod.codigo, "
                                       f"((SELECT quantidade FROM ordemservico where numero = {num_op}) * "
                                       f"(estprod.quantidade)) AS Qtde "
                                       f"FROM estrutura_produto as estprod "
                                       f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                                       f"where estprod.id = {ides_mat};")
                        select_estrut = cursor.fetchall()
                        if select_estrut:
                            id_mat, cod_estrut, qtde_total = select_estrut[0]

                            cursor = conecta.cursor()
                            cursor.execute(f"SELECT max(prod.codigo), max(prod.descricao), "
                                           f"sum(prodser.QTDE_ESTRUT_PROD) as total "
                                           f"FROM estrutura_produto as estprod "
                                           f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                                           f"INNER JOIN produtoos as prodser ON estprod.id = prodser.ID_ESTRUT_PROD "
                                           f"where estprod.id_estrutura = {id_estrutura} "
                                           f"and prodser.numero = {num_op} and estprod.id = {id_mat} "
                                           f"group by prodser.ID_ESTRUT_PROD;")
                            select_os_resumo = cursor.fetchall()
                            if select_os_resumo:
                                for os_cons in select_os_resumo:
                                    cod_cons, descr_cons, qtde_cons_total = os_cons

                                    if qtde_total != qtde_cons_total:
                                        dados = (num_op, cod_pai, descr_pai)
                                        tabela_nova.append(dados)
                            else:
                                dados = (num_op, cod_pai, descr_pai)
                                tabela_nova.append(dados)

            return tabela_nova

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_dados_tabela_venda_final(self, cod_prod):
        try:
            tabela_nova = []

            cursor = conecta.cursor()
            cursor.execute(f"SELECT ped.emissao, ped.id, cli.razao, prodint.qtde, "
                           f"prodint.data_previsao "
                           f"FROM PRODUTOPEDIDOINTERNO as prodint "
                           f"INNER JOIN produto as prod ON prodint.id_produto = prod.id "
                           f"INNER JOIN pedidointerno as ped ON prodint.id_pedidointerno = ped.id "
                           f"INNER JOIN clientes as cli ON ped.id_cliente = cli.id "
                           f"where prodint.status = 'A' and prod.codigo = {cod_prod};")
            dados_pi = cursor.fetchall()

            if dados_pi:
                for i_pi in dados_pi:
                    emissao_pi, num_pi, clie_pi, qtde_pi, entrega_pi = i_pi

                    dados_pi = (num_pi, "", clie_pi)
                    tabela_nova.append(dados_pi)

            cursor = conecta.cursor()
            cursor.execute(f"SELECT oc.data, oc.numero, cli.razao, prodoc.quantidade, prodoc.dataentrega, "
                           f"COALESCE(prodoc.id_pedido, ''), COALESCE(prodoc.id_expedicao, '') "
                           f"FROM PRODUTOORDEMCOMPRA as prodoc "
                           f"INNER JOIN produto as prod ON prodoc.produto = prod.id "
                           f"INNER JOIN ordemcompra as oc ON prodoc.mestre = oc.id "
                           f"INNER JOIN clientes as cli ON oc.cliente = cli.id "
                           f"LEFT JOIN pedidointerno as ped ON prodoc.id_pedido = ped.id "
                           f"where prodoc.quantidade > prodoc.produzido "
                           f"and oc.status = 'A' "
                           f"and oc.entradasaida = 'S' "
                           f"and prod.codigo = {cod_prod};")
            dados_ov = cursor.fetchall()
            if dados_ov:
                for i_ov in dados_ov:
                    emissao_ov, num_ov, clie_ov, qtde_ov, entrega_ov, num_pi_ov, num_exp = i_ov

                    dados = (num_pi_ov, num_ov, clie_ov)
                    tabela_nova.append(dados)

            return tabela_nova

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def envia_email_nao_acha_desenho(self, arquivo_pdf, cod):
        try:
            saudacao, msg_final, to = self.dados_email()

            subject = f'IND - Não foi encontrado o desenho {arquivo_pdf}'

            msg = MIMEMultipart()
            msg['From'] = email_user
            msg['Subject'] = subject

            body = f"{saudacao}\n\n" \
                   f"O desenho {arquivo_pdf} de código {cod} não foi encontrado no cadastro dos produtos.\n\n" \
                   f"{msg_final}"

            msg.attach(MIMEText(body, 'plain'))

            text = msg.as_string()
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(email_user, password)

            server.sendmail(email_user, to, text)
            server.quit()

            print("email enviado sem arquivo pdf desenho")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_se_tem_pdf_desenho(self, ref):
        try:
            s = re.sub(r"[^\d.]", "", ref)  # remove tudo que não é número ou ponto
            s = re.sub(r"\.+$", "", s)  # saída: 47.00.014.07

            caminho_pdf = rf"\\Publico\C\OP\Projetos\{s}.pdf"
            arquivo_pdf = f"{s}.pdf"

            if not os.path.exists(caminho_pdf):
                caminho_pdf = ""
                arquivo_pdf = ""

            return caminho_pdf, arquivo_pdf

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_dados_onde_usa(self, cod_prod):
        try:
            planilha_nova = []

            cursor = conecta.cursor()
            cursor.execute(f"SELECT estprod.id, estprod.id_estrutura, estprod.quantidade "
                           f"from estrutura_produto as estprod "
                           f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                           f" where prod.codigo = {cod_prod};")
            tabela_estrutura = cursor.fetchall()
            for i in tabela_estrutura:
                ides_mat, id_estrutura, qtde = i

                cursor = conecta.cursor()
                cursor.execute(f"SELECT id, codigo "
                               f"from produto "
                               f" where id_versao = {id_estrutura};")
                produto_pai = cursor.fetchall()
                if produto_pai:
                    cod_produto = produto_pai[0][1]

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT prod.codigo, prod.descricao, COALESCE(prod.obs, ''), prod.unidade, "
                                   f"COALESCE(prod.obs2, '') "
                                   f"from estrutura as est "
                                   f"INNER JOIN produto prod ON est.id_produto = prod.id "
                                   f"where prod.codigo = {cod_produto} and prod.tipomaterial = 119;")
                    select_prod = cursor.fetchall()

                    if select_prod:
                        cod, descr, ref, um, obs = select_prod[0]
                        dados = (cod, descr, ref, um, qtde)
                        planilha_nova.append(dados)

            if planilha_nova:
                planilha_nova_ordenada = sorted(planilha_nova, key=lambda x: x[1])

                return planilha_nova_ordenada

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)


chama_classe = EnviaIndustrializacao()
