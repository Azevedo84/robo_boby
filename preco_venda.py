import fdb
import pandas as pd
from comandos.conversores import valores_para_float
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from openpyxl.styles import NamedStyle


conecta = fdb.connect(database=r'C:\HallSys\db\Horus\Suzuki\ESTOQUE.GDB',
                              host='PUBLICO',
                              port=3050,
                              user='sysdba',
                              password='masterkey',
                              charset='ANSI')


def verifica_estrutura_problema(nivel, codigo, qtde):
    cursor = conecta.cursor()
    cursor.execute(f"SELECT prod.id, prod.codigo, prod.descricao, prod.obs, prod.unidade, conj.conjunto, "
                   f"prod.tempo, prod.terceirizado, prod.custounitario, prod.custoestrutura "
                   f"FROM produto as prod "
                   f"LEFT JOIN tipomaterial as tip ON prod.tipomaterial = tip.id "
                   f"INNER JOIN conjuntos conj ON prod.conjunto = conj.id "
                   f"where prod.codigo = {codigo};")
    detalhes_pai = cursor.fetchall()
    id_pai, c_pai, des_pai, ref_pai, um_pai, conj_pai, temp_pai, terc_pai, unit_pai, est_pai = detalhes_pai[0]

    filhos = [(nivel, codigo, des_pai, ref_pai, um_pai, qtde, conj_pai, temp_pai, terc_pai, unit_pai, est_pai)]

    nivel_plus = nivel + 1

    cursor = conecta.cursor()
    cursor.execute(f"SELECT prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, prod.unidade, "
                   f"(mat.quantidade * {qtde}) as qtde "
                   f"FROM materiaprima as mat "
                   f"INNER JOIN produto prod ON mat.produto = prod.id "
                   f"where mestre = {id_pai};")
    dados_estrutura = cursor.fetchall()

    if dados_estrutura:
        for prod in dados_estrutura:
            cod_f, descr_f, ref_f, um_f, qtde_f = prod

            filhos.extend(verifica_estrutura_problema(nivel_plus, cod_f, qtde_f))

    return filhos


dataframe = pd.read_excel('Pasta1.xlsx', sheet_name='Planilha1')
codigos = dataframe['Código']

lista_materiais_erros = []
lista_final_venda = []

for codigo_produto in codigos:
    print("codigo excel:  ", codigo_produto)
    cursor = conecta.cursor()
    cursor.execute(f"SELECT prod.id, prod.codigo, prod.id_versao, prod.custounitario, prod.tempo, prod.terceirizado, "
                   f"tip.tipomaterial "
                   f"FROM produto  as prod "
                   f"LEFT JOIN tipomaterial tip ON prod.tipomaterial = tip.id "
                   f"where prod.codigo = {codigo_produto};")
    select_prod = cursor.fetchall()
    id_pai, cod, id_estrutura, custo_compra, tempo_mao, custo_servico, tipo_material = select_prod[0]

    nova_tabela = []

    cursor = conecta.cursor()
    cursor.execute(f"SELECT prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, "
                   f"prod.conjunto, prod.unidade, "
                   f"(estprod.quantidade * 1) as qtde, prod.terceirizado, prod.custounitario, "
                   f"prod.custoestrutura "
                   f"from estrutura_produto as estprod "
                   f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                   f"INNER JOIN conjuntos conj ON prod.conjunto = conj.id "
                   f"where estprod.id_estrutura = {id_estrutura} "
                   f"order by conj.conjunto DESC, prod.descricao ASC;")
    tabela_estrutura = cursor.fetchall()

    if tabela_estrutura:
        for i in tabela_estrutura:
            cod, descr, ref, conjunto, um, qtde, terc, unit, estrut = i

            qtde_float = valores_para_float(qtde)
            unit_float = valores_para_float(unit)
            estrut_float = valores_para_float(estrut)

            if conjunto == 10:
                total = qtde_float * estrut_float

                dados = (cod, descr, ref, um, qtde_float, estrut_float, total, conjunto)
                nova_tabela.append(dados)
            else:
                total = qtde_float * unit_float

                dados = (cod, descr, ref, um, qtde, unit_float, total, conjunto)
                nova_tabela.append(dados)

    if nova_tabela:
        valor_final = 0.00

        for dados in nova_tabela:
            total = dados[6]
            valor_final = valor_final + total

        valor_totau_dois = ("%.2f" % valor_final)

        tabela_nova_nova = []
        tab_estrut = []
        for itens in nova_tabela:
            cod = itens[0]
            qtde = itens[4]

            estrutura = verifica_estrutura_problema(1, cod, qtde)
            if estrutura:
                for i in estrutura:
                    tab_estrut.append(i)

        tab_ordenada = sorted(tab_estrut, key=lambda x: -x[0])

        for i in tab_ordenada:
            niv, codi, descr, ref, um, qtdi, conj, temp, terc, unit, estrut = i

            if conj == 'PRODUTOS ACABADOS':
                if temp or terc:
                    pass
                else:
                    if estrut:
                        estrut_float = float(estrut)
                    else:
                        estrut_float = 0
                    total = float(qtdi) * estrut_float

                    dados = (codi, descr, ref, um, qtdi, estrut_float, total, conj)
                    tabela_nova_nova.append(dados)
            else:
                if unit:
                    pass
                else:
                    if unit:
                        unit_float = float(unit)
                    else:
                        unit_float = 0

                    total = float(qtdi) * float(unit_float)

                    dados = (codi, descr, ref, um, qtdi, unit_float, total, conj)
                    tabela_nova_nova.append(dados)

        if tabela_nova_nova:
            tabela_nova_ordenada = sorted(tabela_nova_nova, key=lambda x: (x[1], x[0]))

            for ii in tabela_nova_ordenada:
                lista_materiais_erros.append(ii)
        else:
            custo_compra_float = valores_para_float(custo_compra)

            cursor = conecta.cursor()
            cursor.execute("SELECT valorhora FROM valoresmensais WHERE data = (SELECT MAX(data) FROM valoresmensais);")
            valores_mensais = cursor.fetchall()
            custo_mao = f"R$ {(valores_mensais[0][0])}"
            custo_mao_float = valores_para_float(custo_mao)
            tempo_mao_float = valores_para_float(tempo_mao)
            total1 = tempo_mao_float * custo_mao_float
            custo_mao_obra = str("%.2f" % total1)

            if tipo_material == "INDUSTRIALIZACAO":
                if custo_servico:
                    custo_ser = valores_para_float(custo_servico)
                else:
                    custo_ser = 0
            else:
                if custo_mao_obra:
                    custo_ser = valores_para_float(custo_mao_obra)
                else:
                    custo_ser = 0

            if valor_totau_dois:
                custo_materias_float = valores_para_float(valor_totau_dois)
            else:
                custo_materias_float = 0

            totalzao = custo_materias_float + custo_ser
            if totalzao:
                totalz = ("%.2f" % totalzao)
            else:
                totalz = 0

            preco = (valores_para_float(totalz) + (valores_para_float(totalz) * 0.05)) / 0.7663

            valor_totau_dois = ("%.2f" % preco)

            dados = (codigo_produto, valor_totau_dois)
            lista_final_venda.append(dados)

            if codigo_produto == 21423:
                print(codigo_produto, valor_totau_dois)

if lista_materiais_erros:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Preço de Venda"

    headers = ["Código", "Descrição", "Referência", "Conjunto"]
    sheet.append(headers)

    header_row = sheet[1]
    for cell in header_row:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for blabla in lista_materiais_erros:
        cod, descr, ref, um, qtdi, unit_float, total, conj = blabla
        # Adicione os valores formatados com o estilo criado
        sheet.append([cod, descr, ref, conj])

    currency_style = NamedStyle(name='currency')
    currency_style.number_format = 'R$ #,##0.00'

    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=4, max_row=sheet.max_row):
        for cell in row:
            cell.style = currency_style

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if isinstance(cell.value, (int, float)):
                cell_value_str = "{:.2f}".format(cell.value)
            else:
                cell_value_str = str(cell.value)
            if len(cell_value_str) > max_length:
                max_length = len(cell_value_str)

        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=7, max_col=9):
        for cell in row:
            cell.number_format = '0.00'

    desktop = Path.home() / "Desktop"
    desk_str = str(desktop)
    nome_req = '\Problemas Estrutura.xlsx'
    caminho = (desk_str + nome_req)

    workbook.save(caminho)

    print("Excel Problemas salvo!")
else:
    if lista_final_venda:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Preço de Venda"

        headers = ["Código", "Descrição", "Referência", "Preço de Venda"]
        sheet.append(headers)

        header_row = sheet[1]
        for cell in header_row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for blabla in lista_final_venda:
            codigo_produto, valor_venda = blabla

            valor_venda_float = valores_para_float(valor_venda)

            cursor = conecta.cursor()
            cursor.execute(
                f"SELECT codigo, descricao, COALESCE(obs, '') "
                f"FROM produto "
                f"where codigo = {codigo_produto};")
            select_prod = cursor.fetchall()

            cod = select_prod[0][0]
            descr = select_prod[0][1]
            ref = select_prod[0][2]
            # Adicione os valores formatados com o estilo criado
            sheet.append([cod, descr, ref, valor_venda_float])

        # Estilo de moeda para o formato brasileiro diretamente na coluna 'Preço de Venda'
        for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4, max_row=sheet.max_row):
            for cell in row:
                cell.number_format = 'R$ #,##0.00'  # Formato brasileiro

        # Adicionando bordas e alinhamento para todas as células
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ajuste de largura das colunas
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                cell_value_str = str(cell.value) if cell.value is not None else ""
                if len(cell_value_str) > max_length:
                    max_length = len(cell_value_str)
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Salvando o arquivo no Desktop
        desktop = Path.home() / "Desktop"
        caminho = str(desktop / "Preço de Venda.xlsx")
        workbook.save(caminho)

        print("Excel Vendas salvo!")

