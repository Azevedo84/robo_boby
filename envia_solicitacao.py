import sys
from banco_dados.conexao import conecta, conecta_robo
from banco_dados.controle_erros import grava_erro_banco
import os
import inspect
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.header import Header
from email import encoders
from datetime import datetime
import traceback
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Side, Alignment, Border, Font
import openpyxl.styles as styles
from unidecode import unidecode

from dados_email import email_user, password


class EnviaSolicitacaoCompra:
    def __init__(self):
        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)
        self.diretorio_script = os.path.dirname(nome_arquivo_com_caminho)
        nome_base = os.path.splitext(self.nome_arquivo)[0]
        self.arquivo_log = os.path.join(self.diretorio_script, f"{nome_base}_erros.txt")

    def trata_excecao(self, nome_funcao, mensagem, arquivo, excecao):
        try:
            tb = traceback.extract_tb(excecao)
            num_linha_erro = tb[-1][1]

            traceback.print_exc()
            print(f'Houve um problema no arquivo: {arquivo} na função: "{nome_funcao}"\n{mensagem} {num_linha_erro}')

            grava_erro_banco(nome_funcao, mensagem, arquivo, num_linha_erro)

            # 'Log' em arquivo local apenas se houver erro
            with open(self.arquivo_log, "a", encoding="utf-8") as f:
                f.write(f"Erro na função {nome_funcao} do arquivo {arquivo}: {mensagem} (linha {num_linha_erro})\n")

        except Exception as e:
            nome_funcao_trat = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            tb = traceback.extract_tb(exc_traceback)
            num_linha_erro = tb[-1][1]
            print(f'Houve um problema no arquivo: {self.nome_arquivo} na função: "{nome_funcao_trat}"\n'
                  f'{e} {num_linha_erro}')
            grava_erro_banco(nome_funcao_trat, e, self.nome_arquivo, num_linha_erro)

            with open(self.arquivo_log, "a", encoding="utf-8") as f:
                f.write(
                    f"Erro na função {nome_funcao_trat} do arquivo {self.nome_arquivo}: {e} (linha {num_linha_erro})\n")

    def envia_email(self, numero_sol, nome_computador, anexos):
        try:
            cursor = conecta.cursor()
            cursor.execute(f"SELECT versao, responsavel FROM ENVIA_PC where descricao = '{nome_computador}';")
            dados_pc = cursor.fetchall()

            if dados_pc:
                solicitante = dados_pc[0][1]
            else:
                solicitante = "Desconhecido"

            saudacao, msg_final, to = self.dados_email()

            subject = f'Solicitação Nº {numero_sol}'

            msg = MIMEMultipart()
            msg['From'] = email_user
            msg['Subject'] = subject

            body = f"{saudacao}\n\n" \
                   f"Segue solicitação de compra Nº {numero_sol}.\n\n" \
                   f"Solicitado por: {solicitante}\n\n\n\n" \
                   f"{msg_final}"

            msg.attach(MIMEText(body, 'plain'))

            nome_arquivo = f'Solicitação {numero_sol}.xlsx'
            caminho_arquivo = fr'C:\Users\Anderson\PycharmProjects\robo_boby\Solicitação {numero_sol}.xlsx'

            attachment = open(caminho_arquivo, 'rb')

            part = MIMEBase('application', "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment',
                            filename=Header(nome_arquivo, 'utf-8').encode())
            msg.attach(part)

            if anexos:
                print(anexos)
                for anex in anexos:
                    arquivo_final, caminho_final = anex

                    with open(caminho_final, 'rb') as attachment:
                        part = MIMEBase('application', "octet-stream")
                        part.set_payload(attachment.read())
                        encoders.encode_base64(part)
                        part.add_header('Content-Disposition', 'attachment',
                                        filename=Header(arquivo_final, 'utf-8').encode())
                        msg.attach(part)

            text = msg.as_string()
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(email_user, password)

            server.sendmail(email_user, to, text)
            attachment.close()

            server.quit()

            print(f'Solicitação {numero_sol} criada com sucesso!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def dados_email(self):
        try:
            to = ['<maquinas@unisold.com.br>']

            current_time = (datetime.now())
            horario = current_time.strftime('%H')
            hora_int = int(horario)
            saudacao = ""
            if 4 < hora_int < 13:
                saudacao = "Bom Dia!"
            elif 12 < hora_int < 19:
                saudacao = "Boa Tarde!"
            elif hora_int > 18:
                saudacao = "Boa Noite!"
            elif hora_int < 5:
                saudacao = "Boa Noite!"

            msg_final = f"Att,\n" \
                        f"Suzuki Máquinas Ltda\n" \
                        f"Fone (51) 3561.2583/(51) 3170.0965\n\n" \
                        f"Mensagem enviada automaticamente, por favor não responda.\n\n" \
                        f"Se houver algum problema com o recebimento de emails ou conflitos com o arquivo excel, " \
                        f"favor entrar em contato pelo email maquinas@unisold.com.br.\n\n"

            return saudacao, msg_final, to

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def gera_excel(self, num_sol, dados_tabela, dados_solicitacao):
        try:
            data, obs_solicitacao, pc = dados_solicitacao[0]

            dataemissao = data.strftime('%d/%m/%Y')

            if not obs_solicitacao:
                obs_sol = ""
            else:
                obs_sol_maiuscula = obs_solicitacao.upper()
                obs_sol = unidecode(obs_sol_maiuscula)

            d_um = []

            embalagem_sim_rows = []

            for tabi in dados_tabela:
                cod_1, desc_1, ref_1, um_1, qtde_1, destino = tabi
                qtdezinha_float = float(qtde_1)

                cursor = conecta.cursor()
                cursor.execute(f"SELECT id, codigo, embalagem FROM produto where codigo = '{cod_1}';")
                dados_produto = cursor.fetchall()
                if dados_produto:
                    id_produto, codigo, embalagem = dados_produto[0]
                    if embalagem == "SIM" or embalagem == "SER":
                        embalagem_sim_rows.append(len(d_um) - 1)

                dados = (cod_1, desc_1, ref_1, um_1, qtdezinha_float, destino)
                d_um.append(dados)

            df = pd.DataFrame(d_um, columns=['Código', 'Descrição', 'Referência', 'UM', 'Qtde', 'Destino'])

            codigo_int = {'Código': int}
            df = df.astype(codigo_int)
            qtde_float = {'Qtde': float}
            df = df.astype(qtde_float)

            buuks = fr'C:\Users\Anderson\PycharmProjects\robo_boby\Modelo.xlsx'

            nome_arquivo = fr'C:\Users\Anderson\PycharmProjects\robo_boby\Solicitação {num_sol}.xlsx'

            book = load_workbook(buuks)

            writer = pd.ExcelWriter(nome_arquivo, engine='openpyxl')

            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

            linhas_frame = df.shape[0]
            colunas_frame = df.shape[1]

            linhas_certas = linhas_frame + 2 + 9
            colunas_certas = colunas_frame + 1

            ws = book.active

            inicia = 11
            rows = range(inicia, inicia + linhas_frame)
            columns = range(1, colunas_certas)

            ws.row_dimensions[linhas_certas + 2].height = 30
            ws.row_dimensions[linhas_certas + 4].height = 30

            for row in rows:
                for col in columns:
                    ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center',
                                                            wrap_text=True)
                    ws.cell(row, col).border = Border(left=Side(border_style='thin', color='00000000'),
                                                      right=Side(border_style='thin', color='00000000'),
                                                      top=Side(border_style='thin', color='00000000'),
                                                      bottom=Side(border_style='thin', color='00000000'),
                                                      diagonal=Side(border_style='thick', color='00000000'),
                                                      diagonal_direction=0,
                                                      outline=Side(border_style='thin', color='00000000'),
                                                      vertical=Side(border_style='thin', color='00000000'),
                                                      horizontal=Side(border_style='thin', color='00000000'))

            ws.merge_cells(f'A8:D8')
            top_left_cell = ws[f'A8']
            c = ws[f'A8']
            c.alignment = Alignment(horizontal='center',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=14, bold=True)
            top_left_cell.value = 'Solicitação Nº  ' + str(num_sol)

            ws.merge_cells(f'E8:F8')
            top_left_cell = ws[f'E8']
            c = ws[f'E8']
            c.alignment = Alignment(horizontal='center',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=14, bold=True)
            top_left_cell.value = 'Emissão:  ' + str(dataemissao)

            ws.merge_cells(f'B{linhas_certas + 2}:B{linhas_certas + 2}')
            top_left_cell = ws[f'B{linhas_certas + 2}']
            c = ws[f'B{linhas_certas + 2}']
            c.alignment = Alignment(horizontal='right',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=True)
            top_left_cell.value = "Observação:  "

            ws.merge_cells(f'C{linhas_certas + 2}:H{linhas_certas + 2}')
            top_left_cell = ws[f'C{linhas_certas + 2}']
            c = ws[f'C{linhas_certas + 2}']
            c.alignment = Alignment(horizontal='left',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=False)
            top_left_cell.value = obs_sol

            df.to_excel(writer, 'Sheet1', startrow=10, startcol=0, header=False, index=False)

            for row_idx in embalagem_sim_rows:
                row = row_idx + 12
                col = 3
                ws.cell(row, col).fill = styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            writer.save()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_dados_anexo(self, num_solicitacao):
        try:
            tabela_nova = []

            cursor = conecta.cursor()
            cursor.execute(f"SELECT anex.id_solicitacao, anex.caminho, sol.nome_pc "
                           f"FROM solicitacao_anexo as anex "
                           f"LEFT JOIN produtoordemsolicitacao as prodsol ON anex.id_solicitacao = prodsol.mestre "
                           f"LEFT JOIN ordemsolicitacao AS sol ON anex.id_solicitacao = sol.idsolicitacao "
                           f"WHERE prodsol.status = 'A' and prodsol.mestre = {num_solicitacao} "
                           f"group by anex.id_solicitacao, anex.caminho, sol.nome_pc;")
            extrair_sol = cursor.fetchall()

            for dados in extrair_sol:
                num_sol, caminhos, pc = dados

                nomes_de_arquivo = [os.path.basename(caminhos)]

                dadus = (nomes_de_arquivo[0], caminhos)
                tabela_nova.append(dadus)

            return tabela_nova

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_solicitacao_abertas(self):
        try:
            cursor = conecta.cursor()
            cursor.execute(f"SELECT prodsol.mestre, prodsol.produto, prod.obs, prodsol.quantidade "
                           f"FROM produtoordemsolicitacao as prodsol "
                           f"INNER JOIN produto as prod ON prodsol.produto = prod.id "
                           f"WHERE prodsol.status = 'A' "
                           f"ORDER BY prodsol.mestre;")
            extrair_sol = cursor.fetchall()

            numeros_solicitacao = []
            for i in extrair_sol:
                num_sol, id_prod, obs, qtde = i
                if num_sol not in numeros_solicitacao:
                    numeros_solicitacao.append(num_sol)

            if numeros_solicitacao:
                for numero_sol in numeros_solicitacao:
                    cursor = conecta_robo.cursor()
                    cursor.execute(f"SELECT * FROM ENVIA_SOLICITACAO where num_sol = {numero_sol};")
                    select_envio = cursor.fetchall()

                    if not select_envio:
                        cursor = conecta.cursor()
                        cursor.execute(f"SELECT prod.codigo, prod.descricao, COALESCE(prod.obs, ' ') as obs, "
                                       f"prod.unidade, "
                                       f"prodsol.quantidade, prodsol.destino "
                                       f"FROM produtoordemsolicitacao as prodsol "
                                       f"INNER JOIN produto as prod ON prodsol.produto = prod.id "
                                       f"WHERE prodsol.mestre = {numero_sol} "
                                       f"ORDER BY prodsol.mestre;")
                        dados_produtos = cursor.fetchall()

                        cursor = conecta.cursor()
                        cursor.execute(f"SELECT dataemissao, obs, nome_pc "
                                       f"FROM ordemsolicitacao "
                                       f"WHERE idsolicitacao = {numero_sol};")
                        dados_sol = cursor.fetchall()
                        data, obs_solicitacao, pc = dados_sol[0]

                        self.gera_excel(numero_sol, dados_produtos, dados_sol)

                        dados_anexos = self.manipula_dados_anexo(numero_sol)

                        self.envia_email(numero_sol, pc, dados_anexos)

                        self.inserir_no_banco(numero_sol)

                        self.excluir_excel(numero_sol)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def inserir_no_banco(self, num_sol):
        try:
            cursor = conecta_robo.cursor()
            cursor.execute(f"Insert into ENVIA_SOLICITACAO (ID, NUM_SOL) "
                           f"values (GEN_ID(GEN_ENVIA_SOLICITACAO_ID,1), {num_sol});")
            print(f"Nº Solicitação {num_sol} inserido no banco com sucesso!")

            conecta_robo.commit()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excluir_excel(self, num_sol):
        try:
            nome_arquivo = fr'C:\Users\Anderson\PycharmProjects\robo_boby\Solicitação {num_sol}.xlsx'

            if os.path.exists(nome_arquivo):
                os.remove(nome_arquivo)
                print(f'O arquivo {nome_arquivo} foi excluído com sucesso.')
            else:
                print(f'O arquivo {nome_arquivo} não existe.')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)


chama_classe = EnviaSolicitacaoCompra()
chama_classe.manipula_solicitacao_abertas()
