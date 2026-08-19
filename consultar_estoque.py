from core.banco import conecta
from core.erros import trata_excecao
from collections import defaultdict
import re
from openpyxl import Workbook
from openpyxl.styles import Font
from pathlib import Path

class NFComprasPendente:
    def __init__(self):

        self.maquinas = {
            1: "Suzuki Maq",
            2: "Koba Plástico",
            3: "Plástico Suzuki",
            4: "Home Bag",
            5: "Voti Plástico",
            6: "UNISOLD - Picotadeira",
            7: "Plaza Plástico",
            8: "Tachiro Takata",
            9: "AHC",
            10: "UNISOLD - Picotadeira",
            11: "Picotadeira PoliMaq",
            12: "Bobinadeira Pic",
            13: "FlowPack",
            14: "UNISOLD - Triângulo",
            15: "Aglutinadora",
            16: "Injetora Plástico",
            17: "MicroFuro",
            18: "PACIFIL - Extrusora Contínua",
            20: "Bobinadeira",
            21: "UNISOLD - Ponto Bobina",
            22: "UNISOLD - Corte e Solda",
            23: "UNISOLD - Impressora 2 Cores",
            24: "UNISOLD - Corte e Solda 106",
            25: "Elétrica",
            26: "Extrusora PP",
            27: "Extrusora 5C",
            28: "MicroFuro",
            29: "Impressora Canevalli",
            30: "Extrusora 3B",
            31: "Fresa Tábua CNC",
            32: "Estrutura Guincho",
            33: "PACIFIL - Extrusora Pacifil",
            34: "UNISOLD - Impressora Roto",
            35: "Logística",
            36: "Bobinador Tratador",
            37: "PACIFIL - Braço Robótico",
            38: "Extrusora Tábua",
            39: "Furadeira Horin",
            40: "Peças Especiais",
            41: "Móveis Rec",
            42: "UNISOLD - Torre Giratória",
            43: "Esteiras",
            44: "Extrusora Saco Lixo",
            45: "Extrusora Modular",
            46: "PACIFIL - Dobrador de Bolsa",
            47: "PACIFIL - Extrusora 5 Cam",
            48: "PACIFIL - Dobradora Lona",
            49: "PACIFIL - Laboratório",
            51: "PACIFIL - Soldador Bolsa Auto",
            52: "PACIFIL - Armação Bolsa",
            53: "PACIFIL - Mesa Pacifil",
            54: "Seladora Túnel",
            55: "PACIFIL - Mesa Oval B. 5 Pés",
            56: "PACIFIL - Mesa Oval B. 9 Pés",
            57: "PACIFIL - Mesa Oval B. 9/10",
            58: "PACIFIL - Alterações MB",
            59: "PACIFIL - Trole Bolsa",
            60: "PACIFIL - Mesa Oval B. 12 Pés",
            62: "Impressora Rami",
            63: "PACIFIL - Seladora Lacre Bolsa",
            505: "UNISOLD - Triangulo",
            506: "UNISOLD - Triangulo",
            510: "UNISOLD - Triangulo",
            513: "UNISOLD - Triangulo",
            516: "UNISOLD - Triangulo",
            522: "UNISOLD - Triangulo",
            525: "UNISOLD - Triangulo",
            526: "UNISOLD - Triangulo",
            566: "UNISOLD - Triangulo",
            541: "UNISOLD - Triangulo",
            551: "UNISOLD - Triangulo",
            552: "UNISOLD - Triangulo",
            625: "UNISOLD - Impressora Roto",
            61: "PACIFIL - Medidor Espessura",
            64: "Extrusora 5 C",
            66: "Extrusora D",
            67: "PACIFIL - Balança",
            68: "Seladora Embalagem Pequena",
            69: "Alterações seladora",
            71: "PACIFIL - Seladora Grande",
            72: "PACIFIL - EXTRUSORA C",
            73: "PACIFIL - EXTRUSORA CARNEVALI PACIFIL",
            420: "UNISOLD - Extrusora",
        }

        cursor = conecta.cursor()
        cursor.execute(f"SELECT prod.codigo, prod.DESCRICAO, prod.obs, prod.unidade, prod.LOCALIZACAO, "
                       f"prod.CUSTOUNITARIO, prod.CUSTOMATERIAL, prod.QUANTIDADE, tip.TIPOMATERIAL "
                       f"FROM produto as prod "
                       f"LEFT JOIN tipomaterial as tip ON prod.tipomaterial = tip.id "
                       f"INNER JOIN SALDO_ESTOQUE as sald ON prod.id = sald.produto_id "
                       f"WHERE prod.QUANTIDADE > 0 "
                       f"and sald.local_estoque = 1 "
                       f"AND sald.saldo > 0;")
        dados_estoque = cursor.fetchall()

        print("Total de Itens:", len(dados_estoque))

        if dados_estoque:
            itens_pacifil, itens_unisold, itens_sem_uso = self.manipula_dados_tabela_usado(dados_estoque)

            print("Total de Itens Pacifil:", len(itens_pacifil))
            print("Total de Itens Unisold:", len(itens_unisold))
            print("Total de Itens Sem Uso:", len(itens_sem_uso))

            self.gerar_excel(itens_pacifil, itens_unisold, itens_sem_uso)

    def gerar_excel(self, itens_pacifil, itens_unisold, itens_sem_uso):
        try:
            wb = Workbook()

            # Remove a aba padrão
            wb.remove(wb.active)

            def criar_aba(nome, dados):
                ws = wb.create_sheet(title=nome)

                cabecalho = [
                    "Código",
                    "Descrição",
                    "Referência",
                    "UM",
                    "Localização",
                    "Custo Unitário",
                    "Custo Mat",
                    "Saldo",
                    "Tipo Material"
                ]

                for col, titulo in enumerate(cabecalho, start=1):
                    cel = ws.cell(row=1, column=col)
                    cel.value = titulo
                    cel.font = Font(bold=True)

                for linha, item in enumerate(dados, start=2):
                    for coluna, valor in enumerate(item, start=1):
                        ws.cell(row=linha, column=coluna).value = valor

                # Ajusta largura das colunas
                for coluna in ws.columns:
                    tamanho = max(
                        len(str(c.value)) if c.value is not None else 0
                        for c in coluna
                    )
                    ws.column_dimensions[coluna[0].column_letter].width = tamanho + 3


            criar_aba("PACIFIL", itens_pacifil)
            criar_aba("UNISOLD", itens_unisold)
            criar_aba("SEM USO", itens_sem_uso)

            caminho = Path.home() / "Desktop" / "Classificacao_Estoque.xlsx"

            wb.save(caminho)

            print(f"Arquivo salvo em:\n{caminho}")

        except Exception as e:
            trata_excecao(e)
            raise

    def manipula_dados_tabela_usado(self, dados_estoque):
        try:
            itens_pacifil = []
            itens_unisold = []
            itens_sem_uso = []

            for i in dados_estoque:
                cod_prod, descr, ref, um, local, custo, custo_maq, saldo, tipo = i

                cursor = conecta.cursor()
                cursor.execute(f"""
                    SELECT estprod.id,
                           estprod.id_estrutura,
                           estprod.quantidade
                    FROM estrutura_produto estprod
                    INNER JOIN produto prod
                        ON estprod.id_prod_filho = prod.id
                    WHERE prod.codigo = {cod_prod};
                """)
                tabela_estrutura = cursor.fetchall()

                uso_qtde = defaultdict(float)  # quantidade consumida
                uso_maquinas = defaultdict(int)  # quantidade de máquinas

                for ides_mat, id_estrutura, qtde in tabela_estrutura:

                    cursor = conecta.cursor()
                    cursor.execute(f"""
                        SELECT codigo
                        FROM produto
                        WHERE id_versao = {id_estrutura};
                    """)
                    produto_pai = cursor.fetchone()

                    if not produto_pai:
                        continue

                    cod_produto = produto_pai[0]

                    cursor = conecta.cursor()
                    cursor.execute(f"""
                        SELECT prod.codigo,
                               prod.descricao,
                               COALESCE(prod.obs,''),
                               prod.unidade,
                               COALESCE(prod.obs2,'')
                        FROM estrutura est
                        INNER JOIN produto prod
                            ON est.id_produto = prod.id
                        WHERE prod.codigo = {cod_produto};
                    """)
                    select_prod = cursor.fetchone()

                    if not select_prod:
                        continue

                    cod, descr, ref, um, obs = select_prod

                    ref = ref.strip()

                    m = re.search(r"^D\s+(\d+)\.", ref)

                    if not m:
                        continue

                    codigo_maquina = int(m.group(1))

                    nome_maquina = self.maquinas.get(codigo_maquina)

                    if not nome_maquina:
                        continue

                    if "PACIFIL" in nome_maquina:
                        grupo = "PACIFIL"
                    elif "UNISOLD" in nome_maquina:
                        grupo = "UNISOLD"
                    else:
                        grupo = "OUTROS"

                    uso_qtde[grupo] += float(qtde)
                    uso_maquinas[grupo] += 1

                # -----------------------------
                # Resultado
                # -----------------------------

                pac_qtde = uso_qtde["PACIFIL"]
                uni_qtde = uso_qtde["UNISOLD"]

                pac_maquinas = uso_maquinas["PACIFIL"]
                uni_maquinas = uso_maquinas["UNISOLD"]

                # Não é usado por nenhuma máquina da Pacifil ou Unisold
                if pac_qtde == 0 and uni_qtde == 0:
                    itens_sem_uso.append(i)

                else:

                    # Classificação pela quantidade utilizada
                    if pac_qtde > uni_qtde:
                        classificacao = "PACIFIL"

                    elif uni_qtde > pac_qtde:
                        classificacao = "UNISOLD"

                    else:
                        # Empatou na quantidade -> usa número de máquinas
                        if pac_maquinas > uni_maquinas:
                            classificacao = "PACIFIL"
                        else:
                            # Empate nas máquinas também -> UNISOLD
                            classificacao = "UNISOLD"

                    if classificacao == "PACIFIL":
                        itens_pacifil.append(i)
                    else:
                        itens_unisold.append(i)

            return itens_pacifil, itens_unisold, itens_sem_uso

        except Exception as e:
            trata_excecao(e)
            raise

if __name__ == "__main__":
    NFComprasPendente()