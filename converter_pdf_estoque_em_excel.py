import re
from pathlib import Path
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================
# CONFIGURAÇÃO
# ============================================================

NOME_PDF = r"C:\Users\Anderson\Desktop\pendencias\MAQ_06.2026_LOCAL 1 - PROPRIO.pdf"
NOME_EXCEL = r"C:\Users\Anderson\Desktop\estoque_30-06-2026.xlsx"


def numero_br(valor):
    """Converte número no padrão brasileiro para float."""
    valor = valor.strip().replace(".", "").replace(",", ".")
    return float(valor)


def extrair_itens(pdf_path):
    """
    Extrai os itens do Registro de Inventário.

    O PDF possui itens que, em alguns casos, quebram a descrição
    em duas linhas. Por isso a função guarda a primeira parte
    e completa a descrição quando encontra a linha com UN/QTD/VALORES.
    """

    # Linha completa de um item quando não há quebra de descrição.
    padrao_completo = re.compile(
        r"\|\s*(\d{8})\s*\|\s*(\d+)\s+(.+?)\s*\|\s*"
        r"([A-Z0-9²³]+)\s*\|\s*"
        r"([\d.,]+)\s*\|\s*"
        r"([\d.,]+)\s*\|\s*"
        r"([\d.,]+)\s*\|"
    )

    # Primeira linha de item com descrição quebrada.
    padrao_inicio = re.compile(
        r"\|\s*(\d{8})\s*\|\s*(\d+)\s+(.+?)\s*\|\s*\|\s*\|\s*\|\s*\|"
    )

    # Segunda linha, contendo complemento + unidade + valores.
    padrao_final = re.compile(
        r"\|\s*\|\s*(.*?)\s*\|\s*"
        r"([A-Z0-9²³]+)\s*\|\s*"
        r"([\d.,]+)\s*\|\s*"
        r"([\d.,]+)\s*\|\s*"
        r"([\d.,]+)\s*\|"
    )

    itens = []
    pendente = None

    with pdfplumber.open(pdf_path) as pdf:
        for pagina in pdf.pages:
            texto = pagina.extract_text() or ""

            for linha in texto.splitlines():
                linha = linha.strip()

                # Ignora cabeçalhos, totais e linhas vazias.
                if not linha or "REGISTRO DE INVENTARIO" in linha:
                    continue
                if "CLASSI-" in linha or "FICACAO" in linha:
                    continue
                if "TOTAL MATERIAL" in linha:
                    continue

                # Se existe item pendente, tenta completar primeiro.
                if pendente:
                    m_final = padrao_final.search(linha)
                    if m_final:
                        complemento, unidade, qtd, unitario, total = m_final.groups()

                        descricao = (
                                pendente["Descrição"] + " " + complemento
                        ).strip()

                        itens.append({
                            "NCM": pendente["NCM"],
                            "Código": pendente["Código"],
                            "Descrição": descricao,
                            "Unidade": unidade,
                            "Quantidade": numero_br(qtd),
                            "Valor unitário": numero_br(unitario),
                            "Valor total": numero_br(total),
                        })

                        pendente = None
                        continue

                    # Se apareceu outro item antes de completar o anterior,
                    # descarta o pendente para não misturar produtos.
                    if padrao_completo.search(linha) or padrao_inicio.search(linha):
                        pendente = None
                    else:
                        continue

                # Item completo em uma única linha.
                m = padrao_completo.search(linha)
                if m:
                    ncm, codigo, descricao, unidade, qtd, unitario, total = m.groups()

                    # Evita linhas que não são efetivamente produtos.
                    if codigo.isdigit():
                        itens.append({
                            "NCM": ncm,
                            "Código": codigo,
                            "Descrição": descricao.strip(),
                            "Unidade": unidade,
                            "Quantidade": numero_br(qtd),
                            "Valor unitário": numero_br(unitario),
                            "Valor total": numero_br(total),
                        })
                    continue

                # Início de item com descrição quebrada.
                m = padrao_inicio.search(linha)
                if m:
                    ncm, codigo, descricao = m.groups()
                    pendente = {
                        "NCM": ncm,
                        "Código": codigo,
                        "Descrição": descricao.strip(),
                    }

    return itens


def gerar_excel(itens, caminho_excel):
    wb = Workbook()
    ws = wb.active
    ws.title = "Estoque"

    cabecalho = [
        "NCM",
        "Código",
        "Descrição",
        "Unidade",
        "Quantidade",
        "Valor unitário",
        "Valor total",
    ]

    ws.append(cabecalho)

    # Cabeçalho.
    for celula in ws[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor="1F4E78")
        celula.alignment = Alignment(horizontal="center", vertical="center")

    # Dados.
    for item in itens:
        ws.append([
            item["NCM"],
            item["Código"],
            item["Descrição"],
            item["Unidade"],
            item["Quantidade"],
            item["Valor unitário"],
            item["Valor total"],
        ])

    # Formatação.
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for linha in range(2, ws.max_row + 1):
        ws.cell(linha, 5).number_format = '#,##0.###'
        ws.cell(linha, 6).number_format = 'R$ #,##0.00'
        ws.cell(linha, 7).number_format = 'R$ #,##0.00'

    larguras = {
        "A": 14,
        "B": 12,
        "C": 80,
        "D": 12,
        "E": 15,
        "F": 18,
        "G": 18,
    }

    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura

    ws.row_dimensions[1].height = 24

    # Segunda aba com resumo.
    resumo = wb.create_sheet("Resumo")
    resumo["A1"] = "Resumo do estoque"
    resumo["A1"].font = Font(bold=True, size=14)

    resumo["A3"] = "Quantidade de itens"
    resumo["B3"] = len(itens)

    resumo["A4"] = "Valor total do estoque"
    resumo["B4"] = "=SUM(Estoque!G:G)"
    resumo["B4"].number_format = 'R$ #,##0.00'

    resumo["A6"] = "Data do estoque"
    resumo["B6"] = "30/06/2026"

    resumo.column_dimensions["A"].width = 28
    resumo.column_dimensions["B"].width = 22

    wb.save(caminho_excel)


def main():
    pdf_path = Path(NOME_PDF)
    excel_path = Path(NOME_EXCEL)

    if not pdf_path.exists():
        print()
        print("ERRO: não encontrei o PDF.")
        print(f"Esperado: {pdf_path}")
        print()
        input("Pressione ENTER para sair...")
        return

    print("Lendo PDF...")
    itens = extrair_itens(pdf_path)

    print(f"Itens encontrados: {len(itens)}")
    print("Gerando Excel...")

    gerar_excel(itens, excel_path)

    print()
    print("Concluído!")
    print(f"Excel criado em:")
    print(excel_path)
    print()


if __name__ == "__main__":
    main()
