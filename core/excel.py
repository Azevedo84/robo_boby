from core.erros import trata_excecao
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl import Workbook, load_workbook, drawing
from openpyxl.utils import get_column_letter
import pandas as pd


def lanca_dados_mesclado(mp_copy, mesclado, celula, informacao, tam_fonte, negrito):
    try:
        mp_copy.merge_cells(mesclado)
        celula_sup_esq = mp_copy[celula]
        cel = mp_copy[celula]
        edita_alinhamento(cel)
        edita_fonte(cel, tamanho=tam_fonte, negrito=negrito)
        celula_sup_esq.value = informacao

    except Exception as e:
        trata_excecao(e)
        raise

def lanca_dados_coluna(ws, celula, informacao, tam_fonte, negrito):
    try:
        celula_sup_esq = ws[celula]
        cel = ws[celula]
        edita_alinhamento(cel)
        edita_fonte(cel, tamanho=tam_fonte, negrito=negrito)
        celula_sup_esq.value = informacao

    except Exception as e:
        trata_excecao(e)
        raise

def linhas_colunas_p_edicao(sheet, min_linha, max_linha, min_coluna, max_coluna):
    try:
        for row in sheet.iter_rows(min_row=min_linha,
                                   max_row=max_linha,
                                   min_col=min_coluna,
                                   max_col=max_coluna):
            for cell in row:
                yield cell

    except Exception as e:
        trata_excecao(e)
        raise

def edita_alinhamento(cell, ali_horizontal='center', ali_vertical='center', rotacao=0, quebra_linha=False,
                      encolher=False, recuar=0):
    try:
        cell.alignment = Alignment(horizontal=ali_horizontal,
                                   vertical=ali_vertical,
                                   text_rotation=rotacao,
                                   wrap_text=quebra_linha,
                                   shrink_to_fit=encolher,
                                   indent=recuar)

    except Exception as e:
        trata_excecao(e)
        raise

def edita_bordas(cell):
    try:
        cell.border = Border(left=Side(border_style='thin', color='00000000'),
                             right=Side(border_style='thin', color='00000000'),
                             top=Side(border_style='thin', color='00000000'),
                             bottom=Side(border_style='thin', color='00000000'),
                             diagonal=Side(border_style='thick', color='00000000'),
                             diagonal_direction=0,
                             outline=Side(border_style='thin', color='00000000'),
                             vertical=Side(border_style='thin', color='00000000'),
                             horizontal=Side(border_style='thin', color='00000000'))

    except Exception as e:
        trata_excecao(e)
        raise

def edita_preenchimento(cell):
    try:
        cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

    except Exception as e:
        trata_excecao(e)
        raise

def edita_fonte(cell, nome_fonte='Calibri', tamanho=10, negrito=False):
    try:
        cell.font = Font(name=nome_fonte, size=tamanho, bold=negrito)

    except Exception as e:
        trata_excecao(e)
        raise

def carregar_workbook(caminho):
    try:
        book = load_workbook(caminho)

        return book

    except Exception as e:
        trata_excecao(e)
        raise

def letra_coluna(coluna):
    try:
        column_letter = get_column_letter(coluna)

        return column_letter

    except Exception as e:
        trata_excecao(e)
        raise

def ajusta_larg_coluna(planilha, coluna, largura_ajustada):
    try:
        planilha.column_dimensions[get_column_letter(coluna[0].column)].width = largura_ajustada

    except Exception as e:
        trata_excecao(e)
        raise

def criar_workbook():
    try:
        book = Workbook()

        return book

    except Exception as e:
        trata_excecao(e)
        raise

def adiciona_imagem(ws, caminho_arquivo, celula):
    try:
        img = drawing.image.Image(caminho_arquivo)
        ws.add_image(img, celula)

    except Exception as e:
        trata_excecao(e)
        raise

def dataframe_pandas(dados_lista, colunas):
    try:
        df = pd.DataFrame(dados_lista, columns=colunas)

        return df

    except Exception as e:
        trata_excecao(e)
        raise

def escritor_dataframe(caminho):
    try:
        writer = pd.ExcelWriter(caminho, engine='openpyxl')

        return writer

    except Exception as e:
        trata_excecao(e)
        raise

def escritor_direto_dataframe(df, writer, nome_planilha, partida, ini_coluna, cabecalho, indice):
    try:
        df.to_excel(writer,
                    sheet_name=nome_planilha,
                    startrow=partida,
                    startcol=ini_coluna,
                    header=cabecalho,
                    index=indice)

    except Exception as e:
        trata_excecao(e)
        raise
